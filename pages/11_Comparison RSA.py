from utils.auth import require_login

require_login()

# app_streamlit_merged_simple.py
# -*- coding: utf-8 -*-
"""
RSA - PGD Comparison — Simple (only Comparison + Export)
Updates:
- Compare SAP 'Ship-to-Sort1' with Infor 'Customer Number'
- Normalize Infor 'Customer Number' value '--' -> 'ZA30' before grouping
- Styled export visibility improvements
- Normalize 'Segment Attribute Desc' using mapping table
- Add combined 'Infor Customer PO item' column (Line Aggregator + SC Segmentation (new Infor column) + normalized Segment Attribute Desc)
- Add compare: SAP 'Segment Attribute' vs Infor normalized 'Segment Attribute Desc' -> Result_Segment Attribute Desc
- FIX: Delay comparison now normalizes both SAP (XX-XXXX) and Infor (XXXX) to canonical form before comparing
- ADD: 'Sales Channel' and 'Storage Location' columns from Infor data added to output
"""

import re
import io
import sys
from datetime import datetime, timedelta
from contextlib import nullcontext

import numpy as np
import pandas as pd
import streamlit as st

# timezone
try:
    from zoneinfo import ZoneInfo
    _tz = ZoneInfo("Asia/Jakarta")
except Exception:
    _tz = None

# Attempt import openpyxl styling and capture availability/errors
EXCEL_STYLED_AVAILABLE = False
_OPENPYXL_ERROR = None
try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_STYLED_AVAILABLE = True
except Exception as imp_e:
    EXCEL_STYLED_AVAILABLE = False
    _OPENPYXL_ERROR = imp_e

# -------------------
# Config / constants
# -------------------
JOIN_KEYS = ["PO No.(Full)", "CRD_key", "PD_key"]
DATE_COLS_INFOR = ["Issue Date","FPD","LPD","PSDD","PODD","CRD","PD"]
DATE_COLS_SAP   = ["Document Date","FPD","LPD","PSDD","PODD","FCR Date","PO Date","CRD","PD","Actual PGI"]

DESIRED_ORDER = [
    "Client No","Site","Brand FTY Name","SO","Order Type","Order Type Description",
    "PO No.(Full)","infor Order Status","Customer PO item","Infor Customer PO item","Result_Customer PO item",
    "Line Aggregator","Cust Ord No","infor Market PO Number","Result_Cust Ord No","PO No.(Short)",
    "Merchandise Category 2","Quanity","infor Quantity","Result_Quantity",
    "Model Name","Article No","infor Article No","Result Article No","SAP Material",
    "Pattern Code(Up.No.)","Model No","Outsole Mold","Gender","Category 1","Category 2","Category 3",
    "Unit Price","Classification Code","DRC",
    "Delay/Early - Confirmation PD","infor Delay/Early - Confirmation PD","Result Delay/Early - Confirmation PD",
    "Delay/Early - Confirmation CRD","infor Delay/Early - Confirmation CRD","Result Delay/Early - Confirmation CRD",
    "MDP","PDP","SDP","Article Lead time",
    "Ship-to-Sort1","infor Customer Number","Result_Ship-to-Sort1",
    "Ship-to Country","Ship to Name","Shipment Method","infor Shipment Method","Result Shipment Method",
    "Delay - PO PSDD Update","infor Delay - PO PSDD Update","Result Delay - PO PSDD Update",
    "Delay - PO PD Update","infor Delay - PO PD Update","Result Delay - PO PD Update",
    "Document Date","PODD","infor PODD","Result PODD",
    "LPD","infor LPD","Result LPD",
    "PSDD","infor PSDD","Result PSDD",
    "FPD","infor FPD","Result FPD",
    "CRD","infor CRD","Result CRD",
    "PD","infor PD","Result PD",
    "FCR Date","PO Date","Actual PGI",
    # Segment Attribute Desc comparison (both from Infor)
    "infor Segment Attribute","infor Segment Attribute Desc","Result_Segment Attribute Desc",
    "infor SC Segmentation",
    # Sales Channel and Storage Location from Infor
    "infor Sales Channel","infor Storage Location",
    "Segment","S&P LPD","Currency"
]

INFOR_COLOR  = "FFF9F16D"
RESULT_COLOR = "FFC6EFCE"
OTHER_COLOR  = "FFD9D9D9"
DATE_FMT_OPENPYXL = "m/d/yyyy"

BLANKS = {"(blank)", "blank", "", "--", " -- ", " --"}

# -------------------
# Segment Attribute Desc normalization mapping
# -------------------
SEGMENT_ATTR_DESC_MAP = {
    "Launch":                   "10",
    "Launch - Soft Launch":     "13",
    "Launch - Brand Priority":  "11",
    "Launch - Hard Launch":     "12",
    "Flex Track":               "20",
    "Flex Track - LockerRoom":  "21",
    "Responsiveness":           "30",
    "Responsiveness - PR":      "31",
    "Responsiveness - NOS":     "32",
    "Responsiveness - ISC":     "33",
    "Core":                     "40",
    "Core - Control":           "41",
}

def normalize_segment_attr_desc(series: pd.Series) -> pd.Series:
    """Map Segment Attribute Desc values to their numeric code using SEGMENT_ATTR_DESC_MAP.
    Strips whitespace and trailing colons before lookup. Returns original value if not found."""
    def _map(val):
        if pd.isna(val):
            return np.nan
        cleaned = str(val).strip().rstrip(":")
        return SEGMENT_ATTR_DESC_MAP.get(cleaned, cleaned)
    return series.apply(_map)

# -------------------
# Delay code mapping (with zero-padded variants)
# -------------------
# fmt: off
DELAY_CODE_MAPPING = {
    # numeric string (no padding) -> canonical
    '161':  '01-0161', '0161': '01-0161',
    '84':   '03-0084', '0084': '03-0084',
    '68':   '02-0068', '0068': '02-0068',
    '64':   '04-0064', '0064': '04-0064',
    '62':   '02-0062', '0062': '02-0062',
    '61':   '01-0061', '0061': '01-0061',
    '51':   '03-0051', '0051': '03-0051',
    '46':   '03-0046', '0046': '03-0046',
    '7':    '02-0007', '0007': '02-0007',
    '3':    '03-0003', '0003': '03-0003',
    '2':    '01-0002', '0002': '01-0002',
    '1':    '01-0001', '0001': '01-0001',
    '4':    '04-0004', '0004': '04-0004',
    '8':    '02-0008', '0008': '02-0008',
    '10':   '04-0010', '0010': '04-0010',
    '49':   '03-0049', '0049': '03-0049',
    '90':   '04-0090', '0090': '04-0090',
    '63':   '03-0063', '0063': '03-0063',
    '27':   '04-0027', '0027': '04-0027',
    # treat as "no value" -> compare as NaN
    '0':    None,
    '0000': None,
    '':     None,
}
# fmt: on

# Also register the canonical XX-XXXX strings as self-referencing keys
# so that SAP values already in that form can be looked up directly.
for _v in list(DELAY_CODE_MAPPING.values()):
    if _v and _v not in DELAY_CODE_MAPPING:
        DELAY_CODE_MAPPING[_v] = _v


def _extract_delay_lookup_key(val) -> str | None:
    """
    Extract the lookup key for DELAY_CODE_MAPPING from any delay value format:
      - already canonical  "03-0063"  → "03-0063"  (direct match in mapping)
      - zero-padded suffix "0063"     → "0063"
      - bare number        "63"       → "63"
      - blank / NaN / "0000"         → None
    """
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return None
    s = str(val).strip()
    if not s or s.lower() in {b.lower() for b in BLANKS}:
        return None
    # canonical format XX-XXXX: try direct lookup first
    if re.match(r'^\d{2}-\d{4}$', s):
        return s
    # any digit-only string (with or without leading zeros)
    digits_only = re.sub(r'\D', '', s)
    return digits_only if digits_only else None


def normalize_delay_to_canonical(series: pd.Series) -> pd.Series:
    """
    Map any delay value to its canonical 'XX-XXXX' form using DELAY_CODE_MAPPING.
    Returns np.nan for blanks, zeros, or unmapped values → comparison treats as missing.
    """
    def _norm(val):
        key = _extract_delay_lookup_key(val)
        if key is None:
            return np.nan
        result = DELAY_CODE_MAPPING.get(key)
        # result == None means "treat as no-value"
        if result is None:
            return np.nan
        # not in mapping at all → keep original key so mismatch is visible
        return result if result is not None else key
    return series.apply(_norm)


def map_delay_series_for_display(s: pd.Series) -> pd.Series:
    """
    Used to overwrite the infor delay columns with their canonical form
    so the Excel output shows 'XX-XXXX' rather than raw digits.
    (SAP columns are already in XX-XXXX form.)
    """
    return normalize_delay_to_canonical(s).fillna("")


# -------------------
# Helpers (kept)
# -------------------
def to_dt_series(s: pd.Series) -> pd.Series:
    return pd.to_datetime(s, errors="coerce").dt.normalize()

def to_dt_scalar(x):
    ts = pd.to_datetime(x, errors="coerce")
    if pd.isna(ts): return pd.NaT
    return pd.to_datetime(ts).normalize()

def fmt_dt(dt: pd.Timestamp) -> str:
    if pd.isna(dt): return ""
    return dt.strftime("%m/%d/%Y")

def date_concat(series: pd.Series) -> str:
    dts = to_dt_series(series).dropna()
    if dts.empty:
        return np.nan
    uniq = sorted(set(dts))
    return " | ".join(fmt_dt(x) for x in uniq)

def date_to_text_cell(val) -> str:
    return fmt_dt(to_dt_scalar(val))

def sum_keep_nan(s: pd.Series):
    return s.sum(min_count=1)

def keep_or_join(s: pd.Series):
    vals = pd.unique(s.dropna())
    if len(vals) == 0: return np.nan
    if len(vals) == 1: return vals[0]
    return " | ".join(map(str, vals))

def extract_digits(x):
    if pd.isna(x): return np.nan
    s = str(x).strip()
    if s.lower() in {b.lower() for b in BLANKS}: return np.nan
    m = re.search(r"\d+", s)
    return m.group(0) if m else np.nan

def split_date_set(x):
    if pd.isna(x):
        return set()
    parts = [p.strip() for p in str(x).split("|")]
    out = set()
    for p in parts:
        if not p:
            continue
        dt = pd.to_datetime(p, errors="coerce")
        if not pd.isna(dt):
            out.add(pd.to_datetime(dt).normalize())
    return out

def to_excel_col(n):
    s = ""; n += 1
    while n:
        n, r = divmod(n-1, 26)
        s = chr(r + 65) + s
    return s

# -------------------
# File readers (cached)
# -------------------
@st.cache_data(show_spinner=False)
def read_excel_file_bytes(uploaded):
    if hasattr(uploaded, "read"):
        uploaded.seek(0)
        return pd.read_excel(uploaded, engine="openpyxl")
    else:
        return pd.read_excel(str(uploaded), engine="openpyxl")

@st.cache_data(show_spinner=False)
def read_csv_file_bytes(uploaded):
    if hasattr(uploaded, "read"):
        for enc in ("utf-8", "utf-8-sig", "latin1"):
            try:
                uploaded.seek(0)
                return pd.read_csv(uploaded, encoding=enc)
            except Exception:
                continue
        uploaded.seek(0)
        return pd.read_csv(uploaded)
    else:
        return pd.read_csv(str(uploaded))

# -------------------
# match_qty_nearest (fallback)
# -------------------
def match_qty_nearest(df_sap, df_infor, key="PO No.(Full)", qty_col="Quanity", infor_qty_col="infor Quantity"):
    df_sap = df_sap.copy()
    df_sap["___sap_row_id"] = np.arange(len(df_sap))
    merged = df_sap.merge(df_infor, how="left", left_on=key, right_on="PO No.(Full)" if "PO No.(Full)" in df_infor.columns else key)
    if merged.empty:
        return merged
    out_rows = []
    for po, group in merged.groupby(key, sort=False):
        n_sap_rows = group["___sap_row_id"].nunique() if "___sap_row_id" in group.columns else 1
        if n_sap_rows > 1:
            out_rows.extend(group.drop(columns=["___sap_row_id"], errors="ignore").to_dict("records"))
            continue
        if len(group) == 1:
            out_rows.append(group.drop(columns=["___sap_row_id"], errors="ignore").iloc[0].to_dict())
            continue
        try:
            sap_qty = pd.to_numeric(group.iloc[0].get(qty_col, 0), errors="coerce")
        except Exception:
            sap_qty = pd.to_numeric(group.iloc[0].get("Quantity", 0), errors="coerce")
        diffs = (pd.to_numeric(group.get(infor_qty_col, 0), errors="coerce") - sap_qty).abs()
        idx_min = diffs.idxmin() if diffs.notna().any() else group.index[0]
        for idx, row in group.iterrows():
            row = row.copy()
            if idx != idx_min:
                if qty_col in row.index:
                    row[qty_col] = np.nan
                if "Quantity" in row.index:
                    row["Quantity"] = np.nan
            out_rows.append(row.drop(labels=["___sap_row_id"], errors="ignore").to_dict())
    return pd.DataFrame(out_rows)

# -------------------
# Core pipeline
# -------------------
def run_core_pipeline(df_sap_raw, df_infor_raw_all, *,
                     prefer_strict_join=True,
                     fallback_match_qty=True,
                     export_filename_prefix=None):
    df_sap = df_sap_raw.copy()
    df_infor = df_infor_raw_all.copy()

    _today = datetime.now(_tz) if _tz else datetime.now()
    _today_str = _today.strftime("%Y%m%d")
    OUT_JOINED = export_filename_prefix or f"RSA - PGD Comparison Tracking Report - {_today_str}.xlsx"

    # normalize infor columns
    rename_cols = {
        'Order #': 'PO No.(Full)',
        'Line Aggregator': 'Customer PO item',
        'Article Number': 'Article No',
        'Country/Region': 'Ship-to Country',
        'Customer Request Date (CRD)': 'CRD',
        'Plan Date': 'PD',
        'PO Statistical Delivery Date (PSDD)': 'PSDD',
        'First Production Date': 'FPD',
        'Last Production Date': 'LPD',
        'Production Lead Time': 'Lead time',
        'Class Code': 'Classification Code',
        'Delay - Confirmation': 'Delay/Early - Confirmation CRD',
        'Delay - PO Del Update': 'Delay - PO PSDD Update',
        'Grand Total': 'Quantity',
        'Delivery Delay Pd': 'Delay - PO PD Update',
        'Shipment Method': 'Shipment Method',
    }
    df_infor = df_infor.rename(columns=rename_cols)

    # --- Normalization: replace '--' in Customer Number with 'ZA30' ---
    if "Customer Number" in df_infor.columns:
        df_infor["Customer Number"] = df_infor["Customer Number"].replace("--", "ZA30")
    if "infor Customer Number" in df_infor.columns:
        df_infor["infor Customer Number"] = df_infor["infor Customer Number"].replace("--", "ZA30")

    # --- Normalization: Segment Attribute Desc -> numeric code ---
    if "Segment Attribute Desc" in df_infor.columns:
        df_infor["Segment Attribute Desc"] = normalize_segment_attr_desc(df_infor["Segment Attribute Desc"])

    if "Confirmation Delay Pd" in df_infor.columns and "Delay/Early - Confirmation PD" not in df_infor.columns:
        df_infor = df_infor.rename(columns={"Confirmation Delay Pd": "Delay/Early - Confirmation PD"})

    if "Quantity" in df_infor.columns:
        df_infor = df_infor[df_infor["Quantity"].fillna(0) != 0].copy()

    # --- Build 'Infor Customer PO item' per row BEFORE groupby ---
    def _build_infor_po_item_pre(row):
        parts = []
        for col in ["Customer PO item", "SC Segmentation", "Segment Attribute Desc"]:
            val = row.get(col, np.nan)
            if pd.isna(val):
                parts.append("")
            else:
                s = str(val).strip()
                if not s or s in BLANKS:
                    parts.append("")
                else:
                    try:
                        f = float(s)
                        if f == int(f):
                            s = str(int(f))
                    except (ValueError, TypeError):
                        pass
                    parts.append(s)
        return "".join(parts)

    _po_item_src_cols = [c for c in ["Customer PO item", "SC Segmentation", "Segment Attribute Desc"] if c in df_infor.columns]
    if _po_item_src_cols:
        df_infor["Infor Customer PO item"] = df_infor.apply(_build_infor_po_item_pre, axis=1)

    # detect size columns
    size_pat  = re.compile(r'^(?:[1-9]|1[0-8])(?:K|-K|-)?$')
    size_cols = [c for c in df_infor.columns if size_pat.match(str(c))]
    sum_cols   = size_cols + (["Quantity"] if "Quantity" in df_infor.columns else [])
    other_cols = [c for c in df_infor.columns if c not in (["Issue Date","PO No.(Full)","Model Name","Article No","Ship-to Country","CRD","PD"] + sum_cols)]

    agg_infor = {}
    for col in sum_cols:
        agg_infor[col] = sum_keep_nan
    for col in other_cols:
        if col in DATE_COLS_INFOR:
            agg_infor[col] = date_concat
        else:
            agg_infor[col] = keep_or_join

    meta_cols = ["Issue Date", "PO No.(Full)", "Model Name", "Article No", "Ship-to Country", "CRD", "PD"]
    missing = [c for c in meta_cols if c not in df_infor.columns]
    if missing:
        if "PO No.(Full)" in df_infor.columns:
            grp_cols = ["PO No.(Full)"]
            df_infor_grouped = df_infor.groupby(grp_cols, dropna=False).agg({'Quantity':'sum'}).reset_index()
            for c in ["Model Name","Article No","Country/Region","Customer Request Date (CRD)","Plan Date"]:
                if c in df_infor.columns:
                    df_infor_grouped[c] = df_infor.groupby("PO No.(Full)")[c].agg(keep_or_join).values
            if "Customer Request Date (CRD)" in df_infor_grouped.columns:
                df_infor_grouped = df_infor_grouped.rename(columns={"Customer Request Date (CRD)": "CRD"})
            if "Plan Date" in df_infor_grouped.columns:
                df_infor_grouped = df_infor_grouped.rename(columns={"Plan Date": "PD"})
        else:
            raise ValueError(f"Kolom meta tidak ditemukan di Infor: {missing}")
    else:
        df_inf_num = df_infor.copy()
        if sum_cols:
            df_inf_num[sum_cols] = df_inf_num[sum_cols].apply(pd.to_numeric, errors="coerce")
        df_infor_grouped = df_inf_num.groupby(meta_cols, dropna=False).agg(agg_infor).reset_index()
        if "Customer PO item" in df_infor_grouped.columns and "Line Aggregator" not in df_infor_grouped.columns:
            df_infor_grouped["Line Aggregator"] = df_infor_grouped["Customer PO item"]

    # Ensure Customer Number (if present) is carried over to grouped table
    if "Customer Number" in df_infor.columns and "Customer Number" not in df_infor_grouped.columns:
        if "PO No.(Full)" in df_infor.columns:
            df_infor_grouped = df_infor_grouped.merge(
                df_infor.groupby("PO No.(Full)", dropna=False)["Customer Number"].agg(keep_or_join).reset_index(),
                on="PO No.(Full)", how="left"
            )

    # Ensure SC Segmentation, Segment Attribute and Segment Attribute Desc carried over to grouped table
    for seg_col in ["SC Segmentation", "Segment Attribute", "Segment Attribute Desc"]:
        if seg_col in df_infor.columns and seg_col not in df_infor_grouped.columns:
            if "PO No.(Full)" in df_infor.columns:
                df_infor_grouped = df_infor_grouped.merge(
                    df_infor.groupby("PO No.(Full)", dropna=False)[seg_col].agg(keep_or_join).reset_index(),
                    on="PO No.(Full)", how="left"
                )

    # Ensure Sales Channel and Storage Location carried over to grouped table
    for extra_col in ["Sales Channel", "Storage Location"]:
        if extra_col in df_infor.columns and extra_col not in df_infor_grouped.columns:
            if "PO No.(Full)" in df_infor.columns:
                df_infor_grouped = df_infor_grouped.merge(
                    df_infor.groupby("PO No.(Full)", dropna=False)[extra_col].agg(keep_or_join).reset_index(),
                    on="PO No.(Full)", how="left"
                )

    # make keys
    if "CRD" in df_infor_grouped.columns:
        df_infor_grouped["CRD_key"] = to_dt_series(df_infor_grouped["CRD"])
    else:
        df_infor_grouped["CRD_key"] = pd.NaT
    if "PD" in df_infor_grouped.columns:
        df_infor_grouped["PD_key"] = to_dt_series(df_infor_grouped["PD"])
    else:
        df_infor_grouped["PD_key"] = pd.NaT

    # SAP: render display dates as text
    df_sap = df_sap.copy()
    for col in DATE_COLS_SAP:
        if col in df_sap.columns:
            df_sap[col] = df_sap[col].map(date_to_text_cell)
    df_sap["CRD_key"] = to_dt_series(df_sap["CRD"]) if "CRD" in df_sap.columns else pd.NaT
    df_sap["PD_key"]  = to_dt_series(df_sap["PD"]) if "PD" in df_sap.columns else pd.NaT

    # prepare infor pick — includes Sales Channel and Storage Location
    infor_cols_for_merge = [
        "Order Status","Article No","LPD","PODD","PSDD","FPD","CRD","PD",
        "Delay/Early - Confirmation CRD","Delay - PO PSDD Update","Delay - PO PD Update",
        "Quantity","Shipment Method","Issue Date",
        "Customer PO item","Line Aggregator","Infor Customer PO item","Market PO Number",
        "Sales Channel","Storage Location"
    ]
    if "Customer Number" in df_infor_grouped.columns:
        infor_cols_for_merge.append("Customer Number")

    for seg_col in ["SC Segmentation", "Segment Attribute", "Segment Attribute Desc"]:
        if seg_col in df_infor_grouped.columns:
            infor_cols_for_merge.append(seg_col)

    if "Delay/Early - Confirmation PD" in df_infor_grouped.columns:
        infor_cols_for_merge = ["Delay/Early - Confirmation PD"] + infor_cols_for_merge

    inf_pick_cols = [c for c in infor_cols_for_merge if c in df_infor_grouped.columns]
    inf_pick = df_infor_grouped[["PO No.(Full)","CRD_key","PD_key"] + inf_pick_cols].copy()
    pref_map = {c: f"infor {c}" for c in inf_pick_cols}
    pref_map.pop("Infor Customer PO item", None)
    inf_pick = inf_pick.rename(columns=pref_map)

    # Try strict join
    df_join = None
    if prefer_strict_join:
        try:
            df_join = df_sap.merge(inf_pick, on=JOIN_KEYS, how="left")
        except Exception:
            df_join = df_sap.merge(inf_pick, on=["PO No.(Full)"], how="left")

    if df_join is None or df_join.shape[0] == 0 or ("infor Quantity" in df_join.columns and df_join["infor Quantity"].isna().all()):
        if fallback_match_qty:
            df_infor_for_match = df_infor_grouped.copy()
            if "PO No.(Full)" not in df_infor_for_match.columns and "Order #" in df_infor.columns:
                df_infor_for_match = df_infor_for_match.rename(columns={"Order #": "PO No.(Full)"})
            if "Quantity" in df_infor_for_match.columns and "infor Quantity" not in df_infor_for_match.columns:
                df_infor_for_match = df_infor_for_match.rename(columns={"Quantity": "infor Quantity"})
            if "Quanity" not in df_sap.columns and "Quantity" in df_sap.columns:
                df_sap = df_sap.rename(columns={"Quantity": "Quanity"})
            df_join = match_qty_nearest(df_sap, df_infor_for_match, key="PO No.(Full)", qty_col="Quanity", infor_qty_col="infor Quantity")
        else:
            df_join = df_sap.merge(inf_pick, on=["PO No.(Full)"], how="left")

    if df_join is None:
        raise RuntimeError("Merge failed — no join result produced.")

    # ----------------------------------------------------------------
    # Map infor delay columns to canonical XX-XXXX form (for display)
    # NOTE: SAP columns already contain XX-XXXX values.
    # ----------------------------------------------------------------
    for col in [
        "infor Delay/Early - Confirmation PD",
        "infor Delay/Early - Confirmation CRD",
        "infor Delay - PO PSDD Update",
        "infor Delay - PO PD Update",
    ]:
        if col in df_join.columns:
            df_join[col] = normalize_delay_to_canonical(df_join[col]).fillna("")

    # ----------------------------------------------------------------
    # compare logic
    # ----------------------------------------------------------------
    def norm_num(s): return pd.to_numeric(s, errors="coerce")
    def norm_str(s):
        s = s.astype(str).str.strip()
        s = s.replace(list(BLANKS), np.nan)
        return s
    def equal_series(a, b): return a.eq(b) | (a.isna() & b.isna())
    def compare_dates_as_sets(s_left, s_right):
        left_sets  = s_left.apply(split_date_set)
        right_sets = s_right.apply(split_date_set)
        return (left_sets == right_sets).fillna(False)

    df = df_join.copy()

    if "Quanity" in df.columns and "infor Quantity" in df.columns:
        df["Result_Quantity"] = equal_series(norm_num(df["Quanity"]), norm_num(df["infor Quantity"])).fillna(False)

    if "Article No" in df.columns and "infor Article No" in df.columns:
        df["Result Article No"] = equal_series(norm_str(df["Article No"]), norm_str(df["infor Article No"])).fillna(False)

    # ----------------------------------------------------------------
    # Delay comparisons — both sides normalized to canonical XX-XXXX
    # using normalize_delay_to_canonical so format differences (03-0063
    # vs 0063 vs 63) all resolve to the same canonical string.
    # ----------------------------------------------------------------
    delay_pairs = [
        ("Delay/Early - Confirmation PD",  "infor Delay/Early - Confirmation PD",  "Result Delay/Early - Confirmation PD"),
        ("Delay/Early - Confirmation CRD", "infor Delay/Early - Confirmation CRD", "Result Delay/Early - Confirmation CRD"),
        ("Delay - PO PSDD Update",         "infor Delay - PO PSDD Update",         "Result Delay - PO PSDD Update"),
        ("Delay - PO PD Update",           "infor Delay - PO PD Update",           "Result Delay - PO PD Update"),
    ]
    for sap_col, infor_col, result_col in delay_pairs:
        if sap_col in df.columns and infor_col in df.columns:
            sap_norm   = normalize_delay_to_canonical(df[sap_col])
            infor_norm = normalize_delay_to_canonical(df[infor_col])
            df[result_col] = equal_series(sap_norm, infor_norm).fillna(False)

    date_pairs = [
        ("FPD","infor FPD","Result FPD"),
        ("LPD","infor LPD","Result LPD"),
        ("CRD","infor CRD","Result CRD"),
        ("PSDD","infor PSDD","Result PSDD"),
        ("PODD","infor PODD","Result PODD"),
        ("PD","infor PD","Result PD"),
    ]
    for left, right, outcol in date_pairs:
        if left in df.columns and right in df.columns:
            df[outcol] = compare_dates_as_sets(df[left], df[right])

    if "Shipment Method" in df.columns and "infor Shipment Method" in df.columns:
        df["Result Shipment Method"] = equal_series(norm_str(df["Shipment Method"]), norm_str(df["infor Shipment Method"])).fillna(False)

    # --- Compare Ship-to-Sort1 (SAP) with infor Customer Number ---
    if "Ship-to-Sort1" in df.columns and "infor Customer Number" in df.columns:
        df["Result_Ship-to-Sort1"] = equal_series(norm_str(df["Ship-to-Sort1"]), norm_str(df["infor Customer Number"])).fillna(False)

    # --- Compare SAP 'Cust Ord No' vs Infor 'Market PO Number' ---
    if "Cust Ord No" in df.columns and "infor Market PO Number" in df.columns:
        df["Result_Cust Ord No"] = equal_series(norm_str(df["Cust Ord No"]), norm_str(df["infor Market PO Number"])).fillna(False)

    if "Line Aggregator" not in df.columns and "infor Line Aggregator" in df.columns:
        df["Line Aggregator"] = df["infor Line Aggregator"]

    # Drop prefixed infor Customer PO item to avoid duplicate with the pre-built combined column
    if "infor Customer PO item" in df.columns:
        df = df.drop(columns=["infor Customer PO item"])

    # --- Compare Customer PO item (SAP) vs Infor Customer PO item (combined) ---
    if "Customer PO item" in df.columns and "Infor Customer PO item" in df.columns:
        df["Result_Customer PO item"] = df.apply(
            lambda row: (
                True if pd.isna(row.get("Customer PO item")) and pd.isna(row.get("Infor Customer PO item"))
                else False if pd.isna(row.get("Customer PO item")) or pd.isna(row.get("Infor Customer PO item"))
                else str(row["Customer PO item"]).strip() == str(row["Infor Customer PO item"]).strip()
            ), axis=1
        )

    # --- Compare Infor 'Segment Attribute' vs Infor normalized 'Segment Attribute Desc' ---
    if "infor Segment Attribute" in df.columns and "infor Segment Attribute Desc" in df.columns:
        def _norm_single(val):
            if pd.isna(val):
                return np.nan
            s = str(val).strip()
            return np.nan if s in BLANKS else s

        df["Result_Segment Attribute Desc"] = df.apply(
            lambda row: (
                True  if pd.isna(_norm_single(row.get("infor Segment Attribute"))) and pd.isna(_norm_single(row.get("infor Segment Attribute Desc")))
                else False if pd.isna(_norm_single(row.get("infor Segment Attribute"))) or pd.isna(_norm_single(row.get("infor Segment Attribute Desc")))
                else str(row["infor Segment Attribute"]).strip() == str(row["infor Segment Attribute Desc"]).strip()
            ), axis=1
        )

    # --- Add empty placeholder columns for Shipment Method and Result Shipment Method ---
    if "Shipment Method" not in df.columns:
        df["Shipment Method"] = np.nan
    if "Result Shipment Method" not in df.columns:
        df["Result Shipment Method"] = np.nan

    present = [c for c in DESIRED_ORDER if c in df.columns]
    rest     = [c for c in df.columns if c not in present]
    df_final = df[present + rest]

    # Primary export (xlsxwriter) with per-cell date handling
    DATE_FMT = "mm/dd/yyyy"
    date_display_cols = [
        "Document Date","FPD","LPD","PSDD","PODD","FCR Date","PD","PO Date","Actual PGI",
        "infor FPD","infor LPD","infor CRD","infor PSDD","infor PODD","infor PD",
        "CRD"
    ]
    date_display_cols = [c for c in date_display_cols if c in df_final.columns]
    def is_single_date_text(s: str) -> bool:
        if not isinstance(s, str): return False
        if " | " in s: return False
        dt = pd.to_datetime(s, errors="coerce")
        return not pd.isna(dt)

    out = io.BytesIO()
    try:
        import xlsxwriter
        with pd.ExcelWriter(out, engine="xlsxwriter") as writer:
            df_final.to_excel(writer, index=False, sheet_name="Data", startrow=0, startcol=0, header=True)
            wb  = writer.book
            ws  = writer.sheets["Data"]

            fmt_bool_T = wb.add_format({"bg_color": "#C6EFCE", "font_color": "#006100"})
            fmt_bool_F = wb.add_format({"bg_color": "#FFC7CE", "font_color": "#9C0006"})
            fmt_date   = wb.add_format({"num_format": DATE_FMT})

            # Header color formats
            fmt_hdr_infor  = wb.add_format({"bg_color": "#F9F16D", "bold": True, "font_name": "Calibri", "font_size": 9, "align": "center", "valign": "vcenter", "border": 1})
            fmt_hdr_result = wb.add_format({"bg_color": "#C6EFCE", "bold": True, "font_name": "Calibri", "font_size": 9, "align": "center", "valign": "vcenter", "border": 1})
            fmt_hdr_other  = wb.add_format({"bg_color": "#D9D9D9", "bold": True, "font_name": "Calibri", "font_size": 9, "align": "center", "valign": "vcenter", "border": 1})

            # Write header row with colors
            for cidx, col in enumerate(df_final.columns):
                col_lower = col.lower()
                if col_lower.startswith("infor ") or col == "Infor Customer PO item":
                    fmt = fmt_hdr_infor
                elif col.startswith("Result ") or col.startswith("Result_"):
                    fmt = fmt_hdr_result
                else:
                    fmt = fmt_hdr_other
                ws.write(0, cidx, col, fmt)

            nrows, ncols = df_final.shape
            for cidx, col in enumerate(df_final.columns):
                if col in date_display_cols:
                    for ridx, val in enumerate(df_final[col].tolist(), start=1):
                        if pd.isna(val):
                            ws.write_string(ridx, cidx, "")
                        else:
                            s = str(val).strip()
                            if is_single_date_text(s):
                                dt = pd.to_datetime(s, errors="coerce")
                                if not pd.isna(dt):
                                    ws.write_datetime(ridx, cidx, dt.to_pydatetime(), fmt_date)
                                else:
                                    ws.write_string(ridx, cidx, s)
                            else:
                                ws.write_string(ridx, cidx, s)

            res_cols = [c for c in df_final.columns if c.startswith("Result " ) or c.startswith("Result_")]
            for col in res_cols:
                cidx = df_final.columns.get_loc(col)
                rng  = f"{to_excel_col(cidx)}2:{to_excel_col(cidx)}{nrows+1}"
                ws.conditional_format(rng, {"type":"cell","criteria":"==","value":"TRUE","format":fmt_bool_T})
                ws.conditional_format(rng, {"type":"cell","criteria":"==","value":"FALSE","format":fmt_bool_F})

            for idx, col in enumerate(df_final.columns, start=1):
                if col.startswith("Result " ) or col.startswith("Result_"):
                    ws.set_column(idx-1, idx-1, 16)
                elif col in date_display_cols:
                    ws.set_column(idx-1, idx-1, 12)
                else:
                    ws.set_column(idx-1, idx-1, 18)

            ws.freeze_panes(1, 0)
            ws.autofilter(0, 0, nrows, ncols-1)
        out.seek(0)
    except Exception:
        out = io.BytesIO()
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            df_final.to_excel(writer, index=False, sheet_name="Data")
        out.seek(0)

    # Styled export using openpyxl (only if import succeeded)
    styled_bytes = None
    if EXCEL_STYLED_AVAILABLE:
        try:
            bio = io.BytesIO()
            with pd.ExcelWriter(bio, engine="openpyxl") as writer:
                df_final.to_excel(writer, index=False, sheet_name="Report")
                ws = writer.sheets["Report"]

                # header styling
                header_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))[0]
                idx_by_name = {c.value: i+1 for i, c in enumerate(header_cells)}
                for cell in header_cells:
                    col_name = str(cell.value) if cell.value is not None else ""
                    if col_name.lower().startswith("infor "):
                        cell.fill = PatternFill("solid", fgColor=INFOR_COLOR)
                    elif col_name.startswith("Result ") or col_name.startswith("Result_"):
                        cell.fill = PatternFill("solid", fgColor=RESULT_COLOR)
                    elif col_name == "Infor Customer PO item":
                        cell.fill = PatternFill("solid", fgColor=INFOR_COLOR)
                    else:
                        cell.fill = PatternFill("solid", fgColor=OTHER_COLOR)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = Font(name="Calibri", size=9, bold=True)

                # body formatting
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.font = Font(name="Calibri", size=9)

                # date columns
                for date_col in date_display_cols:
                    if date_col in idx_by_name:
                        cidx = idx_by_name[date_col]
                        for r in range(2, ws.max_row + 1):
                            cell = ws.cell(row=r, column=cidx)
                            if cell.value not in ("", None):
                                try:
                                    if isinstance(cell.value, str):
                                        dt = pd.to_datetime(cell.value, errors="coerce")
                                        if not pd.isna(dt):
                                            cell.value = dt.to_pydatetime()
                                    cell.number_format = DATE_FMT_OPENPYXL
                                except Exception:
                                    pass

                # auto width
                for col_idx in range(1, ws.max_column + 1):
                    col_letter = get_column_letter(col_idx)
                    maxlen = 0
                    for cell in ws[col_letter]:
                        v = "" if cell.value is None else str(cell.value)
                        maxlen = max(maxlen, len(v))
                    ws.column_dimensions[col_letter].width = min(max(9, maxlen + 2), 60)

                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions

            bio.seek(0)
            styled_bytes = bio
        except Exception as e:
            styled_bytes = None
            print("Styled export error:", e, file=sys.stderr)

    return df_final, out, styled_bytes, OUT_JOINED

# -------------------
# Streamlit UI
# -------------------
st.set_page_config(page_title="RSA - PGD Comparison (Simple)", layout="wide")
st.title("RSA - PGD Comparison — Simple (only Comparison + Export)")

with st.sidebar:
    st.header("Upload files")
    sap_file = st.file_uploader("Upload SAP Excel (.xlsx/.xls/.csv)", type=["xlsx","xls","csv"], key="sap_file")
    infor_files = st.file_uploader("Upload Infor files (CSV/XLSX) - multiple allowed", type=["csv","xlsx","xls"], accept_multiple_files=True, key="infor_files")
    st.markdown("---")
    st.header("Options")
    prefer_strict = st.checkbox("Prefer strict join (PO + CRD_key + PD_key)", value=True)
    fallback_match = st.checkbox("Enable fallback tolerant matching (match_qty_nearest)", value=True)
    st.markdown("---")
    if EXCEL_STYLED_AVAILABLE:
        st.success("openpyxl available — styled export enabled")
    else:
        st.warning("openpyxl NOT available — styled export disabled")
        if _OPENPYXL_ERROR is not None:
            st.caption(f"Import error: {_OPENPYXL_ERROR}")

st.markdown("### Upload & Run Pipeline")
if sap_file and infor_files:
    try:
        if sap_file.name.lower().endswith((".xls",".xlsx")):
            sap_df_load = read_excel_file_bytes(sap_file)
        else:
            sap_df_load = read_csv_file_bytes(sap_file)
        st.write("SAP preview:")
        st.dataframe(sap_df_load.head(20))
    except Exception as e:
        st.error("Gagal membaca SAP file.")
        st.exception(e)

    try:
        infor_dfs = []
        for f in infor_files:
            if f.name.lower().endswith((".xls",".xlsx")):
                infor_dfs.append(read_excel_file_bytes(f))
            else:
                infor_dfs.append(read_csv_file_bytes(f))
        df_infor_all = pd.concat(infor_dfs, ignore_index=True)
        st.write("Infor combined preview:")
        st.dataframe(df_infor_all.head(20))
    except Exception as e:
        st.error("Gagal membaca Infor files.")
        st.exception(e)

    if st.button("Run pipeline"):
        try:
            with st.spinner("Running pipeline..."):
                df_final, bytes_xlsx, styled_bytes, out_name = run_core_pipeline(
                    sap_df_load, df_infor_all,
                    prefer_strict_join=prefer_strict,
                    fallback_match_qty=fallback_match
                )
            st.success("Pipeline selesai.")
            st.subheader("Preview result (top 200 rows)")
            st.dataframe(df_final.head(200), use_container_width=True)

            st.download_button("⬇️ Download Excel (xlsxwriter, per-cell dates)", data=bytes_xlsx.getvalue(),
                               file_name=out_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            if EXCEL_STYLED_AVAILABLE and styled_bytes is not None:
                st.download_button("⬇️ Download Excel (styled, openpyxl)", data=styled_bytes.getvalue(),
                                   file_name=out_name.replace(".xlsx","_styled.xlsx"),
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                st.success("Styled export created successfully (openpyxl). Open in Microsoft Excel (desktop) to view colors.")
            else:
                if EXCEL_STYLED_AVAILABLE and styled_bytes is None:
                    st.error("Styled export attempted but failed during creation. Check server logs.")
                else:
                    st.info("Styled export unavailable (openpyxl not installed).")

            csv_name = out_name.replace(".xlsx", ".csv")
            st.download_button("⬇️ Download CSV (basic)", data=df_final.to_csv(index=False).encode("utf-8"), file_name=csv_name, mime="text/csv")

            st.markdown("### Summary Result*")
            res_bool_cols = [c for c in df_final.columns if c.startswith("Result " ) or c.startswith("Result_")]
            if res_bool_cols:
                summary = {}
                for c in res_bool_cols:
                    col = df_final[c]
                    true_count  = int(col.eq(True).sum())
                    false_count = int(col.eq(False).sum())
                    summary[c] = {"TRUE": true_count, "FALSE": false_count}
                st.json(summary)
            else:
                st.info("Tidak ditemukan kolom Result* di hasil.")
        except Exception as e:
            st.error("Pipeline gagal.")
            st.exception(e)
else:
    st.info("Unggah SAP file dan minimal 1 Infor file di sidebar untuk mulai.")

with st.expander("🛠 Debug info"):
    try:
        import platform
        st.write("Python:", sys.version)
        st.write("Platform:", platform.platform())
        st.write("Streamlit:", st.__version__)
        st.write("Pandas:", pd.__version__)
        import numpy as np2
        st.write("NumPy:", np2.__version__)
        st.write("openpyxl available:", EXCEL_STYLED_AVAILABLE)
        if _OPENPYXL_ERROR is not None:
            st.write("openpyxl import error:", _OPENPYXL_ERROR)
    except Exception as e:
        st.write("Failed to fetch debug info:", e)
