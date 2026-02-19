import streamlit as st
import pandas as pd
from io import BytesIO

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Pivot Report Generator",
    page_icon="📊",
    layout="wide",
)

# ── Custom CSS ─────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;600&display=swap');

html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif; }
.stApp { background: #0f1117; color: #e8e8e8; }
h1, h2, h3 { font-family: 'IBM Plex Mono', monospace !important; color: #00d4aa !important; }
.block-container { padding-top: 2rem; max-width: 1600px; }

section[data-testid="stFileUploadDropzone"] {
    border: 2px dashed #00d4aa !important;
    background: #161b27 !important;
    border-radius: 8px;
}
.stButton > button {
    background: #00d4aa !important; color: #0f1117 !important;
    font-family: 'IBM Plex Mono', monospace !important; font-weight: 600;
    border: none; border-radius: 4px; padding: 0.5rem 1.5rem; transition: all 0.2s;
}
.stButton > button:hover { background: #00f0c3 !important; transform: translateY(-1px); }
.stMultiSelect [data-baseweb="tag"] {
    background-color: #00d4aa22 !important; color: #00d4aa !important;
    border: 1px solid #00d4aa55 !important;
}
.metric-card {
    background: #161b27; border: 1px solid #00d4aa33;
    border-radius: 8px; padding: 1rem 1.5rem; text-align: center;
}
.metric-value { font-family: 'IBM Plex Mono', monospace; font-size: 1.8rem; font-weight: 600; color: #00d4aa; }
.metric-label { font-size: 0.75rem; color: #888; text-transform: uppercase; letter-spacing: 1px; margin-top: 0.25rem; }
.header-bar {
    background: linear-gradient(135deg, #00d4aa22, #0077ff11);
    border-left: 4px solid #00d4aa; border-radius: 0 8px 8px 0;
    padding: 0.75rem 1.25rem; margin-bottom: 1.5rem;
}
.stDownloadButton > button {
    background: #161b27 !important; color: #00d4aa !important;
    border: 1px solid #00d4aa !important; font-family: 'IBM Plex Mono', monospace !important;
    font-weight: 600; border-radius: 4px;
}
.stDownloadButton > button:hover { background: #00d4aa22 !important; }
div[data-testid="stSelectbox"] label, div[data-testid="stMultiSelect"] label {
    color: #aaa !important; font-size: 0.78rem; text-transform: uppercase;
    letter-spacing: 0.8px; font-family: 'IBM Plex Mono', monospace !important;
}
hr { border-color: #00d4aa22 !important; }
.col-list {
    background: #161b27; border: 1px solid #333; border-radius: 6px;
    padding: 0.75rem 1rem; font-family: 'IBM Plex Mono', monospace;
    font-size: 0.76rem; color: #aaa; line-height: 1.8;
}
</style>
""", unsafe_allow_html=True)

# ── Default row fields ─────────────────────────────────────────────────────────
DEFAULT_ROW_FIELDS = [
    "Issue Date", "Order #", "Order Status", "Line Aggregator",
    "Model Name", "Article Number", "Customer Number", "Country/Region",
    "Customer Request Date (CRD)", "Plan Date",
    "PO Statistical Delivery Date (PSDD)",
    "First Production Date", "Last Production Date", "PODD",
    "Delay - Confirmation", "Confirmation Delay Pd", "Delivery Delay Pd",
    "Delay - PO Del Update", "Production Lead Time", "Shipment Method",
    "Segment Attribute", "Segment Attribute Desc",
    "Sales Channel", "Sales Channel Description", "Storage Location",
]

# ── Session state ──────────────────────────────────────────────────────────────
for key, default in [("df_raw", None), ("pivot_df", None), ("pivot_meta", {})]:
    if key not in st.session_state:
        st.session_state[key] = default

# ── Header ─────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="header-bar">
    <h1 style="margin:0; font-size:1.6rem;">📊 PIVOT REPORT GENERATOR</h1>
    <div style="color:#888; font-size:0.8rem; margin-top:0.25rem; font-family:'IBM Plex Mono',monospace;">
        Tabular Form · Repeat All Labels · No Subtotals
    </div>
</div>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("### ⚙️ Configuration")
    st.markdown("---")

    uploaded = st.file_uploader("Upload Excel / CSV", type=["xlsx", "xls", "csv"])

    if uploaded is not None:
        try:
            if uploaded.name.lower().endswith(".csv"):
                df_raw = pd.read_csv(uploaded, dtype=str)
            else:
                xls = pd.ExcelFile(uploaded)
                sheet = st.selectbox("Select Sheet", xls.sheet_names)
                df_raw = pd.read_excel(xls, sheet_name=sheet, dtype=str)
            st.session_state.df_raw = df_raw
            st.success(f"✅ {len(df_raw):,} rows · {df_raw.shape[1]} cols")
        except Exception as e:
            st.error(f"❌ Load error: {e}")
            st.session_state.df_raw = None

    st.markdown("---")

    # ── Field selectors (hanya jika data ada) ─────────────────────────────────
    if st.session_state.df_raw is not None:
        all_cols = list(st.session_state.df_raw.columns)

        st.markdown("**📌 Column Field**")
        col_field = st.selectbox(
            "col_field",
            all_cols,
            index=all_cols.index("Manufacturing Size") if "Manufacturing Size" in all_cols else 0,
            label_visibility="collapsed",
        )

        st.markdown("**🔢 Value Field (Sum)**")
        rem = [c for c in all_cols if c != col_field]
        val_idx = 0
        for cand in ["Quantity", "Qty", "Sum of Quantity", "QTY"]:
            if cand in rem:
                val_idx = rem.index(cand)
                break
        value_field = st.selectbox("value_field", rem, index=val_idx, label_visibility="collapsed")

        st.markdown("**📑 Row Fields** *(urutan = urutan kolom)*")
        avail = [c for c in all_cols if c not in [col_field, value_field]]
        default_rows = [r for r in DEFAULT_ROW_FIELDS if r in avail] or avail[:5]
        row_fields = st.multiselect("row_fields", avail, default=default_rows, label_visibility="collapsed")

        st.markdown("---")
        generate_btn = st.button("🚀 Generate Pivot", use_container_width=True)

        with st.expander("🔍 Semua kolom di file"):
            st.markdown(
                "<div class='col-list'>" + "<br>".join(f"• {c}" for c in all_cols) + "</div>",
                unsafe_allow_html=True,
            )
    else:
        generate_btn = False
        col_field = "Manufacturing Size"
        value_field = "Quantity"
        row_fields = []

# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

if st.session_state.df_raw is None:
    c1, c2, c3 = st.columns(3)
    for col, icon, lbl in [(c1, "📋", "Tabular Form"), (c2, "🔁", "Repeat All Labels"), (c3, "🚫", "No Subtotals")]:
        with col:
            st.markdown(f"""<div class="metric-card">
                <div class="metric-value">{icon}</div>
                <div class="metric-label">{lbl}</div></div>""", unsafe_allow_html=True)
    st.markdown("---")
    st.info("⬆️ Upload file Excel/CSV di sidebar, lalu klik **Generate Pivot**.")
    st.stop()

# ── Raw data stats ─────────────────────────────────────────────────────────────
df = st.session_state.df_raw
mfg_unique = df[col_field].nunique() if col_field in df.columns else "—"
c1, c2, c3, c4 = st.columns(4)
for col, val, lbl in [
    (c1, f"{len(df):,}", "Total Rows"),
    (c2, str(df.shape[1]), "Columns"),
    (c3, str(len(row_fields)), "Row Fields Dipilih"),
    (c4, str(mfg_unique), f"{col_field} Unik"),
]:
    with col:
        st.markdown(f"""<div class="metric-card">
            <div class="metric-value">{val}</div>
            <div class="metric-label">{lbl}</div></div>""", unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# GENERATE
# ══════════════════════════════════════════════════════════════════════════════

def build_pivot(df, row_fields, col_field, value_field):
    """Build tabular pivot: all labels repeated, no subtotals."""
    # Cek kolom yang benar-benar ada
    valid_rows = [r for r in row_fields if r in df.columns]
    missing = [r for r in row_fields if r not in df.columns]

    if missing:
        st.warning(f"⚠️ Field berikut tidak ada di data (diabaikan): `{'`, `'.join(missing)}`")
    if not valid_rows:
        st.error("❌ Tidak ada row field yang valid. Periksa nama kolom di file kamu.")
        return None, None
    if col_field not in df.columns:
        st.error(f"❌ Column field **'{col_field}'** tidak ditemukan. Kolom yang ada: `{'`, `'.join(df.columns[:10])}`...")
        return None, None
    if value_field not in df.columns:
        st.error(f"❌ Value field **'{value_field}'** tidak ditemukan. Kolom yang ada: `{'`, `'.join(df.columns[:10])}`...")
        return None, None

    df = df.copy()

    # ── FIX UTAMA: isi NaN di semua row field dengan "(blank)" ─────────────
    # pivot_table secara default DROP baris yang ada NaN di index columns!
    for r in valid_rows:
        df[r] = df[r].fillna("(blank)").astype(str).str.strip()
        df[r] = df[r].replace({"nan": "(blank)", "NaN": "(blank)", "": "(blank)", "None": "(blank)"})

    # Isi NaN di col_field juga
    df[col_field] = df[col_field].fillna("(blank)").astype(str).str.strip()
    df[col_field] = df[col_field].replace({"nan": "(blank)", "": "(blank)", "None": "(blank)"})

    # Convert value ke numerik
    df[value_field] = pd.to_numeric(df[value_field], errors="coerce").fillna(0)

    try:
        pivot = df.pivot_table(
            index=valid_rows,
            columns=col_field,
            values=value_field,
            aggfunc="sum",
            fill_value=0,
            observed=True,

        )
    except Exception as e:
        st.error(f"❌ Pivot error: {e}")
        return None, None

    # Flatten MultiIndex kolom jika ada
    if isinstance(pivot.columns, pd.MultiIndex):
        pivot.columns = [" | ".join(str(v) for v in c) for c in pivot.columns]
    else:
        pivot.columns = [str(c) for c in pivot.columns]

    # Reset index → tabular (semua label terisi, tidak ada blank)
    pivot = pivot.reset_index()

    # Bersihkan & format kolom row fields
    def format_date_value(val):
        if not val or val in ("(blank)", "nan", "None", "", "<NA>"):
            return val
        try:
            return pd.to_datetime(val).strftime("%m/%d/%Y")
        except Exception:
            return val

    for r in valid_rows:
        if r in pivot.columns:
            col_vals = pivot[r].astype(str)
            # Deteksi apakah kolom berisi tanggal (format YYYY-MM-DD)
            sample = col_vals[col_vals.str.strip().str.match(r"^[0-9]{4}-[0-9]{2}-[0-9]{2}")]
            if len(sample) > 0:
                pivot[r] = col_vals.apply(format_date_value)
            else:
                pivot[r] = col_vals.replace({"nan": "", "None": "", "<NA>": ""})

    # Grand Total
    size_cols = [c for c in pivot.columns if c not in valid_rows]
    pivot["Grand Total"] = pivot[size_cols].sum(axis=1)

    return pivot, valid_rows


if generate_btn:
    if not row_fields:
        st.warning("⚠️ Pilih minimal satu Row Field di sidebar.")
    else:
        with st.spinner("🔄 Membangun pivot table..."):
            pivot, valid_rows = build_pivot(df, row_fields, col_field, value_field)

        if pivot is not None:
            st.session_state.pivot_df = pivot
            st.session_state.pivot_meta = {
                "valid_rows": valid_rows,
                "col_field": col_field,
                "value_field": value_field,
            }

# ── Tampilkan hasil ────────────────────────────────────────────────────────────
if st.session_state.pivot_df is not None:
    pivot      = st.session_state.pivot_df
    meta       = st.session_state.pivot_meta
    valid_rows = meta.get("valid_rows", [])

    row_order  = [c for c in valid_rows if c in pivot.columns]
    size_cols  = [c for c in pivot.columns if c not in valid_rows and c != "Grand Total"]
    size_order = sorted(size_cols)
    gt_col     = ["Grand Total"] if "Grand Total" in pivot.columns else []

    # Reorder columns
    final_order = row_order + size_order + gt_col
    pivot = pivot[[c for c in final_order if c in pivot.columns]]

    # Style
    def style_pivot(data):
        s = pd.DataFrame("", index=data.index, columns=data.columns)
        for c in row_order:
            if c in s.columns:
                s[c] = "background-color:#161b27; color:#ccc; font-size:0.76rem;"
        for c in size_order:
            if c in s.columns:
                s[c] = "background-color:#0f1117; color:#e8e8e8; text-align:right; font-family:'IBM Plex Mono',monospace; font-size:0.8rem;"
        if "Grand Total" in s.columns:
            s["Grand Total"] = (
                "background-color:#00d4aa18; color:#00d4aa; font-weight:600;"
                "text-align:right; font-family:'IBM Plex Mono',monospace; font-size:0.8rem;"
            )
        return s

    num_cols = size_order + gt_col
    styled = (
        pivot.style
        .apply(style_pivot, axis=None)
        .format({c: "{:,.0f}" for c in num_cols if c in pivot.columns}, na_rep="—")
    )

    st.success(f"✅ Pivot berhasil: **{len(pivot):,} baris** · **{len(size_cols)} Manufacturing Size** · **{len(row_order)} row fields**")
    st.markdown("### 📊 Hasil Pivot")
    st.dataframe(styled, use_container_width=True, height=600)

    # Summary
    st.markdown("---")
    grand_sum = pivot["Grand Total"].sum() if "Grand Total" in pivot.columns else 0
    s1, s2, s3 = st.columns(3)
    for col, val, lbl in [
        (s1, f"{len(pivot):,}", "Data Rows"),
        (s2, f"{grand_sum:,.0f}", "Grand Total Qty"),
        (s3, f"{len(size_cols)}", "Mfg Size Categories"),
    ]:
        with col:
            st.markdown(f"""<div class="metric-card">
                <div class="metric-value">{val}</div>
                <div class="metric-label">{lbl}</div></div>""", unsafe_allow_html=True)

    # Download
    st.markdown("<br>", unsafe_allow_html=True)
    d1, d2 = st.columns(2)
    with d1:
        csv_bytes = pivot.to_csv(index=False).encode("utf-8-sig")
        st.download_button("⬇️ Download CSV", data=csv_bytes,
            file_name="pivot_report.csv", mime="text/csv", use_container_width=True)
    with d2:
        buf = BytesIO()
        try:
            from openpyxl.styles import numbers as xl_numbers
            from openpyxl.utils import get_column_letter

            # Siapkan DataFrame untuk Excel: konversi kolom tanggal ke datetime
            df_excel = pivot.copy()
            DATE_FMT = "MM/DD/YYYY"
            date_col_indices = {}  # col letter -> True

            for c in df_excel.columns:
                # Deteksi kolom yang isinya string tanggal MM/DD/YYYY
                sample = df_excel[c].dropna().astype(str)
                sample = sample[sample.str.match(r"^[0-9]{2}/[0-9]{2}/[0-9]{4}$")]
                if len(sample) > 0:
                    df_excel[c] = pd.to_datetime(df_excel[c], format="%m/%d/%Y", errors="coerce")
                    date_col_indices[c] = True

            with pd.ExcelWriter(buf, engine="openpyxl", datetime_format="MM/DD/YYYY") as writer:
                df_excel.to_excel(writer, index=False, sheet_name="Pivot Report")
                ws = writer.sheets["Pivot Report"]

                # Ambil mapping nama kolom -> huruf kolom Excel
                header_map = {ws.cell(1, i).value: i for i in range(1, ws.max_column + 1)}

                for col_name, is_date in date_col_indices.items():
                    if col_name in header_map:
                        col_idx = header_map[col_name]
                        col_letter = get_column_letter(col_idx)
                        # Terapkan format Short Date ke semua cell data (bukan header)
                        for row in range(2, ws.max_row + 1):
                            cell = ws.cell(row=row, column=col_idx)
                            if cell.value is not None:
                                cell.number_format = "MM/DD/YYYY"

                # Auto width semua kolom
                for col_cells in ws.columns:
                    max_len = max((len(str(c.value)) for c in col_cells if c.value), default=8)
                    ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 40)

            buf.seek(0)
            st.download_button("⬇️ Download Excel", data=buf,
                file_name="pivot_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
        except Exception as e:
            st.warning(f"Excel export error: {e}")

    # Breakdown
    if size_cols:
        with st.expander("📈 Breakdown per Manufacturing Size"):
            bd = pivot[size_order].sum().reset_index()
            bd.columns = [meta.get("col_field", "Column"), "Total Quantity"]
            bd = bd.sort_values("Total Quantity", ascending=False)

            # Manual bar color tanpa matplotlib
            max_val = bd["Total Quantity"].max() if bd["Total Quantity"].max() > 0 else 1
            def highlight_bar(row):
                pct = row["Total Quantity"] / max_val
                g = int(180 + 75 * pct)
                return [
                    "background-color:#161b27; color:#ccc;",
                    f"background-color:rgba(0,{g},100,0.25); color:#e8e8e8; "
                    f"text-align:right; font-family:'IBM Plex Mono',monospace;"
                ]
            st.dataframe(
                bd.style.apply(highlight_bar, axis=1)
                  .format({"Total Quantity": "{:,.0f}"}),
                use_container_width=True, hide_index=True,
            )

elif not generate_btn:
    st.info("👈 Pilih field di sidebar lalu klik **🚀 Generate Pivot**")
