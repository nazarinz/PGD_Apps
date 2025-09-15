# ======================================================
# Reorder UK_* Size Columns in Excel (Upload + Keep Format) + Auto Download
# ======================================================

import re
from io import BytesIO
from datetime import datetime
from pathlib import Path
from copy import copy  # agar style bukan StyleProxy

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
import streamlit as st
from utils import set_page, header, footer

set_page("PGD Apps — Susun Sizelist", "📐")
header("📐 Susun Sizelist", "Reorder kolom UK_* sambil menjaga format Excel (openpyxl)")
uploaded = st.file_uploader("Unggah file Excel (.xlsx)", type=["xlsx"], accept_multiple_files=False)
run = st.button("⚙️ Proses & Siapkan Unduhan")

# --- Deteksi & aturan urut kolom ---
SIZE_REGEX = re.compile(r"^UK_(\d+)(?:(-K)|K|(-))?$")

def parse_size_col(colname: str):
    name = str(colname).strip()
    m = SIZE_REGEX.match(name)
    if not m:
        return (False, None, None)
    num = int(m.group(1))
    dash_k = m.group(2)
    dash_only = m.group(3)
    if dash_k:
        suffix = "-K"
    elif name.endswith("K") and not name.endswith("-K"):
        suffix = "K"
    elif dash_only:
        suffix = "-"
    else:
        suffix = ""
    return (True, num, suffix)

def compute_sorted_size_cols(cols):
    size_cols, numbers, by_key = [], set(), {}
    for c in cols:
        is_size, num, suffix = parse_size_col(c)
        if is_size:
            size_cols.append(c)
            numbers.add(num)
            by_key.setdefault(num, set()).add(suffix)
    if not size_cols:
        return []
    ordered, nums = [], sorted(numbers)
    # 1) K-first
    for n in nums:
        if "K" in by_key[n]:
            col = f"UK_{n}K"
            if col in cols: ordered.append(col)
        if "-K" in by_key[n]:
            col = f"UK_{n}-K"
            if col in cols: ordered.append(col)
    # 2) Non-K
    for n in nums:
        if "" in by_key[n]:
            col = f"UK_{n}"
            if col in cols: ordered.append(col)
        if "-" in by_key[n]:
            col = f"UK_{n}-"
            if col in cols: ordered.append(col)
    return ordered

# --- Copy util untuk menjaga style/format ---
def copy_cell(src, dst):
    dst.value = src.value
    if src.has_style:
        if src.font:          dst.font = copy(src.font)
        if src.border:        dst.border = copy(src.border)
        if src.fill:          dst.fill = copy(src.fill)
        if src.number_format: dst.number_format = src.number_format
        if src.protection:    dst.protection = copy(src.protection)
        if src.alignment:     dst.alignment = copy(src.alignment)

def copy_col_dimension(src_ws: Worksheet, dst_ws: Worksheet, src_col_letter: str, dst_col_letter: str):
    src_dim = src_ws.column_dimensions.get(src_col_letter)
    if not src_dim:
        return
    dst_dim = dst_ws.column_dimensions[dst_col_letter]
    dst_dim.width = src_dim.width
    dst_dim.hidden = src_dim.hidden
    dst_dim.outlineLevel = src_dim.outlineLevel
    try:
        dst_dim.bestFit = bool(getattr(src_dim, "bestFit", False))
    except Exception:
        pass

def copy_row_dimensions(src_ws: Worksheet, dst_ws: Worksheet):
    for idx, dim in src_ws.row_dimensions.items():
        dd = dst_ws.row_dimensions[idx]
        dd.height = dim.height
        dd.hidden = dim.hidden
        dd.outlineLevel = dim.outlineLevel

def map_header_order(ws: Worksheet, header_row: int = 1):
    max_col = ws.max_column
    headers = []
    for j in range(1, max_col + 1):
        val = ws.cell(row=header_row, column=j).value
        headers.append("" if val is None else str(val))
    sorted_sizes = compute_sorted_size_cols(headers)
    size_set = set(sorted_sizes)
    non_sizes = [h for h in headers if (h not in size_set and not parse_size_col(h)[0])]
    return headers, {name: i for i, name in enumerate(headers)}, non_sizes, sorted_sizes

def build_new_col_order(headers, non_sizes, sorted_sizes):
    size_set = set(sorted_sizes)
    leftovers = [h for h in headers if parse_size_col(h)[0] and h not in size_set]
    return non_sizes + sorted_sizes + leftovers

def create_reordered_sheet(src_ws: Worksheet, dst_ws: Worksheet, new_order_names, header_row: int = 1):
    headers = [str(src_ws.cell(row=header_row, column=j).value or "") for j in range(1, src_ws.max_column + 1)]
    name_to_idx = {name: i for i, name in enumerate(headers)}

    new_indices = []
    for nm in new_order_names:
        if nm in name_to_idx:
            new_indices.append(name_to_idx[nm])

    # Row dims & freeze panes
    copy_row_dimensions(src_ws, dst_ws)
    dst_ws.freeze_panes = src_ws.freeze_panes

    max_row = src_ws.max_row
    for new_j, src_idx in enumerate(new_indices, start=1):
        src_col_letter = get_column_letter(src_idx + 1)
        dst_col_letter = get_column_letter(new_j)
        copy_col_dimension(src_ws, dst_ws, src_col_letter, dst_col_letter)
        for i in range(1, max_row + 1):
            s = src_ws.cell(row=i, column=src_idx + 1)
            d = dst_ws.cell(row=i, column=new_j)
            copy_cell(s, d)

    # Tangani merged cells (vertikal ok, horizontal dilewati)
    col_old_to_new = {src_idx + 1: new_j for new_j, src_idx in enumerate(new_indices, start=1)}
    for m in src_ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = m.min_col, m.min_row, m.max_col, m.max_row
        if min_col == max_col:
            if min_col in col_old_to_new:
                new_c = col_old_to_new[min_col]
                dst_ws.merge_cells(start_row=min_row, start_column=new_c, end_row=max_row, end_column=new_c)
        else:
            # skip horizontal merges to avoid corruption after reordering
            pass

def reorder_workbook_keep_format(wb: "openpyxl.workbook.workbook.Workbook"):
    new_wb = Workbook()
    if new_wb.active and len(wb.worksheets) > 0:
        new_wb.remove(new_wb.active)
    for src_ws in wb.worksheets:
        title = src_ws.title
        new_title = title if all(ws.title != title for ws in new_wb.worksheets) else f"{title}_reordered"
        dst_ws = new_wb.create_sheet(title=new_title)

        headers, _, non_sizes, sorted_sizes = map_header_order(src_ws, header_row=1)
        new_order = build_new_col_order(headers, non_sizes, sorted_sizes)
        create_reordered_sheet(src_ws, dst_ws, new_order, header_row=1)
    return new_wb

def get_uploaded_streamlit(file):
    if not file:
        return None, None
    try:
        return file.name, file.read()
    except Exception:
        return None, None

if run:
    fname, content = get_uploaded_streamlit(uploaded)
    if not content:
        st.error("Silakan unggah file .xlsx lebih dulu.")
    else:
        bio = BytesIO(content)
        wb = load_workbook(bio, data_only=False)
        st.info(f"File: {fname} | Sheets: {[ws.title for ws in wb.worksheets]}")
        new_wb = reorder_workbook_keep_format(wb)
        out_buf = BytesIO()
        new_wb.save(out_buf)
        out_buf.seek(0)
        out_name = f"{Path(fname).stem}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        st.success("Selesai disusun. Silakan unduh file hasil.")
        st.download_button(
            label="⬇️ Download Hasil",
            data=out_buf.getvalue(),
            file_name=out_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        footer()
