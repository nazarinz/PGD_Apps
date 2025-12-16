import streamlit as st
import pandas as pd
import zipfile
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Mix-Packing Separator (ZIP)",
    page_icon="🧱",
    layout="wide"
)

st.title("🧱 Mix-Packing Separator — ZIP Processor")
st.markdown("""
Upload **ZIP berisi banyak file PO Import Table**.  
Jika **mix-packing terdeteksi**, sistem akan **menyisipkan separator**.  
Jika **tidak**, file **tetap original**.
""")

# ==========================
# EXCEL WRITER (ALL BORDER)
# ==========================
def write_styled_excel(df, sheet_name="Result"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold = Font(bold=True)

    # header
    for c, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=c, value=str(col))
        cell.font = bold
        cell.border = border

    # body
    for r, (_, row) in enumerate(df.iterrows(), start=2):
        for c, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=r, column=c, value=str(row[col]))
            cell.border = border

    # autosize
    for i, col in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(40, len(col) + 2))

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# ==========================
# MIX-PACKING LOGIC
# ==========================
def apply_mixpacking_separator(df):
    required = ["PO No.", "Packing Rule No."]
    for c in required:
        if c not in df.columns:
            raise ValueError(f"Missing column: {c}")

    df = df.fillna("").astype(str).reset_index(drop=True)

    grp = (
        df.groupby(["PO No.", "Packing Rule No."], dropna=False)
        .size()
        .rename("grp_size")
        .reset_index()
    )

    df = df.merge(grp, on=["PO No.", "Packing Rule No."], how="left")

    has_mixpacking = (df["grp_size"] > 1).any()
    if not has_mixpacking:
        return df.drop(columns=["grp_size"]), False

    result = []
    inserted = set()

    for _, row in df.iterrows():
        key = (row["PO No."], row["Packing Rule No."])

        if row["grp_size"] > 1 and key not in inserted:
            if result:
                result.append({c: "" for c in df.columns if c != "grp_size"})
            inserted.add(key)

        result.append(row.drop("grp_size").to_dict())

    return pd.DataFrame(result), True


# ==========================
# UI
# ==========================
uploaded_zip = st.file_uploader(
    "Upload ZIP (PO Import Tables)",
    type=["zip"]
)

if uploaded_zip:
    if st.button("🚀 Process ZIP"):
        out_zip = BytesIO()

        with zipfile.ZipFile(uploaded_zip, "r") as zin, \
             zipfile.ZipFile(out_zip, "w", zipfile.ZIP_DEFLATED) as zout:

            processed = 0
            separated = 0

            for name in zin.namelist():
                if not name.lower().endswith(".xlsx"):
                    continue

                with zin.open(name) as f:
                    df = pd.read_excel(f, dtype=str)

                df_out, is_mix = apply_mixpacking_separator(df)

                excel_bytes = write_styled_excel(df_out)

                zout.writestr(name, excel_bytes.read())

                processed += 1
                if is_mix:
                    separated += 1

        out_zip.seek(0)

        st.success(f"""
        ✅ Processing complete  
        📄 Files processed: {processed}  
        🧱 Mix-packing detected: {separated}
        """)

        st.download_button(
            "📦 Download Result ZIP",
            data=out_zip,
            file_name="PO_Import_With_Separators.zip",
            mime="application/zip"
        )
