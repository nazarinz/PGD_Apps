from utils.auth import require_login

require_login()

"""
PO Compare Tool — Streamlit App
Jalankan: streamlit run app.py
"""

import streamlit as st
import re
from datetime import date
from collections import defaultdict
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

st.set_page_config(page_title="PO Compare Tool", page_icon="📦", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=DM+Sans:wght@300;400;500;600&display=swap');
html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
.stApp { background: #0f1117; color: #e8eaf0; }
.main-header { background: linear-gradient(135deg,#1a1f2e,#151921); border:1px solid #2a3040; border-radius:12px; padding:28px 32px; margin-bottom:24px; }
.main-header h1 { font-family:'DM Mono',monospace; font-size:1.6rem; font-weight:500; color:#7eb8f7; margin:0 0 4px 0; }
.main-header p { color:#6b7280; margin:0; font-size:0.9rem; }
.upload-label { font-family:'DM Mono',monospace; font-size:0.75rem; color:#4a90d9; letter-spacing:1px; text-transform:uppercase; margin-bottom:8px; display:block; }
.stat-card { background:#151921; border:1px solid #2a3040; border-radius:10px; padding:16px 20px; text-align:center; }
.stat-number { font-family:'DM Mono',monospace; font-size:2rem; font-weight:500; line-height:1; margin-bottom:4px; }
.stat-label { font-size:0.75rem; color:#6b7280; text-transform:uppercase; letter-spacing:0.5px; }
.warn-box { background:#1a1500; border:1px solid #3a3000; border-radius:8px; padding:12px 16px; color:#c0a030; font-family:'DM Mono',monospace; font-size:0.8rem; margin-bottom:12px; }
.log-box { background:#0a0d12; border:1px solid #1e2535; border-radius:8px; padding:16px; font-family:'DM Mono',monospace; font-size:0.78rem; color:#5a7a9a; line-height:1.8; max-height:280px; overflow-y:auto; }
.stButton > button { background:#4a90d9 !important; color:white !important; border:none !important; border-radius:8px !important; font-weight:500 !important; }
.stDownloadButton > button { background:#1a4a2a !important; color:#4caf7d !important; border:1px solid #2a6a3a !important; border-radius:8px !important; font-family:'DM Mono',monospace !important; }
hr { border-color:#2a3040 !important; }
[data-testid="stExpander"] { background:#151921 !important; border:1px solid #2a3040 !important; border-radius:8px !important; }
.section-label { font-family:'DM Mono',monospace; font-size:0.7rem; color:#3a5a7a; letter-spacing:1px; text-transform:uppercase; padding:6px 12px; background:#0d1520; border-left:2px solid #2a4060; margin-bottom:0; }
</style>
""", unsafe_allow_html=True)

# ── CONSTANTS ─────────────────────────────────

COUNTRY_NORM = {
    'usa': 'United States', 'united states': 'United States', 'us': 'United States',
    'uk': 'United Kingdom', 'united kingdom': 'United Kingdom', 'united': 'United Kingdom',
    'germany': 'Germany', 'deutschland': 'Germany',
    'italy': 'Italy', 'italia': 'Italy',
    'france': 'France', 'netherlands': 'Netherlands', 'spain': 'Spain',
    'singapore': 'Singapore', 'malaysia': 'Malaysia', 'vietnam': 'Vietnam',
    'india': 'India', 'taiwan': 'Taiwan', 'indonesia': 'Indonesia',
    'macao': 'Macao', 'china': 'China', 'japan': 'Japan', 'korea': 'Korea',
    'thailand': 'Thailand', 'philippines': 'Philippines',
    'australia': 'Australia', 'new zealand': 'New Zealand',
    'canada': 'Canada', 'mexico': 'Mexico', 'brazil': 'Brazil',
    'austria': 'Austria', 'belgium': 'Belgium', 'switzerland': 'Switzerland',
    'poland': 'Poland', 'turkey': 'Turkey', 'ireland': 'Ireland',
    'portugal': 'Portugal', 'sweden': 'Sweden', 'norway': 'Norway',
    'denmark': 'Denmark', 'finland': 'Finland', 'greece': 'Greece',
    'hungary': 'Hungary', 'romania': 'Romania',
}

# Multi-word countries — checked first as substrings
MULTI_WORD_COUNTRIES = ['united kingdom', 'united states', 'new zealand']

# Single-word countries — length > 4 to avoid matching short tokens
SINGLE_WORD_COUNTRIES = {
    k: v for k, v in COUNTRY_NORM.items()
    if ' ' not in k and len(k) > 4 and k not in ('united',)
}

def normalize_country(raw):
    if not raw: return None
    key = str(raw).strip().lower()
    if key in COUNTRY_NORM:
        return COUNTRY_NORM[key]
    for k, v in COUNTRY_NORM.items():
        if len(k) > 3 and k in key:
            return v
    return str(raw).strip().title()


# ── COUNTRY DETECTION (Infor) ─────────────────
#
# Key insight from PDF structure (page 1):
#
#   [ship-to address]        ← e.g. "FORT WORTH, TX, 76177"
#   [COUNTRY]                ← e.g. "United states"   ← THIS is what we want
#   adidas                   ← customs consignee, always present
#   International
#   Trading AG
#   Ocean or as...
#
# The ship-to country ALWAYS appears immediately before "adidas international
# trading ag" in the extracted text. We just look for the last country name
# in the text that comes before that anchor phrase.
#
def extract_shipto_country(text):
    t = text.lower()

    # KEY INSIGHT from actual pdfplumber output:
    # Multi-column PDF table causes character interleaving, e.g.:
    #   "a d i d a s O cean or as" — spaces between letters from adjacent columns!
    # So searching for "ocean or as" or "adidas international trading" is UNRELIABLE.
    #
    # What IS reliable: "PO Line Details" is a section header below the table,
    # outside the multi-column area, so it always appears clean in extracted text.
    #
    # Actual extracted text structure (page 1):
    #   ...FORT WORTH, TX, 76177
    #   United states          ← ship-to country, last clean line before section header
    #   PO Line Details        ← reliable anchor ✓
    #
    # Strategy: find "po line details", look 300 chars before it,
    # pick the LAST (rightmost) country name = ship-to country.

    anchor_pos = t.find('po line details')
    if anchor_pos == -1:
        m = re.search(r'po line\s+details', t)
        anchor_pos = m.start() if m else -1
    if anchor_pos == -1:
        return None

    window_start = max(0, anchor_pos - 300)
    window = t[window_start:anchor_pos]

    # Build country pattern: multi-word first (longest first to avoid partial match)
    all_candidates = sorted(
        list(MULTI_WORD_COUNTRIES) + list(SINGLE_WORD_COUNTRIES.keys()),
        key=len, reverse=True
    )
    country_pattern = '|'.join(r'\b' + re.escape(c) + r'\b' for c in all_candidates)

    best_pos = -1
    best_country = None
    for m in re.finditer(country_pattern, window):
        if m.start() > best_pos:
            best_pos = m.start()
            raw = m.group(0).strip()
            best_country = COUNTRY_NORM.get(raw, raw.title())

    return best_country


# ── PARSERS ──────────────────────────────────

def detect_type(text):
    if "PURCHASE ORDER as of" in text and "Infor Nexus" in text: return "infor"
    if "Carton Form(" in text and "Cust.PO" in text: return "sap"
    return None

def split_infor_pages(pdf_bytes):
    pages = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for p in pdf.pages:
            pages.append(p.extract_text() or '')
    groups, i = [], 0
    while i < len(pages):
        if 'P.1 of' in pages[i]:
            combined = pages[i]; j = i + 1
            while j < len(pages) and 'P.1 of' not in pages[j]:
                combined += '\n' + pages[j]; j += 1
            groups.append(combined); i = j
        else:
            i += 1
    return groups

def parse_infor_block(text, filename):
    data = {'filename': filename}

    # PO Number
    header = text[:600]
    m = re.search(r'Phone: \+41.*?(\d{10})\s+\d+', header, re.DOTALL)
    if not m: m = re.search(r'SWITZERLAND\s+(\d{10})\s+\d+', text)
    data['po_number'] = m.group(1) if m else None

    # Article Number
    m = re.search(r'Article Number\s*\n\S+\s+\S+\s+(\S+)', text)
    data['article_number'] = m.group(1).strip() if m else None

    # Short Description + GPS Customer Number (same line)
    m = re.search(r'Gps Customer Number\s*\n(.+?)\s+(\d{6})\s*\n', text)
    data['short_desc']   = m.group(1).strip() if m else None
    data['gps_customer'] = m.group(2) if m else None

    # Ship Mode
    m = re.search(r'(Ocean|Air) or as', text)
    data['ship_mode'] = m.group(1) if m else None

    # Latest Date
    m = re.search(r'Ocean or as\s+(\d{4}-\d{2}-\d{2})', text)
    if not m: m = re.search(r'(\d{4}-\d{2}-\d{2})\s+ID\w+\s+\d{3}', text)
    data['latest_date'] = m.group(1) if m else None

    # Destination Country — zone-based rfind approach
    data['dest_country'] = extract_shipto_country(text)

    # Total Qty
    m = re.search(r'Total Item Quantity\s+([\d.]+)', text)
    data['total_qty'] = float(m.group(1)) if m else None

    # Qty by Sourcing/Mfg Size
    qty = {}
    for m in re.finditer(r'1\s+\d+\s+\S+\s+(?:\S+\s+)?([\d\-]+)\s+([\d\-]+)\s+\S+\s+T1\s+\d{10,13}\s+([\d.]+)', text):
        qty[m.group(2).strip()] = float(m.group(3))
    data['qty_by_size'] = qty
    return data

def parse_infor_pdf(pdf_bytes, filename):
    results = []
    for block in split_infor_pages(pdf_bytes):
        d = parse_infor_block(block, filename)
        if d['po_number']: results.append(d)
    return results

def parse_sap_page(text, filename):
    data = {'filename': filename}

    # PO Number
    m = re.search(r'Cust\.PO\s*:\s*(\d+)', text)
    data['po_number'] = m.group(1) if m else None

    # Article / Cust.Mat
    m = re.search(r'Cust\.Mat:\s*(\S+)', text)
    data['cust_mat'] = m.group(1) if m else None

    # Model name
    m = re.search(r'Model\s*:\s*(.+?)(?:\s{2,}|Arr\.Po)', text)
    data['model'] = m.group(1).strip() if m else None

    # Ship-to code (number in parentheses)
    m = re.search(r'Ship-to:\s*.+?\((\d+)\)', text)
    data['ship_to_code'] = m.group(1) if m else None

    # Destination country
    m = re.search(r'Coun:\s*(\S+)', text)
    data['country'] = m.group(1) if m else None

    # Ship Type
    m = re.search(r'ShipType:\s*\d+\s+(\w+)', text)
    data['ship_type'] = m.group(1) if m else None

    # PODD
    m = re.search(r'PODD\s*:\s*(\d{4}/\d{2}/\d{2})', text)
    data['podd'] = m.group(1).replace('/', '-') if m else None

    # Qty by size (SSP + MSP with continuation lines)
    lines = text.split('\n')
    merged = []
    for line in lines:
        s = line.strip()
        if merged and re.match(r'^[\d\-]+\(\d+\)', s) and not re.match(r'^\d+-\d+\s', s):
            merged[-1] += ',' + s
        else:
            merged.append(s)
    qty = defaultdict(float)
    for line in merged:
        m = re.match(r'(\d+-\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+([\d\-\(\),]+)', line)
        if not m: continue
        ctns, total_prs, sizes_raw = int(m.group(2)), int(m.group(4)), line
        if '(' in m.group(5):
            for sm in re.finditer(r'([\d\-]+)\((\d+)\)', sizes_raw):
                qty[sm.group(1)] += int(sm.group(2)) * ctns
        else:
            qty[m.group(5).strip()] += total_prs
    data['qty_by_size'] = dict(qty)
    return data

def parse_sap_pdf(pdf_bytes, filename):
    results = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ''
            if 'Carton Form(' in text and 'Cust.PO' in text:
                d = parse_sap_page(text, filename)
                if d['po_number']: results.append(d)
    return results

def compare_po(infor, sap_list):
    rows = []
    po = infor['po_number']
    merged_sap_qty = defaultdict(float)
    sap_filenames = []
    for sap in sap_list:
        for size, qty in sap['qty_by_size'].items(): merged_sap_qty[size] += qty
        sap_filenames.append(sap['filename'])
    total_sap = sum(merged_sap_qty.values())
    sap_str = ', '.join(dict.fromkeys(sap_filenames))
    sap0 = sap_list[0] if sap_list else {}

    def hrow(field, iv, sv, section='header', custom_ok=None):
        if custom_ok is not None:
            ok = custom_ok
        else:
            ok = bool(iv and sv and str(iv).strip().lower() == str(sv).strip().lower())
        return {
            'PO Number': po, 'Infor File': infor['filename'], 'SAP File(s)': sap_str,
            'Field': field, 'Section': section,
            'Infor Value': str(iv) if iv else '-',
            'SAP Value':   str(sv) if sv else '-',
            'Status': "✅ MATCH" if ok else "❌ MISMATCH", 'Size': ''
        }

    def qrow(field, iv, sv, size=''):
        try: ok = abs(float(iv or 0) - float(sv or 0)) < 0.01
        except: ok = False
        return {
            'PO Number': po, 'Infor File': infor['filename'], 'SAP File(s)': sap_str,
            'Field': field, 'Section': 'qty',
            'Infor Value': iv if iv is not None else '-',
            'SAP Value':   sv if sv is not None else '-',
            'Status': "✅ MATCH" if ok else "❌ MISMATCH", 'Size': size
        }

    # ── 1. Article Number ──────────────────────
    rows.append(hrow("Article Number", infor.get('article_number'), sap0.get('cust_mat')))

    # ── 2. Model / Short Description ───────────
    infor_desc = (infor.get('short_desc') or '').upper()
    sap_model  = (sap0.get('model') or '').strip().upper()
    model_ok   = bool(sap_model and infor_desc.startswith(sap_model))
    rows.append(hrow("Model / Short Desc", infor.get('short_desc'), sap0.get('model'), custom_ok=model_ok))

    # ── 3. GPS Customer Code ───────────────────
    rows.append(hrow("GPS Customer Code", infor.get('gps_customer'), sap0.get('ship_to_code')))

    # ── 4. Destination Country ─────────────────
    infor_ctr = normalize_country(infor.get('dest_country'))
    sap_ctr   = normalize_country(sap0.get('country'))
    ctr_ok    = bool(infor_ctr and sap_ctr and infor_ctr.lower() == sap_ctr.lower())
    rows.append(hrow("Destination Country", infor_ctr or '-', sap_ctr or '-', custom_ok=ctr_ok))

    # ── 5. Ship Mode ───────────────────────────
    infor_ship = (infor.get('ship_mode') or '').strip().lower()
    sap_ship   = (sap0.get('ship_type') or '').strip().lower()
    ship_ok    = infor_ship in sap_ship or sap_ship in infor_ship
    rows.append(hrow("Ship Mode", infor.get('ship_mode'), sap0.get('ship_type'), custom_ok=ship_ok))

    # ── 6. Latest Date / PODD ─────────────────
    rows.append(hrow("Latest Date / PODD", infor.get('latest_date'), sap0.get('podd')))

    # ── 7. Total Qty ───────────────────────────
    rows.append(qrow("Total Qty (Pairs)", infor.get('total_qty'), total_sap))

    # ── 8. Qty per Size ───────────────────────
    all_sizes = sorted(
        set(list(infor['qty_by_size'].keys()) + list(merged_sap_qty.keys())),
        key=lambda x: float(x.replace('-', '.5')) if re.match(r'^\d+\-?$', x) else 99
    )
    for size in all_sizes:
        rows.append(qrow(f'Qty Size {size}', infor['qty_by_size'].get(size, 0), float(merged_sap_qty.get(size, 0)), size))

    return rows

# ── EXCEL ────────────────────────────────────

GREEN="C6EFCE"; DGREEN="375623"; RED="FFC7CE"; DRED="9C0006"
DBLUE="1F3864"; WHITE="FFFFFF"; LGRAY="F2F2F2"; LBLUE="D6E4F0"

def mf(h): return PatternFill("solid", start_color=h, fgColor=h)
def fn(bold=False, color="000000", size=10): return Font(bold=bold, color=color, size=size, name="Arial")
def tb():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)
def ca(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
def la(): return Alignment(horizontal='left', vertical='center', wrap_text=True)

def build_excel(all_rows, summary):
    wb = Workbook()

    # ── Summary sheet ──
    ws = wb.active; ws.title = "Summary"
    ws.merge_cells("A1:G1")
    ws["A1"] = "PO COMPARE SUMMARY — adidas Infor vs SAP Carton"
    ws["A1"].font = fn(True,WHITE,13); ws["A1"].fill = mf(DBLUE); ws["A1"].alignment = ca()
    ws.row_dimensions[1].height = 30
    for c,h in enumerate(["PO Number","Infor File","SAP File(s)","Total Fields","✅ Match","❌ Mismatch","Result"],1):
        cell=ws.cell(2,c,h); cell.font=fn(True,WHITE); cell.fill=mf(DBLUE); cell.alignment=ca(); cell.border=tb()
    for r,(po,info) in enumerate(summary.items(),3):
        ok=info['mismatch']==0
        vals=[po,info['infor_file'],info['sap_files'],info['total'],info['match'],info['mismatch'],"✅ ALL OK" if ok else f"❌ {info['mismatch']} ISSUE(S)"]
        for c,v in enumerate(vals,1):
            cell=ws.cell(r,c,v); cell.border=tb()
            cell.alignment=ca() if c in [1,4,5,6,7] else la()
            if c==7: cell.fill=mf(GREEN) if ok else mf(RED); cell.font=fn(True,DGREEN if ok else DRED)
            elif r%2==0: cell.fill=mf(LGRAY)
    for i,w in enumerate([18,38,45,13,10,13,20],1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes="A3"

    # ── Detail sheet ──
    wd=wb.create_sheet("Detail")
    wd.merge_cells("A1:H1")
    wd["A1"]="PO COMPARE DETAIL — Field by Field"
    wd["A1"].font=fn(True,WHITE,12); wd["A1"].fill=mf(DBLUE); wd["A1"].alignment=ca()
    wd.row_dimensions[1].height = 25
    hdrs = ["PO Number","Infor File","SAP File(s)","Field","Infor Value","SAP Value","Status","Size"]
    for c,h in enumerate(hdrs,1):
        cell=wd.cell(2,c,h); cell.font=fn(True,WHITE); cell.fill=mf(DBLUE); cell.alignment=ca(); cell.border=tb()

    for r,row in enumerate(all_rows,3):
        is_ok  = "MATCH" in row['Status'] and "MIS" not in row['Status']
        is_err = "MISMATCH" in row['Status']
        is_header_section = row.get('Section') == 'header'

        for c,key in enumerate(hdrs,1):
            val = row.get(key,'')
            if key == 'Section': continue
            cell=wd.cell(r,c,val)
            cell.border=tb(); cell.alignment=ca() if c in [1,7,8] else la()
            if is_ok and c in [5,6,7]: cell.fill=mf(GREEN)
            if is_ok and c==7: cell.font=fn(True,DGREEN)
            if is_err and c in [5,6,7]: cell.fill=mf(RED)
            if is_err and c==7: cell.font=fn(True,DRED)
            if is_header_section and c in [4] and not is_ok and not is_err:
                cell.fill=mf(LBLUE)
            if not is_ok and not is_err and r%2==0: cell.fill=mf(LGRAY)

    for i,w in enumerate([18,38,45,22,25,25,16,8],1): wd.column_dimensions[get_column_letter(i)].width=w
    wd.freeze_panes="A3"

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

# ── UI ───────────────────────────────────────

st.markdown("""
<div class="main-header">
    <h1>📦 PO Compare Tool</h1>
    <p>adidas Infor PO vs SAP Carton Form — Upload PDF, compare otomatis by PO number</p>
</div>
""", unsafe_allow_html=True)

col1, col2 = st.columns(2)
with col1:
    st.markdown('<span class="upload-label">📄 Infor PO Files</span>', unsafe_allow_html=True)
    infor_files = st.file_uploader("Upload Infor PDF", type="pdf", accept_multiple_files=True, key="infor", label_visibility="collapsed")
    if infor_files:
        for f in infor_files:
            st.markdown(f'<div style="font-family:DM Mono,monospace;font-size:0.75rem;color:#4a90d9;padding:2px 0">✓ {f.name}</div>', unsafe_allow_html=True)

with col2:
    st.markdown('<span class="upload-label">📋 SAP Carton Files</span>', unsafe_allow_html=True)
    sap_files = st.file_uploader("Upload SAP PDF", type="pdf", accept_multiple_files=True, key="sap", label_visibility="collapsed")
    if sap_files:
        for f in sap_files:
            st.markdown(f'<div style="font-family:DM Mono,monospace;font-size:0.75rem;color:#4caf7d;padding:2px 0">✓ {f.name}</div>', unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)
run_btn = st.button("▶  Run Compare", disabled=not (infor_files and sap_files))

if run_btn and infor_files and sap_files:
    log_lines = []
    infor_by_po = {}
    sap_by_po = defaultdict(list)

    with st.spinner("Parsing PDFs..."):
        for f in infor_files:
            pdf_bytes = f.read()
            pos = parse_infor_pdf(pdf_bytes, f.name)
            log_lines.append(f"[Infor] {f.name} → {len(pos)} PO(s): {[p['po_number'] for p in pos]}")
            for d in pos:
                if d['po_number'] and d['po_number'] not in infor_by_po:
                    infor_by_po[d['po_number']] = d

        for f in sap_files:
            pdf_bytes = f.read()
            cartons = parse_sap_pdf(pdf_bytes, f.name)
            log_lines.append(f"[SAP] {f.name} → {len(cartons)} Carton(s): {[c['po_number'] for c in cartons]}")
            for d in cartons:
                if d['po_number']: sap_by_po[d['po_number']].append(d)

    all_rows, summary = [], {}
    no_sap, no_infor = [], []
    all_pos = sorted(set(list(infor_by_po.keys()) + list(sap_by_po.keys())))

    for po in all_pos:
        infor = infor_by_po.get(po)
        sap_list = sap_by_po.get(po, [])
        if not infor: no_infor.append(po); continue
        if not sap_list: no_sap.append(po); continue
        rows = compare_po(infor, sap_list)
        all_rows.extend(rows)
        mc = sum(1 for r in rows if "MATCH" in r['Status'] and "MIS" not in r['Status'])
        ec = sum(1 for r in rows if "MISMATCH" in r['Status'])
        log_lines.append(f"  PO {po}: {mc} match, {ec} mismatch")
        summary[po] = {
            'infor_file': infor['filename'],
            'sap_files': ', '.join(dict.fromkeys(s['filename'] for s in sap_list)),
            'total': len(rows), 'match': mc, 'mismatch': ec,
        }

    # ── Stats ──
    total_po = len(summary)
    ok_po    = sum(1 for v in summary.values() if v['mismatch'] == 0)
    err_po   = total_po - ok_po
    not_matched = len(no_sap) + len(no_infor)

    st.markdown("<br>", unsafe_allow_html=True)
    c1, c2, c3, c4 = st.columns(4)
    with c1: st.markdown(f'<div class="stat-card"><div class="stat-number" style="color:#7eb8f7">{total_po}</div><div class="stat-label">PO Compared</div></div>', unsafe_allow_html=True)
    with c2: st.markdown(f'<div class="stat-card"><div class="stat-number" style="color:#4caf7d">{ok_po}</div><div class="stat-label">All OK</div></div>', unsafe_allow_html=True)
    with c3: st.markdown(f'<div class="stat-card"><div class="stat-number" style="color:#f47171">{err_po}</div><div class="stat-label">Has Mismatch</div></div>', unsafe_allow_html=True)
    with c4: st.markdown(f'<div class="stat-card"><div class="stat-number" style="color:#f0c050">{not_matched}</div><div class="stat-label">No Pair Found</div></div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Warnings ──
    if no_sap or no_infor:
        with st.expander(f"⚠️  {not_matched} PO tidak punya pasangan"):
            if no_sap:
                st.markdown(f'<div class="warn-box">SAP Carton tidak ditemukan untuk {len(no_sap)} PO:<br>{"  |  ".join(no_sap)}</div>', unsafe_allow_html=True)
            if no_infor:
                st.markdown(f'<div class="warn-box">Infor PO tidak ditemukan untuk {len(no_infor)} PO:<br>{"  |  ".join(no_infor)}</div>', unsafe_allow_html=True)

    # ── Results ──
    st.markdown("### Results")
    for po, info in summary.items():
        ok = info['mismatch'] == 0
        label = f"✅ PO {po}  —  all {info['total']} fields match" if ok else f"❌ PO {po}  —  {info['mismatch']} MISMATCH dari {info['total']} fields"
        with st.expander(label):
            st.markdown(f'<div style="font-family:DM Mono,monospace;font-size:0.72rem;color:#6b7280;margin-bottom:12px">Infor: {info["infor_file"]}<br>SAP&nbsp;&nbsp;: {info["sap_files"]}</div>', unsafe_allow_html=True)

            po_rows = [r for r in all_rows if r['PO Number'] == po]
            header_rows = [r for r in po_rows if r.get('Section') == 'header']
            qty_rows    = [r for r in po_rows if r.get('Section') == 'qty']

            def render_table(rows_subset):
                t = '<table style="width:100%;border-collapse:collapse;font-size:0.8rem;font-family:DM Mono,monospace">'
                t += '<thead><tr style="background:#1a2030;color:#6b8aaa">'
                t += '<th style="padding:8px 12px;text-align:left;border-bottom:1px solid #2a3040;width:30%">Field</th>'
                t += '<th style="padding:8px 12px;text-align:left;border-bottom:1px solid #2a3040">Infor</th>'
                t += '<th style="padding:8px 12px;text-align:left;border-bottom:1px solid #2a3040">SAP</th>'
                t += '<th style="padding:8px 12px;text-align:center;border-bottom:1px solid #2a3040;width:120px">Status</th>'
                t += '</tr></thead><tbody>'
                for row in rows_subset:
                    is_ok  = "MATCH" in row['Status'] and "MIS" not in row['Status']
                    is_err = "MISMATCH" in row['Status']
                    bg = "#0d1e10" if is_ok else ("#1e0d0d" if is_err else "transparent")
                    st_html = '<span style="color:#4caf7d">✅ MATCH</span>' if is_ok else '<span style="color:#f47171">❌ MISMATCH</span>'
                    t += f'<tr style="background:{bg};border-bottom:1px solid #1a2030">'
                    t += f'<td style="padding:7px 12px;color:#7a9ab8">{row["Field"]}</td>'
                    t += f'<td style="padding:7px 12px;color:#e0e8f0">{row["Infor Value"]}</td>'
                    t += f'<td style="padding:7px 12px;color:#e0e8f0">{row["SAP Value"]}</td>'
                    t += f'<td style="padding:7px 12px;text-align:center">{st_html}</td>'
                    t += '</tr>'
                t += '</tbody></table>'
                return t

            if header_rows:
                st.markdown('<div class="section-label">📋 Header Fields</div>', unsafe_allow_html=True)
                st.markdown(render_table(header_rows), unsafe_allow_html=True)

            if qty_rows:
                st.markdown('<div class="section-label" style="margin-top:12px">📦 Quantity</div>', unsafe_allow_html=True)
                st.markdown(render_table(qty_rows), unsafe_allow_html=True)

    # ── Log ──
    st.markdown("<br>", unsafe_allow_html=True)
    with st.expander("📋 Processing Log"):
        log_html = "<br>".join(
            f'<span style="color:{"#4caf7d" if "[Infor]" in l else "#4a90d9" if "[SAP]" in l else "#6b7280"}">{l}</span>'
            for l in log_lines if l.strip()
        )
        st.markdown(f'<div class="log-box">{log_html}</div>', unsafe_allow_html=True)

    # ── Download ──
    st.markdown("<br>", unsafe_allow_html=True)
    if all_rows:
        excel_bytes = build_excel(all_rows, summary)
        today = date.today().strftime("%Y-%m-%d")
        st.download_button(
            label=f"⬇  Download Compare PDF_{today}.xlsx",
            data=excel_bytes,
            file_name=f"Compare PDF_{today}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
