import streamlit as st
import pandas as pd
import re
from datetime import datetime
from io import BytesIO

from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="MCS Recap Generator", layout="wide")
st.title("MCS Recap Generator (Smart Auto + Manual Override)")

# =====================================================
# PREPARE FUNCTIONS
# =====================================================

def prepare_mcs(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip()

    df = df.rename(columns={
        "Article No": "Article/No",
        "1ST Ex-factory date": "1st Ex-Factory Date",
        "Gender": "Age Group/Gender",
        "Bottom Tooling No.": "Bottom Tooling No",
        "Upper Tooling No.": "Upper Tooling No",
    })

    if "Size" not in df.columns:
        df["Size"] = None
    if "Factory Priority in Origo System" not in df.columns:
        df["Factory Priority in Origo System"] = None
    return df


def prepare_original(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip()

    df = df.rename(columns={
        "Article": "Article/No",
        "Age Group": "Age Group/Gender",
    })

    if "Size" not in df.columns:
        df["Size"] = None
    if "Factory Priority in Origo System" not in df.columns:
        df["Factory Priority in Origo System"] = None
    return df


final_cols = [
    "Season",
    "Category",
    "Model Name",
    "Model No",
    "Article/No",
    "Status",
    "Factory Priority in Origo System",
    "Development Type",
    "Developer",
    "Bottom Tooling No",
    "Upper Tooling No",
    "Last",
    "1st Ex-Factory Date",
    "LT",
    "Age Group/Gender",
    "Size Run",
    "Size"
]

soccer_keywords = [
    "PREDATOR", "COPA", "CRAZYFAST", "F50", "MESSI", "EDGE", "ACCURACY", "FREAK",
    "SALA", "ADIPURE", "SAMBA TEAM", "SUPERSTAR JUDE",
]
soccer_pattern = "|".join(re.escape(k) for k in soccer_keywords)

normalize_mapping = {
    "U-JUNIOR": "JUNIOR",
    "JUNIOR": "JUNIOR",
    "UNISEX": "UNISEX",
    "M": "MEN",
    "MALE": "MEN",
    "MEN": "MEN",
    "W": "WOMEN",
    "FEMALE": "WOMEN",
    "WOMEN": "WOMEN",
    "CHILDREN": "CHILDREN",
    "C": "CHILDREN",
    "U-INFANTS": "INFANT",
    "INFANT": "INFANT",
    "ADULT": "ADULT",
    "KIDS": "KIDS",
}

size_mapping = {
    "MEN": "8-",
    "WOMEN": "5-",
    "JUNIOR": "3",
    "CHILDREN": "11-K",
    "INFANT": "5-K",
    "KIDS": "11-K",
}

# =====================================================
# SIZE RUN NORMALIZATION
# =====================================================

def clean_token(token: str) -> str:
    token = token.strip().upper().replace(" ", "")
    return token

def expand_range_numeric(start, end, is_kids=False):
    sizes = []
    start = int(float(start))
    end = int(float(end))
    for i in range(start, end + 1):
        if is_kids:
            sizes.append(f"{i}K")
            sizes.append(f"{i}-K")
        else:
            sizes.append(f"{i}")
            sizes.append(f"{i}-")
    return sizes

def normalize_size_run(size_run):
    if pd.isna(size_run):
        return None

    raw = str(size_run).upper().replace(";", ",").replace(" ", "")
    tokens = re.split(r",", raw)
    final_sizes = []

    for token in tokens:
        token = clean_token(token)
        if not token:
            continue

        kids_match = re.match(r"(\d+\.?\d*)K-(\d+\.?\d*)K", token)
        if kids_match:
            start, end = kids_match.groups()
            final_sizes.extend(expand_range_numeric(start, end, is_kids=True))
            continue

        adult_match = re.match(r"(\d+\.?\d*)-(\d+\.?\d*)", token)
        if adult_match and "K" not in token:
            start, end = adult_match.groups()
            final_sizes.extend(expand_range_numeric(start, end, is_kids=False))
            continue

        single_k_match = re.match(r"(\d+)K", token)
        if single_k_match:
            num = single_k_match.group(1)
            final_sizes.append(f"{num}K")
            final_sizes.append(f"{num}-K")
            continue

        single_match = re.match(r"(\d+)", token)
        if single_match:
            num = single_match.group(1)
            final_sizes.append(f"{num}")
            final_sizes.append(f"{num}-")
            continue

    seen = set()
    normalized = []
    for s in final_sizes:
        if s not in seen:
            seen.add(s)
            normalized.append(s)
    return ",".join(normalized)

# =====================================================
# EXCEL EXPORT (STYLED)
# =====================================================

def export_styled_excel(df: pd.DataFrame, sheet_name="MCS Recap") -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        font_style = Font(name="Calibri", size=9)
        alignment_style = Alignment(horizontal="center", vertical="center")

        for row in ws.iter_rows():
            for cell in row:
                cell.font = font_style
                cell.alignment = alignment_style

        for col_cells in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 15

    return output.getvalue()

# =====================================================
# SMART DETECTION (TYPE + SHEET)
# =====================================================

MCS_KEY_COLS = {
    "Article No", "1ST Ex-factory date", "Gender", "Bottom Tooling No.", "Upper Tooling No.",
    "Season", "Category", "Model Name", "Model No", "Status", "Size Run"
}
ORIG_KEY_COLS = {
    "Article", "Age Group", "Season", "Category", "Model Name", "Model No", "Status", "Size Run"
}

def score_columns(cols, keyset):
    cols_norm = {str(c).strip() for c in cols}
    return len(cols_norm.intersection(keyset))

def smart_pick_sheet_and_type(uploaded_file):
    """
    Return: (best_sheet, detected_type, debug_scores)
    detected_type: "MCS" or "Original Football"
    """
    xls = pd.ExcelFile(uploaded_file)
    debug = []

    best_score = -1
    best_sheet = None
    best_type = None
    for sh in xls.sheet_names:
        try:
            # read just header (0 rows) -> cepat
            tmp = pd.read_excel(uploaded_file, sheet_name=sh, nrows=0)
            cols = tmp.columns

            mcs_score = score_columns(cols, MCS_KEY_COLS)
            orig_score = score_columns(cols, ORIG_KEY_COLS)

            # pilih type score tertinggi
            if mcs_score >= orig_score:
                chosen_type = "MCS"
                chosen_score = mcs_score
            else:
                chosen_type = "Original Football"
                chosen_score = orig_score

            debug.append((sh, chosen_type, chosen_score, mcs_score, orig_score))

            if chosen_score > best_score:
                best_score = chosen_score
                best_sheet = sh
                best_type = chosen_type
        except Exception:
            debug.append((sh, "ERROR", -1, -1, -1))
            continue

    if best_sheet is None or best_type is None:
        # fallback
        return (xls.sheet_names[0], "MCS", debug)

    return (best_sheet, best_type, debug)

# =====================================================
# UI
# =====================================================

uploaded_files = st.file_uploader(
    "Upload semua file MCS (dan/atau Original Football) sekaligus",
    type=["xlsx", "xls"],
    accept_multiple_files=True
)

preview_rows = st.number_input("Preview rows", min_value=5, max_value=200, value=20, step=5)

if not uploaded_files:
    st.info("Upload minimal 1 file Excel.")
    st.stop()

st.subheader("Konfigurasi tiap file (Auto default, bisa override)")

configs = []
for i, f in enumerate(uploaded_files):
    with st.expander(f"⚙️ {f.name}", expanded=(i == 0)):
        xls = pd.ExcelFile(f)
        sheets = xls.sheet_names

        auto_sheet, auto_type, debug = smart_pick_sheet_and_type(f)

        # default index untuk selectbox sheet
        sheet_index = sheets.index(auto_sheet) if auto_sheet in sheets else 0

        c1, c2 = st.columns([1, 1])

        with c1:
            file_type = st.selectbox(
                "Tipe file",
                ["MCS", "Original Football"],
                index=0 if auto_type == "MCS" else 1,
                key=f"type_{i}",
                help=f"Auto-detect: {auto_type}"
            )

        with c2:
            sheet = st.selectbox(
                "Sheet",
                sheets,
                index=sheet_index,
                key=f"sheet_{i}",
                help=f"Auto-pick: {auto_sheet}"
            )

        with st.popover("Lihat skor auto-detect"):
            st.write("Semakin tinggi score = makin cocok.")
            df_dbg = pd.DataFrame(
                debug,
                columns=["Sheet", "Chosen Type", "Chosen Score", "MCS Score", "Original Score"]
            ).sort_values(["Chosen Score", "Sheet"], ascending=[False, True])
            st.dataframe(df_dbg, use_container_width=True)

        configs.append({"file": f, "type": file_type, "sheet": sheet})

run = st.button("Generate MCS Recap", type="primary", disabled=(len(configs) == 0))

if run:
    try:
        prepared_dfs = []

        for cfg in configs:
            df = pd.read_excel(cfg["file"], sheet_name=cfg["sheet"])

            if cfg["type"] == "MCS":
                df_std = prepare_mcs(df).reindex(columns=final_cols)
            else:
                df_std = prepare_original(df).reindex(columns=final_cols)

            # Standardize types
            df_std["Bottom Tooling No"] = pd.to_numeric(df_std["Bottom Tooling No"], errors="coerce")
            df_std["Upper Tooling No"] = pd.to_numeric(df_std["Upper Tooling No"], errors="coerce")
            df_std["LT"] = pd.to_numeric(df_std["LT"], errors="coerce")
            df_std["1st Ex-Factory Date"] = pd.to_datetime(df_std["1st Ex-Factory Date"], errors="coerce")

            df_std["Age Group/Gender"] = (
                df_std["Age Group/Gender"]
                .fillna("")
                .astype(str)
                .str.upper()
                .str.strip()
            )

            df_std["Size"] = df_std["Size"].replace(r"^\s*$", pd.NA, regex=True)

            prepared_dfs.append(df_std)

        df_all = pd.concat(prepared_dfs, ignore_index=True)

        # Soccer detection
        cat_soccer = (
            df_all["Category"]
            .fillna("")
            .astype(str)
            .str.upper()
            .str.contains(r"FOOTBALL\s*/\s*SOCCER", regex=True, na=False)
        )

        name_soccer = (
            df_all["Model Name"]
            .fillna("")
            .astype(str)
            .str.upper()
            .str.contains(soccer_pattern, na=False, regex=True)
        )

        df_all["Is_Soccer"] = cat_soccer | name_soccer

        # Normalize age group/gender
        df_all["Age Group/Gender Normalized"] = df_all["Age Group/Gender"].map(normalize_mapping)

        # Size generation (fill blanks only)
        generated_size = df_all["Age Group/Gender Normalized"].map(size_mapping)

        mask_special = (
            (df_all["Age Group/Gender Normalized"] == "KIDS") &
            (df_all["Is_Soccer"] == True)
        )
        generated_size.loc[mask_special] = "3"

        existing_size = df_all["Size"].replace(r"^\s*$", pd.NA, regex=True)
        df_all["Size"] = existing_size.combine_first(generated_size)

        # Size run normalized
        df_all["Size Run Normalized"] = df_all["Size Run"].apply(normalize_size_run)

        output_cols = final_cols + ["Is_Soccer", "Age Group/Gender Normalized", "Size Run Normalized"]
        df_all = df_all.reindex(columns=output_cols)

        st.subheader("Preview Output")
        st.dataframe(df_all.head(int(preview_rows)), use_container_width=True)

        today_str = datetime.today().strftime("%Y%m%d")
        file_name = f"MCS Recap - {today_str}.xlsx"
        excel_bytes = export_styled_excel(df_all, sheet_name="MCS Recap")

        st.download_button(
            label="Download Excel (Styled)",
            data=excel_bytes,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.success(f"✅ Generated: {file_name}")

    except Exception as e:
        st.exception(e)
