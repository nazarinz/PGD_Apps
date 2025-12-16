# NOTE: Full script preserved. Only ADDITION is the mix-packing separator function
# and its invocation at the correct point (AFTER merge & final string-normalization).
# No existing logic removed.

import streamlit as st
import pandas as pd
import numpy as np
import os
import re
from io import BytesIO
import zipfile
from datetime import datetime

# tambahan untuk styling Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Packing Plan Processor",
    page_icon="📦",
    layout="wide"
)

st.title("📦 Packing Plan Processor")
st.markdown("Process packing plan files and merge with lookup data")

# Sidebar configuration
with st.sidebar:
    st.header("⚙️ Settings")
    debug_mode = st.checkbox("Debug Mode", value=True)
    st.markdown("---")
    st.info("Upload your packing plan files and lookup file to begin processing")

# ============= EXCEL WRITE HELPER =============
def write_workbook_bytes(sheets: dict):
    wb = Workbook()
    thin = Side(border_style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold_font = Font(bold=True)

    first = True
    for sheet_name, df in sheets.items():
        ws = wb.active if first else wb.create_sheet(title=sheet_name)
        ws.title = sheet_name
        first = False

        cols = list(df.columns)

        for c_idx, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=c_idx, value=str(col_name))
            cell.font = bold_font
            cell.border = border_all

        for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
            for c_idx, col_name in enumerate(cols, start=1):
                v = row[col_name]
                out_v = "" if pd.isna(v) else str(v)
                cell = ws.cell(row=r_idx, column=c_idx, value=out_v)
                cell.border = border_all

        for i, col_name in enumerate(cols, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(12, min(40, len(str(col_name)) + 2))

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

# ================= MIX-PACKING SEPARATOR (ADDED) =================
def insert_mixpacking_separators(df):
    """
    Insert blank row BEFORE each mix-packing group.
    Mix-packing = same PO No. + Packing Rule No. with >1 row.
    """

    df = df.copy().reset_index(drop=True)

    # hitung jumlah baris per PO + Packing Rule
    grp_size = (
        df.groupby(["PO No.", "Packing Rule No."], dropna=False)
        .size()
        .rename("grp_size")
        .reset_index()
    )

    df = df.merge(grp_size, on=["PO No.", "Packing Rule No."], how="left")

    result = []
    seen_mix_rule = set()

    for idx, row in df.iterrows():
        po = row["PO No."]
        rule = row["Packing Rule No."]
        size = row["grp_size"]

        key = (po, rule)

        # === INSERT SEPARATOR BEFORE FIRST ROW OF EACH MIX-PACKING RULE ===
        if size > 1 and key not in seen_mix_rule:
            # jangan taruh separator di baris paling atas
            if len(result) > 0:
                result.append({c: "" for c in df.columns if c != "grp_size"})
            seen_mix_rule.add(key)

        result.append(row.drop(labels="grp_size").to_dict())

    return pd.DataFrame(result, columns=[c for c in df.columns if c != "grp_size"])



# ============= HELPER FUNCTIONS =============

def find_header_row(df, keywords):
    for idx in df.index:
        row = " ".join(df.iloc[idx].astype(str).str.lower().tolist())
        score = sum(1 for k in keywords if k in row)
        if score >= 2:
            return idx
    return None

def find_col_precise(cols, candidates, data, prefer_integer=False):
    cols_lc = [str(c) for c in cols]
    patterns = [re.compile(r"(^|\W)"+re.escape(c.lower())+r"($|\W)") for c in candidates]
    matches = []
    for i,c in enumerate(cols_lc):
        for p in patterns:
            if p.search(c.lower()):
                matches.append(cols[i])
                break
    if not matches:
        for i,c in enumerate(cols_lc):
            for cand in candidates:
                if cand in c.lower():
                    matches.append(cols[i])
                    break
    if not matches:
        return None
    if len(matches) == 1:
        return matches[0]

    best = None
    best_score = -1
    for col in matches:
        series = data[col].dropna().astype(str).str.strip()
        if len(series) == 0:
            score = 0
        else:
            num_prop = series.str.match(r'^-?\d+(\.\d+)?$').mean() if len(series)>0 else 0
            int_prop = series.str.match(r'^\d+$').mean() if len(series)>0 else 0
            uniq_prop = series.nunique() / max(1, len(series))
            nonnull_prop = series.notna().mean()
            if prefer_integer:
                score = int_prop * 0.7 + num_prop * 0.2 + uniq_prop * 0.1
            else:
                score = nonnull_prop * 0.55 + uniq_prop * 0.35 + num_prop * 0.1
        if score > best_score:
            best_score = score
            best = col
    return best

def extract_package_detail(file_bytes, filename, debug=False):
    try:
        x = pd.read_excel(BytesIO(file_bytes), sheet_name="Detail", header=None, dtype=str)
    except Exception as e:
        if debug:
            st.warning(f"❌ Cannot open sheet 'Detail' in {filename}: {e}")
        return pd.DataFrame()

    df = x.copy()
    pkg_rows = df.index[df.apply(lambda r: r.astype(str).str.contains("package detail", case=False).any(), axis=1)].tolist()
    header_row = None

    if pkg_rows:
        for offset in (1,2,3):
            cand = pkg_rows[0] + offset
            if cand < len(df):
                rowtxt = " ".join(df.iloc[cand].astype(str).str.lower().tolist())
                if any(k in rowtxt for k in ["range","serial","buyer","po","pkg","gross","qty","manufacturing","size"]):
                    header_row = cand
                    if debug:
                        st.info(f"✓ {filename}: header at row {cand}")
                    break

    if header_row is None:
        header_keywords = ["range","serial","buyer","po","pkg","gross","qty per pkg","manufacturing size","buyer item","size"]
        header_row = find_header_row(df, header_keywords)

    if header_row is None:
        if debug:
            st.warning(f"⚠️ {filename}: header not found. Skipping.")
        return pd.DataFrame()

    header = df.iloc[header_row].fillna("").astype(str).tolist()
    data = df.iloc[header_row+1:].copy()
    data.columns = header

    candidate_patterns = [r'pkg', r'inner', r'qty', r'item', r'item qty', r'qty per', r'serial', r'\bfrom\b', r'\bto\b']
    fill_cols = []
    for c in data.columns:
        lc = str(c).lower()
        if any(re.search(p, lc) for p in candidate_patterns):
            fill_cols.append(c)

    if fill_cols:
        data[fill_cols] = data[fill_cols].ffill(axis=0).bfill(axis=0)

    data = data.dropna(how='all').copy()
    if data.empty:
        return pd.DataFrame()

    cols = data.columns.tolist()

    mapping_candidates = {
        "PO #": (["po number","po #","po no","^po$","purchase order"], False),
        "Range": (["^range$","^range\\b","range"], False),
        "Buyer Item #": (["buyer item","buyer item #","article","sku","style"], False),
        "Manufacturing Size": (["^manufacturing size$","manufacturing size","mfg size"], False),
        "Qty per Pkg/Inner Pack": (["qty per pkg","qty per inner","qty per pkg/inner","qty per pkg/inner pack","qty per pack","qtyperpkg"], True),
        "Pkg Count": (["pkg count","package count","pkg_count","pkgcount","inner pkg count","inner pkg","inner_pack","innerpack","inner pack","pkg"], True),
        "Gross": (["gross","gross weight","grossweight","gross_weight"], False)
    }

    out = pd.DataFrame(index=data.index)
    mapping_log = {}

    for outcol, (cands, prefer_int) in mapping_candidates.items():
        if outcol == "Pkg Count":
            out[outcol] = ""
            mapping_log[outcol] = None
            continue
        found = find_col_precise(cols, cands, data, prefer_integer=prefer_int)
        mapping_log[outcol] = found
        # KEEP AS STRING - NO CONVERSION
        out[outcol] = data[found].astype(str).replace(r'^\s*$', "", regex=True).replace('nan', '') if (found and found in data.columns) else ""

    # Pkg Count detection - 4-step approach
    pkg_count_col = None
    
    # Step 1: Exact match
    for c in cols:
        lc = str(c).strip().lower()
        if re.match(r'^(pkg\s*count|package\s*count|pkgcount)$', lc):
            pkg_count_col = c
            if debug:
                st.info(f"✓ Pkg Count exact: '{c}'")
            break
    
    # Step 2: Keyword match
    if pkg_count_col is None:
        for c in cols:
            lc = str(c).strip().lower()
            if 'pkg' in lc and 'count' in lc and 'qty' not in lc and 'per' not in lc:
                pkg_count_col = c
                if debug:
                    st.info(f"✓ Pkg Count keyword: '{c}'")
                break
    
    # Step 3: Position-based
    if pkg_count_col is None and mapping_log.get("Qty per Pkg/Inner Pack"):
        try:
            qty_col = mapping_log["Qty per Pkg/Inner Pack"]
            qty_idx = cols.index(qty_col)
            for offset in range(1, 4):
                idx = qty_idx + offset
                if idx < len(cols):
                    cand = cols[idx]
                    lc = str(cand).lower()
                    if any(skip in lc for skip in ['gross', 'weight', 'kg', 'lbs', 'dimension']):
                        continue
                    if ('pkg' in lc or 'count' in lc or 'inner' in lc) or lc.strip() == '':
                        pkg_count_col = cand
                        if debug:
                            st.info(f"✓ Pkg Count position: '{cand}' (+{offset})")
                        break
        except Exception:
            pass
    
    # Step 4: Pattern analysis
    if pkg_count_col is None:
        candidates = []
        qty_col = mapping_log.get("Qty per Pkg/Inner Pack")
        
        for c in cols:
            if c == qty_col or c in [mapping_log.get(k) for k in mapping_log if k != "Pkg Count"]:
                continue
            
            lc = str(c).lower()
            if any(skip in lc for skip in ['gross', 'weight', 'kg', 'dimension', 'serial', 'from', 'to', 'style', 'item', 'color', 'po']):
                continue
            
            ser = data[c].dropna().astype(str).str.strip()
            if ser.empty:
                continue
            
            nums = pd.to_numeric(ser.str.replace(r'[^\d\-\.]', '', regex=True), errors='coerce').dropna()
            if nums.empty:
                continue
            
            int_prop = nums.apply(lambda x: float(x).is_integer()).mean()
            prop_small = (nums <= 20).mean()
            prop_positive = (nums > 0).mean()
            variance = nums.var() if len(nums) > 1 else 0
            
            score = (int_prop * 0.4) + (prop_small * 0.3) + (prop_positive * 0.2) + (min(variance, 50) / 50 * 0.1)
            candidates.append((c, score, nums.mean()))
        
        if candidates:
            candidates.sort(key=lambda x: x[1], reverse=True)
            best_col, best_score, best_mean = candidates[0]
            
            if best_score > 0.4:
                pkg_count_col = best_col
                if debug:
                    st.info(f"✓ Pkg Count pattern: '{best_col}' ({best_score:.2f})")
    
    # Store Pkg Count - KEEP AS STRING
    if pkg_count_col and pkg_count_col in data.columns:
        raw_pkg = data[pkg_count_col].astype(str).str.strip()
        out["_original_pkg_empty"] = (raw_pkg == '') | (raw_pkg == 'nan')
        out["Pkg Count"] = raw_pkg.replace('nan', '').replace('', '')
        mapping_log["Pkg Count"] = pkg_count_col
        if debug:
            st.success(f"✓ Pkg Count: '{pkg_count_col}'")
    else:
        out["_original_pkg_empty"] = True
        out["Pkg Count"] = ""
        mapping_log["Pkg Count"] = None
        if debug:
            st.warning("⚠️ Pkg Count not found")

    # ALL COLUMNS AS STRING
    for c in out.columns:
        if c != "_original_pkg_empty":
            out[c] = out[c].astype(str).str.strip().replace('nan', '')

    return out.reset_index(drop=True)

def group_packing_rules(df):
    """Group rows without valid Packing Rule No. with last valid one."""
    df = df.copy()
    
    if "_is_separator" in df.columns:
        separator_mask = df["_is_separator"] == True
        separator_rows = df[separator_mask].copy()
        data_rows = df[~separator_mask].copy()
    else:
        separator_rows = pd.DataFrame()
        data_rows = df.copy()
    
    data_rows["Packing Rule No."] = data_rows["Packing Rule No."].replace('', '')
    valid_rule_mask = data_rows["Packing Rule No."].astype(str).str.match(r'^\d{4}$', na=False)
    data_rows["_ffill_key"] = data_rows["Packing Rule No."].where(valid_rule_mask, '')
    data_rows["_ffill_key"] = data_rows["_ffill_key"].replace('', pd.NA).ffill().fillna('')

    detail_row_mask = data_rows["_original_pkg_empty"] == True
    data_rows["Packing Rule No."] = data_rows["_ffill_key"]

    data_rows.loc[detail_row_mask, "Pkg Count"] = ""
    data_rows.loc[detail_row_mask, "Gross Weight"] = ""
    data_rows = data_rows.drop(columns=["_ffill_key"])
    
    if not separator_rows.empty:
        result = pd.concat([data_rows, separator_rows], ignore_index=True)
        result = result.sort_index()
    else:
        result = data_rows
    
    return result

def apply_multirow_pack_rules(df, debug=False):
    """Apply multi-row pack rules and insert separators."""
    df = df.copy()
    for c in ["PO No.","Packing Rule No.","Pkg Count","Gross Weight"]:
        if c not in df.columns:
            df[c] = ""

    if "_is_separator" not in df.columns:
        df["_is_separator"] = False

    existing_separators = df[df["_is_separator"] == True].copy()
    data_rows = df[df["_is_separator"] == False].copy()
    data_rows = data_rows.reset_index(drop=True)

    grp_keys = data_rows.groupby(["PO No.","Packing Rule No."], dropna=False).size().rename("grp_size")
    data_rows = data_rows.merge(grp_keys.reset_index(), on=["PO No.","Packing Rule No."], how="left")

    # Clear Pkg Count & Gross for rows 2+ in multi-row groups
    for (po, pr), group in data_rows.groupby(["PO No.","Packing Rule No."], dropna=False, sort=False):
        if len(group) > 1:
            idxs = group.index.tolist()
            for i in idxs[1:]:
                data_rows.at[i, "Pkg Count"] = ""
                data_rows.at[i, "Gross Weight"] = ""

    # Mark first row of multi-row groups
    data_rows['_group_first'] = False
    for (po, pr), group in data_rows.groupby(["PO No.","Packing Rule No."], dropna=False, sort=False):
        if len(group) > 1:
            first_idx = group.index[0]
            data_rows.at[first_idx, '_group_first'] = True

    insert_indices = data_rows[data_rows['_group_first'] == True].index.tolist()
    data_rows = data_rows.drop(columns=["grp_size", "_group_first"])

    # # Insert separator rows
    # if insert_indices:
    #     result_parts = []
    #     last_idx = 0
    #     separator_count = 0
        
    #     for insert_idx in insert_indices:
    #         result_parts.append(data_rows.iloc[last_idx:insert_idx])
    #         empty_row = pd.DataFrame([{col: "" for col in data_rows.columns}])
    #         empty_row["_is_separator"] = True
    #         result_parts.append(empty_row)
    #         separator_count += 1
    #         last_idx = insert_idx
        
    #     result_parts.append(data_rows.iloc[last_idx:])
    #     data_rows = pd.concat(result_parts, ignore_index=True)
        
    #     if debug:
    #         st.info(f"✓ Inserted {separator_count} separators")

    if not existing_separators.empty:
        for col in data_rows.columns:
            if col not in existing_separators.columns:
                existing_separators[col] = ""
        result = pd.concat([data_rows, existing_separators], ignore_index=True)
    else:
        result = data_rows
    
    return result

# ============= STREAMLIT UI =============

col1, col2 = st.columns(2)

with col1:
    st.subheader("📁 Packing Plan Files")
    packing_files = st.file_uploader(
        "Upload Packing Plan Files (.xlsx, .xls)",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
        key="packing"
    )
    if packing_files:
        st.success(f"✓ {len(packing_files)} file(s) uploaded")

with col2:
    st.subheader("📋 Lookup File")
    lookup_file = st.file_uploader(
        "Upload Lookup File (.xlsx)",
        type=["xlsx"],
        key="lookup"
    )
    if lookup_file:
        st.success(f"✓ {lookup_file.name} uploaded")

st.markdown("---")

if st.button("🚀 Process Files", type="primary", disabled=not (packing_files and lookup_file)):
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    try:
        # Extract
        status_text.text("📦 Extracting...")
        progress_bar.progress(10)
        
        results = []
        for idx, file in enumerate(packing_files):
            file_bytes = file.read()
            df_ex = extract_package_detail(file_bytes, file.name, debug=debug_mode)
            if not df_ex.empty:
                results.append(df_ex)
                if debug_mode:
                    st.info(f"✓ {file.name}: {len(df_ex)} rows")
            progress_bar.progress(10 + (idx + 1) * 20 // len(packing_files))
        
        if not results:
            st.error("❌ No data extracted")
            st.stop()
        
        df_final = pd.concat(results, ignore_index=True)
        
        # Normalize
        status_text.text("🔄 Normalizing...")
        progress_bar.progress(35)
        
        rename_map = {
            "PO #": "PO No.",
            "Range": "Packing Rule No.",
            "Buyer Item #": "Style",
            "Manufacturing Size": "Size",
            "Qty per Pkg/Inner Pack": "Qty per Pkg/Inner Pack",
            "Pkg Count": "Pkg Count",
            "Gross": "Gross Weight",
            "_original_pkg_empty": "_original_pkg_empty"
        }
        df_norm = df_final.rename(columns=rename_map)
        
        if "_is_separator" not in df_norm.columns:
            df_norm["_is_separator"] = False
        
        for c in ["PO No.", "Packing Rule No.", "Style", "Size", "Qty per Pkg/Inner Pack", "Pkg Count", "Gross Weight"]:
            if c not in df_norm.columns:
                df_norm[c] = ""
        
        if "_original_pkg_empty" not in df_norm.columns:
            df_norm["_original_pkg_empty"] = False
        
        if "Color" not in df_norm.columns:
            df_norm["Color"] = ""
        if "Item No." not in df_norm.columns:
            df_norm["Item No."] = ""
        
        cols_order = ["PO No.","Packing Rule No.","Style","Color","Size","Item No.","Qty per Pkg/Inner Pack","Pkg Count","Gross Weight","_original_pkg_empty","_is_separator"]
        df_norm = df_norm[[c for c in cols_order if c in df_norm.columns]]
        
        # Group
        status_text.text("📋 Grouping...")
        progress_bar.progress(45)
        df_norm = group_packing_rules(df_norm)
        
        # Multi-row rules BEFORE merge
        status_text.text("📊 Multi-row rules...")
        progress_bar.progress(50)
        df_norm = apply_multirow_pack_rules(df_norm, debug=debug_mode)
        
        # Read lookup
        status_text.text("📖 Reading lookup...")
        progress_bar.progress(60)
        
        lookup_df = pd.read_excel(lookup_file, dtype=str)
        
        required_cols = ["Order #","Material Color Description","Manufacturing Size","UPC/EAN (GTIN)","Country/Region"]
        missing = [c for c in required_cols if c not in lookup_df.columns]
        if missing:
            st.error(f"❌ Missing columns: {missing}")
            st.stop()
        
        for c in required_cols:
            lookup_df[c] = lookup_df[c].fillna('').astype(str)
        
        lookup_df["__Order_norm"] = lookup_df["Order #"].astype(str).str.strip()
        lookup_df["__Size_norm"] = lookup_df["Manufacturing Size"].astype(str).str.strip()
        
        lookup_df = lookup_df[
            (lookup_df["__Order_norm"] != '') &
            (lookup_df["__Order_norm"] != 'nan') &
            (lookup_df["__Size_norm"] != '') &
            (lookup_df["__Size_norm"] != 'nan')
        ].copy()
        
        lookup_df = lookup_df.drop_duplicates(subset=["__Order_norm","__Size_norm"], keep="first")
        
        # Merge
        status_text.text("🔗 Merging...")
        progress_bar.progress(75)
        
        df = df_norm.copy()
        df["__PO_norm"] = df["PO No."].astype(str).str.strip()
        df["__Size_norm"] = df["Size"].astype(str).str.strip()
        
        separator_rows = df[df["_is_separator"] == True].copy()
        data_rows = df[df["_is_separator"] == False].copy()
        
        df_valid = data_rows[
            (data_rows["__PO_norm"] != '') &
            (data_rows["__PO_norm"] != 'nan') &
            (data_rows["__Size_norm"] != '') &
            (data_rows["__Size_norm"] != 'nan')
        ].copy()
        
        df_invalid = data_rows[
            (data_rows["__PO_norm"] == '') |
            (data_rows["__PO_norm"] == 'nan') |
            (data_rows["__Size_norm"] == '') |
            (data_rows["__Size_norm"] == 'nan')
        ].copy()
        
        merged = df_valid.merge(
            lookup_df[["__Order_norm","__Size_norm","Material Color Description","UPC/EAN (GTIN)","Country/Region"]],
            left_on=["__PO_norm","__Size_norm"],
            right_on=["__Order_norm","__Size_norm"],
            how="left",
            indicator=True
        )
        
        merged = merged.drop(columns=["_merge"])
        
        if len(df_invalid) > 0:
            df_invalid["Material Color Description"] = ""
            df_invalid["UPC/EAN (GTIN)"] = ""
            df_invalid["Country/Region"] = ""
            merged = pd.concat([merged, df_invalid], ignore_index=True)
        
        if len(separator_rows) > 0:
            for col in merged.columns:
                if col not in separator_rows.columns:
                    separator_rows[col] = ""
            merged = pd.concat([merged, separator_rows], ignore_index=True)
            merged = merged.sort_index()
        
        # KEEP AS STRING
        merged["Color"] = merged["Material Color Description"].where(
            (merged["Material Color Description"] != '') & (merged["Material Color Description"] != 'nan'),
            ""
        )
        merged["Item No."] = merged["UPC/EAN (GTIN)"].where(
            (merged["UPC/EAN (GTIN)"] != '') & (merged["UPC/EAN (GTIN)"] != 'nan'),
            ""
        )
        
        final_cols = ["PO No.","Packing Rule No.","Style","Color","Size","Item No.","Qty per Pkg/Inner Pack","Pkg Count","Gross Weight","Country/Region"]
        final_cols = [c for c in final_cols if c in merged.columns]
        df_final_with_lookup = merged[final_cols].copy()
        
        # Ensure all are strings
        for col in df_final_with_lookup.columns:
            df_final_with_lookup[col] = df_final_with_lookup[col].fillna('').astype(str).replace('nan', '')
       
        # === INSERT MIX-PACKING SEPARATORS (BLANK ROWS) ===
        df_final_with_lookup = insert_mixpacking_separators(df_final_with_lookup)

        # Unmatched
        status_text.text("🔍 Checking...")
        progress_bar.progress(90)
        
        unmatched = df_final_with_lookup[
            (df_final_with_lookup["PO No."] != '') &
            ((df_final_with_lookup["Color"] == '') | (df_final_with_lookup["Item No."] == ''))
        ].copy()
        
        # Downloads (styled)
        status_text.text("💾 Preparing styled downloads...")
        progress_bar.progress(95)

        # 1) Main report workbook (Rekap + Unmatched)
        sheets = {"Rekap": df_final_with_lookup}
        if not unmatched.empty:
            sheets["Unmatched"] = unmatched
        report_bio = write_workbook_bytes(sheets)

        # 2) PO exports into ZIP
        po_country_map = lookup_df.set_index("__Order_norm")["Country/Region"].to_dict()
        unique_pos = [po for po in df_final_with_lookup["PO No."].unique() if po != '']

        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for po in unique_pos:
                df_po = df_final_with_lookup[df_final_with_lookup["PO No."] == po].copy()
                if "Country/Region" in df_po.columns:
                    df_po = df_po.drop(columns=["Country/Region"])
                po_bio = write_workbook_bytes({"Rekap": df_po})
                country = po_country_map.get(str(po), "Unknown")
                country = str(country).strip() if country != '' else "Unknown"
                safe_country = country.replace('/', '-')
                file_name = f"Po import table_ {po} {safe_country}.xlsx"
                zip_file.writestr(file_name, po_bio.read())
        zip_buffer.seek(0)

        progress_bar.progress(100)
        status_text.text("✅ Complete!")
        
        st.success("🎉 Processing completed!")
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Total Rows", len(df_final_with_lookup))
        with col2:
            st.metric("Color Filled", (df_final_with_lookup["Color"] != '').sum())
        with col3:
            st.metric("Item No. Filled", (df_final_with_lookup["Item No."] != '').sum())
        with col4:
            st.metric("Unmatched", len(unmatched))
        
        st.markdown("---")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.download_button(
                label="📥 Download Styled Report",
                data=report_bio,
                file_name=f"rekap_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        with col2:
            st.download_button(
                label="📦 Download Styled PO Exports",
                data=zip_buffer,
                file_name=f"po_exports_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                mime="application/zip"
            )
        
        st.markdown("---")
        st.subheader("📊 Preview")
        
        tab1, tab2 = st.tabs(["Complete Data", "Unmatched"])
        
        with tab1:
            st.dataframe(df_final_with_lookup.head(100), use_container_width=True)
        
        with tab2:
            if not unmatched.empty:
                st.dataframe(unmatched.head(100), use_container_width=True)
            else:
                st.info("No unmatched rows!")
        
    except Exception as e:
        st.error(f"❌ Error: {str(e)}")
        if debug_mode:
            st.exception(e)

st.markdown("---")
st.markdown("<div style='text-align: center; color: #666;'><p>Packing Plan Processor v1.0</p></div>", unsafe_allow_html=True)
