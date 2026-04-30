from utils.auth import require_login

require_login()

import streamlit as st
import pandas as pd
import io
import re
from collections import defaultdict
from datetime import datetime, timedelta
import traceback

st.set_page_config(page_title="Loading Plan Checker", page_icon="📋", layout="wide")


# =========================================================
# STANDARD LOADING PLAN READER (Header baris ke-3)
# =========================================================
def read_daily_loading_plan(uploaded_file):
    name = uploaded_file.name.lower()
    ext = "." + name.split(".")[-1]

    if hasattr(uploaded_file, "getvalue"):
        raw = uploaded_file.getvalue()
    else:
        uploaded_file.seek(0)
        raw = uploaded_file.read()

    if ext in [".xlsx", ".xlsm"]:
        engine = "openpyxl"
    elif ext == ".xls":
        engine = "xlrd"
    elif ext == ".xlsb":
        engine = "pyxlsb"
    elif ext == ".ods":
        engine = "odf"
    else:
        raise ValueError(f"Format {ext} tidak didukung")

    sheets = pd.read_excel(io.BytesIO(raw), sheet_name=None, header=None, engine=engine)

    sheet_name = None
    for sh in sheets.keys():
        if str(sh).strip().lower().replace(" ", "_") == "loading_plan":
            sheet_name = sh
            break
    if sheet_name is None:
        sheet_name = list(sheets.keys())[0]

    df_raw = sheets[sheet_name]

    header_row_index = 2
    raw_header = [str(x).strip() if x is not None else "" for x in df_raw.iloc[header_row_index].tolist()]

    seen = set()
    clean_cols = []
    for c in raw_header:
        base = c if c else "col"
        new = base
        i = 1
        while new in seen:
            i += 1
            new = f"{base}_{i}"
        seen.add(new)
        clean_cols.append(new)

    df = df_raw.iloc[header_row_index + 1:].copy()
    df.columns = clean_cols
    df = df.dropna(how="all")
    df = df.dropna(axis=1, how="all")
    df = df.reset_index(drop=True)

    return df, sheet_name


# =========================================================
# STANDARD SO DETECTOR
# =========================================================
CANDIDATES = [
    "sap_odr_no", "sap_odrno", "sap_order_no", "sap_odr_number",
    "sales_order", "salesorder", "so", "so_no", "sono", "so_number",
]

FVB_CANDIDATES = [
    "fvb_so", "fvbso", "fvb_so", "fvb_sales_order", "fvb",
]


def norm_col_name(c: str) -> str:
    c = str(c).strip().lower()
    c = re.sub(r"\s+", " ", c)
    c = re.sub(r"[^0-9a-z_ ]+", "", c)
    c = c.replace(" ", "_")
    c = re.sub(r"_+", "_", c)
    return c


def detect_so_column(df: pd.DataFrame):
    cols = [norm_col_name(c) for c in df.columns]

    for cand in CANDIDATES:
        if cand in cols:
            return df.columns[cols.index(cand)]

    for i, cn in enumerate(cols):
        if ("sap" in cn) and ("odr" in cn or "order" in cn):
            return df.columns[i]
        if ("sales" in cn) and ("order" in cn):
            return df.columns[i]

    return None


def detect_fvb_so_column(df: pd.DataFrame):
    cols_norm = [norm_col_name(c) for c in df.columns]

    for cand in FVB_CANDIDATES:
        nc = norm_col_name(cand)
        if nc in cols_norm:
            return df.columns[cols_norm.index(nc)]

    for i, cn in enumerate(cols_norm):
        if "fvb" in cn:
            return df.columns[i]

    return None


def clean_so_series(series: pd.Series) -> pd.Series:
    return (
        series
        .astype(str)
        .str.strip()
        .str.replace(r"\.0$", "", regex=True)
    )


# =========================================================
def extract_date_from_filename_ddmm(filename):
    name = filename.rsplit(".", 1)[0]
    try:
        parts = name.split(".")
        if len(parts) != 2:
            return None
        day, month = map(int, parts)
        today = pd.Timestamp.today()
        year = today.year
        if month > today.month:
            year -= 1
        return pd.Timestamp(year=year, month=month, day=day)
    except:
        return None


# =========================================================
def process_files(zrsd_file, plan_files, podd_start: pd.Timestamp, podd_end: pd.Timestamp):
    text_columns = {
        'PO No.(Full)': str,
        'PO No.(Short)': str,
        'Article No': str,
        'SAP Material': str,
        'Pattern Code(Up.No.)': str,
        'Model No': str,
        'Outsole Mold': str,
        'SO': str,
        'Material': str
    }

    df_013 = pd.read_excel(io.BytesIO(zrsd_file.getvalue()), dtype=text_columns)

    fcr_date_check = pd.to_datetime(df_013['FCR Date'], errors='coerce').isna()
    podd_parsed    = pd.to_datetime(df_013['PODD'], errors='coerce')

    # ── Filter PODD sesuai rentang yang dipilih user ──────────────────────
    df_filtered = df_013[
        fcr_date_check &
        (podd_parsed >= podd_start) &
        (podd_parsed <= podd_end)
    ].copy()

    known_date_columns = [
        'Document Date', 'FPD', 'LPD', 'CRD', 'PSDD', 'FCR Date',
        'PODD', 'PD', 'PO Date', 'Actual PGI', 'Delivery Date',
        'Ship Date', 'Created Date', 'Modified Date', 'Invoice Date'
    ]

    for col in df_filtered.columns:
        if any(col.lower() == known.lower() for known in known_date_columns):
            if not pd.api.types.is_datetime64_any_dtype(df_filtered[col]):
                df_filtered[col] = pd.to_datetime(df_filtered[col], errors='coerce')

    # plan_so_map: SO (int) → list of (source, date) tuples
    plan_so_map = defaultdict(list)

    # ===================== READ LOADING PLAN =====================
    for file in plan_files:
        try:
            st.info(f"📂 Membaca {file.name}...")

            df_plan, sheet = read_daily_loading_plan(file)
            plan_date = extract_date_from_filename_ddmm(file.name)

            if plan_date is None:
                st.error(f"❌ {file.name}: Format nama file salah (harus DD.MM seperti 22.01)")
                continue

            so_col_sap = detect_so_column(df_plan)
            so_col_fvb = detect_fvb_so_column(df_plan)

            if not so_col_sap and not so_col_fvb:
                st.error(f"❌ {file.name}: Tidak ada kolom SAP SO maupun FVB SO")
                st.write("Kolom yang tersedia:", list(df_plan.columns))
                continue

            sap_label = f"`{so_col_sap}`" if so_col_sap else "tidak ditemukan"
            fvb_label = f"`{so_col_fvb}`" if so_col_fvb else "tidak ditemukan"
            st.success(
                f"✅ {file.name} → SAP SO: {sap_label} | "
                f"FVB SO: {fvb_label} | Tanggal: {plan_date.date()}"
            )

            if so_col_sap and so_col_sap in df_plan.columns:
                df_plan["__SO_sap__"] = clean_so_series(df_plan[so_col_sap])
                for so in df_plan["__SO_sap__"].dropna().unique():
                    if str(so).isdigit():
                        plan_so_map[int(so)].append(("SAP", plan_date.date()))

            if so_col_fvb and so_col_fvb in df_plan.columns:
                df_plan["__SO_fvb__"] = clean_so_series(df_plan[so_col_fvb])
                for so in df_plan["__SO_fvb__"].dropna().unique():
                    if str(so).isdigit():
                        plan_so_map[int(so)].append(("FVB", plan_date.date()))

        except Exception as e:
            st.error(f"❌ {file.name}: {str(e)}")
            with st.expander("🔍 Detail Error"):
                st.code(traceback.format_exc())
            continue

    # ===================== CHECK RESULT =====================
    def check_remark(row):
        try:
            so_raw = str(row['SO']).strip()
            so = int(so_raw.replace(".0", ""))
            podd_date = pd.to_datetime(row['PODD'], errors='coerce').date()
        except:
            return "⚠️ Invalid Data", "invalid", "", ""

        if so not in plan_so_map:
            return "❌ NOT IN LOADING PLAN", "not_found", "", ""

        entries = plan_so_map[so]

        all_dates = sorted({d for _, d in entries})
        all_dates_str = ", ".join(d.strftime("%-m/%-d/%Y") for d in all_dates)
        all_dates_remark = ", ".join(str(d) for d in all_dates)

        sources_found = []
        if any(src == "SAP" for src, _ in entries):
            sources_found.append("SAP")
        if any(src == "FVB" for src, _ in entries):
            sources_found.append("FVB")
        source_label = "+".join(sources_found)

        if podd_date in all_dates:
            return "✅ MATCH – Date Match", "match", all_dates_str, source_label
        else:
            return (
                f"⚠️ IN PLAN – Date Mismatch (Plan: {all_dates_remark})",
                "mismatch",
                all_dates_str,
                source_label,
            )

    result_df = df_filtered.apply(
        lambda row: pd.Series(check_remark(row)), axis=1
    )

    if result_df.empty or 0 not in result_df.columns:
        df_filtered['Remark Loading Plan'] = pd.Series(dtype=str)
        df_filtered['Status']              = pd.Series(dtype=str)
        df_filtered['Plan Dates']          = pd.Series(dtype=str)
        df_filtered['SO Source']           = pd.Series(dtype=str)
    else:
        df_filtered['Remark Loading Plan'] = result_df[0]
        df_filtered['Status']              = result_df[1]
        df_filtered['Plan Dates']          = result_df[2]
        df_filtered['SO Source']           = result_df[3]

    return df_filtered


# =========================================================
def build_excel(df: pd.DataFrame) -> bytes:
    """Buat file Excel dengan formatting."""
    output = io.BytesIO()

    podd_idx = df.columns.get_loc('PODD') if 'PODD' in df.columns else len(df.columns) - 1
    df_export = df.copy()
    df_export.insert(podd_idx + 1, 'DN',  "")
    df_export.insert(podd_idx + 2, 'FGR', "")

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False)

        from openpyxl.styles import Font, PatternFill, Alignment

        worksheet = writer.sheets['Sheet1']

        known_date_columns = [
            'Document Date', 'FPD', 'LPD', 'CRD', 'PSDD', 'FCR Date',
            'PODD', 'PD', 'PO Date', 'Actual PGI', 'Delivery Date',
            'Ship Date', 'Created Date', 'Modified Date', 'Invoice Date'
        ]

        yellow_columns = {'DN', 'FGR', 'Remark Loading Plan', 'Status', 'Plan Dates', 'SO Source'}

        YELLOW     = PatternFill(fill_type="solid", fgColor="FFFF00")
        GREY       = PatternFill(fill_type="solid", fgColor="D9D9D9")
        NO_FILL    = PatternFill(fill_type=None)
        FONT       = Font(name="Calibri", size=9)
        FONT_HDR   = Font(name="Calibri", size=9, bold=True)
        ALIGN_HDR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ALIGN_DATA = Alignment(horizontal="center", vertical="center", wrap_text=False)

        for cell in worksheet[1]:
            col_name       = df_export.columns[cell.column - 1]
            cell.font      = FONT_HDR
            cell.alignment = ALIGN_HDR
            cell.fill      = YELLOW if col_name in yellow_columns else GREY

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
            for cell in row:
                col_name       = df_export.columns[cell.column - 1]
                cell.font      = FONT
                cell.alignment = ALIGN_DATA
                cell.fill      = NO_FILL

                if (any(col_name.lower() == k.lower() for k in known_date_columns)
                        and cell.value):
                    cell.number_format = 'M/D/YYYY'

        worksheet.freeze_panes = "H2"

    return output.getvalue()


# =========================================================
# ======================= UI ================================
# =========================================================
st.title("📋 Loading Plan Checker")

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ Pengaturan")
    st.markdown("---")

    zrsd_file = st.file_uploader("📁 Upload ZRSD", type=["xlsx", "xls"])
    plan_files = st.file_uploader(
        "📁 Upload Loading Plan (nama: DD.MM)",
        type=["ods", "xlsx", "xls"],
        accept_multiple_files=True,
    )

    st.markdown("---")
    st.subheader("📅 Filter Tanggal PODD")

    today = datetime.today().date()

    # Batas bawah: tidak dibatasi (semua PODD ≤ hari ini otomatis masuk)
    # Batas atas: diatur user, default H+3
    podd_start = pd.Timestamp.min  # tidak ada batas bawah

    podd_end = st.date_input(
        "Tampilkan PODD sampai",
        value=today + timedelta(days=3),
        help="PODD ≤ hari ini selalu masuk. Atur seberapa jauh ke depan yang ingin dicek.",
    )

    date_valid = True
    days_ahead = (podd_end - today).days
    if days_ahead < 0:
        st.caption(
            f"Filter aktif: semua PODD s/d **{podd_end.strftime('%d %b %Y')}** "
            f"({abs(days_ahead)} hari sebelum hari ini)"
        )
    elif days_ahead == 0:
        st.caption("Filter aktif: semua PODD s/d **hari ini**")
    else:
        st.caption(
            f"Filter aktif: semua PODD s/d **{podd_end.strftime('%d %b %Y')}** "
            f"(H+{days_ahead})"
        )

    st.markdown("---")

    run_btn = st.button(
        "🚀 Proses Data",
        use_container_width=True,
        disabled=not date_valid,
    )

    if run_btn:
        if zrsd_file and plan_files:
            with st.spinner("Memproses data..."):
                st.session_state['results'] = process_files(
                    zrsd_file,
                    plan_files,
                    podd_start,          # sudah pd.Timestamp.min
                    pd.Timestamp(podd_end),
                )
                st.session_state['filter_label'] = (
                    f"{podd_start.strftime('%d %b %Y')} – {podd_end.strftime('%d %b %Y')}"
                )
        else:
            st.warning("⚠️ Upload semua file yang diperlukan")


# ── Main area ─────────────────────────────────────────────────────────────────
if 'results' in st.session_state:
    df = st.session_state['results']
    filter_label = st.session_state.get('filter_label', '')

    st.subheader(f"📊 Hasil Pengecekan  —  PODD: {filter_label}")

    if df.empty:
        st.info(
            "ℹ️ Tidak ada data yang memenuhi filter:\n"
            "- FCR Date kosong, DAN\n"
            f"- PODD dalam rentang {filter_label}"
        )
    else:
        # ── Ringkasan statistik ───────────────────────────────────────────
        status_counts = df['Status'].value_counts() if 'Status' in df.columns else pd.Series(dtype=int)

        n_match     = int(status_counts.get('match',     0))
        n_mismatch  = int(status_counts.get('mismatch',  0))
        n_not_found = int(status_counts.get('not_found', 0))
        n_invalid   = int(status_counts.get('invalid',   0))
        n_total     = len(df)

        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("Total Order",  n_total)
        c2.metric("✅ Match",     n_match)
        c3.metric("⚠️ Mismatch",  n_mismatch)
        c4.metric("❌ Not Found", n_not_found)
        c5.metric("⚠️ Invalid",   n_invalid)

        st.markdown("---")

        # ── Tab view: All / Not Found / Mismatch ─────────────────────────
        tab_all, tab_not_found, tab_mismatch, tab_match = st.tabs([
            f"Semua ({n_total})",
            f"❌ Not Found ({n_not_found})",
            f"⚠️ Mismatch ({n_mismatch})",
            f"✅ Match ({n_match})",
        ])

        with tab_all:
            st.dataframe(df, use_container_width=True)

        with tab_not_found:
            df_nf = df[df['Status'] == 'not_found'] if 'Status' in df.columns else df.iloc[0:0]
            if df_nf.empty:
                st.success("Tidak ada order yang Not Found 🎉")
            else:
                st.dataframe(df_nf, use_container_width=True)

        with tab_mismatch:
            df_mm = df[df['Status'] == 'mismatch'] if 'Status' in df.columns else df.iloc[0:0]
            if df_mm.empty:
                st.success("Tidak ada Date Mismatch 🎉")
            else:
                st.dataframe(df_mm, use_container_width=True)

        with tab_match:
            df_ok = df[df['Status'] == 'match'] if 'Status' in df.columns else df.iloc[0:0]
            if df_ok.empty:
                st.info("Belum ada order yang Match.")
            else:
                st.dataframe(df_ok, use_container_width=True)

        st.markdown("---")

        # ── Download ──────────────────────────────────────────────────────
        excel_bytes = build_excel(df)
        file_date   = datetime.now().strftime('%Y%m%d_%H%M')

        st.download_button(
            label="📥 Download Excel",
            data=excel_bytes,
            file_name=f"Loading Plan Check - {file_date}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
