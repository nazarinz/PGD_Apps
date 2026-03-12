"""
Invoice Comparison Tool — FMS vs FVB
Auto-detect FMS/FVB dari SO/PO No:
  - Diawali "82..." → FMS
  - Diawali "10..." → FVB

Run: streamlit run app.py
"""

import io
import re
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

# ── CONFIG ────────────────────────────────────────────────────────────────────

HEADER_FIELDS = [
    "Invoice No.", "Export Type", "Brand Name", "Buyer Name",
    "Consignee", "Ex-Go Date", "PEB No", "PEB Date", "Amount", "Gross Weight"
]
COMPARE_HEADER_FIELDS = ["Amount", "Gross Weight"]

ITEM_COLS = [
    "Item Seq.", "Fact No", "SO / PO No", "Mat No", "Mat Name",
    "Hscode No", "Qty", "Jenis Satuan", "Unit Price", "Amount",
    "FOC Mark", "Pack Qty", "Pack Name", "Net Weight", "Gross Weight", "Volume"
]
SKIP_ITEM_COMPARE = {"SO / PO No"}
COMPARE_ITEM_COLS = [c for c in ITEM_COLS if c not in SKIP_ITEM_COMPARE]

# Colors
C_FMS_COL   = "2471A3"
C_FVB_COL   = "D68910"
C_FMS_LIGHT = "D6E4F0"
C_FVB_LIGHT = "FFF3CD"
C_GREEN     = "D4EDDA"
C_RED       = "F8D7DA"
C_YELLOW    = "FFF9C4"
C_WHITE     = "FFFFFF"
C_SUMMARY   = "1A1A2E"

# ── DATA MODELS ───────────────────────────────────────────────────────────────

@dataclass
class InvoiceData:
    invoice_no: str
    file_type: str  # "FMS" or "FVB"
    header: dict[str, str] = field(default_factory=dict)
    items: list[dict[str, Any]] = field(default_factory=list)
    total_row: list[Any] | None = None
    source_file: str = ""

@dataclass
class Diff:
    invoice_no: str
    scope: str
    field: str
    item_seq: str
    fms_val: str
    fvb_val: str

# ── PARSER ────────────────────────────────────────────────────────────────────

def _norm(val: Any) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    # Remove .0 suffix from numeric strings
    if re.match(r'^\d+\.0$', s):
        return s[:-2]
    return s


def _detect_type(items: list[dict]) -> str:
    """
    Detect FMS or FVB from SO/PO No:
      - Starts with '82' → FMS
      - Starts with '10' → FVB
    """
    for item in items:
        po = str(item.get("SO / PO No", "")).strip()
        if po.startswith("82"):
            return "FMS"
        if po.startswith("10"):
            return "FVB"
    return "UNKNOWN"


def _build_header_lookup() -> dict[str, str]:
    lookup: dict[str, str] = {}
    for f in HEADER_FIELDS:
        key = f.strip().rstrip(":").rstrip(".").strip().lower()
        lookup[key] = f
    lookup["invoice no"]   = "Invoice No."
    lookup["gross weight"] = "Gross Weight"
    return lookup

_HEADER_LOOKUP = _build_header_lookup()


def _clean_key(cell: Any) -> str:
    """Normalize a cell value to a lookup key.
    Strips colons, dots, spaces in any order so 'Invoice No. :' → 'invoice no'
    """
    s = str(cell).strip()
    # Remove trailing colon + dot combinations e.g. " :", ". :", "."
    s = s.rstrip(": ").rstrip(".").rstrip(": ").rstrip(".").strip()
    return s.lower()


def _next_val(row: list, j: int) -> str:
    """Return first non-empty value to the right of position j."""
    for offset in range(1, len(row) - j):
        v = _norm(row[j + offset])
        if v:
            return v
    return ""


def parse_invoice(file_obj, filename: str) -> InvoiceData:
    df_raw = pd.read_excel(file_obj, header=None, dtype=str)
    rows = df_raw.fillna("").values.tolist()

    header: dict[str, str] = {}
    item_header_row = -1

    # ── Pass 1: scan every cell for header labels ─────────────────────────────
    for i, row in enumerate(rows):
        for j, cell in enumerate(row):
            key = _clean_key(cell)
            if key in _HEADER_LOOKUP:
                f = _HEADER_LOOKUP[key]
                if not header.get(f):           # keep first match
                    header[f] = _next_val(row, j)

        if any(str(c).strip() == "Item Seq." for c in row):
            item_header_row = i

    # ── Pass 2: dedicated hunt for Invoice No. if still missing ───────────────
    # Scan all cells; when we find a label-like cell containing "invoice no",
    # grab the nearest non-empty value in the same row OR the row below.
    if not header.get("Invoice No."):
        for i, row in enumerate(rows):
            for j, cell in enumerate(row):
                if "invoice no" in _clean_key(cell):
                    # Try same row first
                    val = _next_val(row, j)
                    # Try next row at same column if still empty
                    if not val and i + 1 < len(rows):
                        val = _norm(rows[i + 1][j]) or _next_val(rows[i + 1], j - 1)
                    if val:
                        header["Invoice No."] = val
                        break
            if header.get("Invoice No."):
                break

    # ── Parse item rows ───────────────────────────────────────────────────────
    items: list[dict[str, Any]] = []
    total_row = None

    if item_header_row >= 0:
        col_map = {str(rows[item_header_row][k]).strip(): k
                   for k in range(len(rows[item_header_row]))}
        for row in rows[item_header_row + 1:]:
            if all(str(c).strip() == "" for c in row):
                break
            seq = str(row[0]).strip() if row else ""
            if seq == "" and any(str(c).strip() != "" for c in row):
                total_row = row
                continue
            if seq and not seq.replace(".", "").isdigit():
                break
            item = {col: _norm(row[col_map[col]]) if col in col_map else ""
                    for col in ITEM_COLS}
            items.append(item)

    file_type = _detect_type(items)

    # Final invoice number — strip filename suffix if value came from stem
    raw_inv = header.get("Invoice No.", "").strip() or Path(filename).stem
    inv_no = re.sub(r'[\s_]+(FMS|FVB)$', '', raw_inv, flags=re.IGNORECASE).strip()
    header["Invoice No."] = inv_no   # ensure Excel output always shows it

    return InvoiceData(
        invoice_no=inv_no,
        file_type=file_type,
        header=header,
        items=items,
        total_row=total_row,
        source_file=filename
    )

# ── COMPARISON ────────────────────────────────────────────────────────────────

def compare_pair(fms: InvoiceData, fvb: InvoiceData) -> list[Diff]:
    diffs: list[Diff] = []
    inv = fms.invoice_no

    for f in COMPARE_HEADER_FIELDS:
        fv, bv = fms.header.get(f, ""), fvb.header.get(f, "")
        if fv != bv:
            diffs.append(Diff(inv, "header", f, "", fv, bv))

    for idx, fms_item in enumerate(fms.items):
        fvb_item = fvb.items[idx] if idx < len(fvb.items) else {}
        seq = fms_item.get("Item Seq.", str(idx + 1))
        for col in COMPARE_ITEM_COLS:
            fv = fms_item.get(col, "")
            bv = _norm(fvb_item.get(col, ""))
            if fv != bv:
                diffs.append(Diff(inv, "item", col, seq, fv, bv))

    if fms.total_row and fvb.total_row:
        for ci, (fv, bv) in enumerate(zip(fms.total_row, fvb.total_row)):
            if _norm(fv) != _norm(bv) and (_norm(fv) or _norm(bv)):
                col_name = ITEM_COLS[ci] if ci < len(ITEM_COLS) else f"Col{ci}"
                diffs.append(Diff(inv, "total", col_name, "TOTAL", _norm(fv), _norm(bv)))

    return diffs

# ── EXCEL BUILDER ─────────────────────────────────────────────────────────────

def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)

def _border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _font(bold=True, size=10, color="000000") -> Font:
    return Font(bold=bold, size=size, color=color)


def _write_header_block(ws, r: int, label: str, data: InvoiceData, lc: str, fc: str) -> int:
    ws.cell(r, 1, label).font = _font(color="FFFFFF", size=11)
    ws.cell(r, 1).fill = _fill(lc)
    ws.cell(r, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    for f in HEADER_FIELDS:
        r += 1
        ws.cell(r, 2, f"{f} :").font = _font()
        ws.cell(r, 2).fill = _fill(fc)
        ws.cell(r, 3, data.header.get(f, "")).fill = _fill(C_WHITE)
    return r + 1


def _write_items_block(ws, r: int, data: InvoiceData, col_color: str, diffs: list[Diff]) -> int:
    diff_keys  = {(d.item_seq, d.field) for d in diffs if d.scope == "item"}
    diff_total = {d.field for d in diffs if d.scope == "total"}

    for ci, col in enumerate(ITEM_COLS, 1):
        c = ws.cell(r, ci, col)
        c.font = _font(color="FFFFFF", size=9)
        c.fill = _fill(col_color)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = _border()
    r += 1

    for item in data.items:
        for ci, col in enumerate(ITEM_COLS, 1):
            c = ws.cell(r, ci, item.get(col, ""))
            c.border = _border()
            c.alignment = Alignment(horizontal="center")
            if (item.get("Item Seq.", ""), col) in diff_keys:
                c.fill = _fill(C_RED)
            elif col in SKIP_ITEM_COMPARE:
                c.fill = _fill(C_YELLOW)
        r += 1

    if data.total_row:
        for ci, val in enumerate(data.total_row[:len(ITEM_COLS)], 1):
            col_name = ITEM_COLS[ci - 1] if ci - 1 < len(ITEM_COLS) else ""
            c = ws.cell(r, ci, _norm(val))
            c.font = _font(size=9)
            c.border = _border()
            c.alignment = Alignment(horizontal="center")
            if col_name in diff_total:
                c.fill = _fill(C_RED)
        r += 1

    return r + 1


def _write_ok_row(ws, r: int, diffs: list[Diff]) -> None:
    # Item-level diffs (by column name)
    item_diff_cols   = {d.field for d in diffs if d.scope in ("item", "total")}
    # Header-level diffs (Amount / Gross Weight)
    header_diff_cols = {d.field for d in diffs if d.scope == "header"}

    ws.cell(r, 1, "CHECK").font = _font(color="FFFFFF", size=9)
    ws.cell(r, 1).fill = _fill("444444")
    ws.cell(r, 1).alignment = Alignment(horizontal="center")

    # Header compare result (Amount + Gross Weight) shown in the SO/PO No cell
    header_ok = not header_diff_cols  # True if Amount AND Gross Weight both match

    for ci, col in enumerate(ITEM_COLS, 1):
        if col in SKIP_ITEM_COMPARE:
            # Repurpose SO/PO No cell → shows header Amount & Gross Weight compare result
            lbl = "OK" if header_ok else "DIFF"
            clr = C_GREEN if header_ok else C_RED
        elif col in item_diff_cols:
            lbl, clr = "DIFF", C_RED
        else:
            lbl, clr = "OK", C_GREEN
        c = ws.cell(r, ci, lbl)
        c.font = _font(size=9)
        c.fill = _fill(clr)
        c.alignment = Alignment(horizontal="center")
        c.border = _border()


def build_excel(pairs: list[tuple[InvoiceData, InvoiceData, list[Diff]]]) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    summary_rows: list[tuple] = []
    col_widths = [8, 18, 36, 16, 30, 12, 8, 10, 10, 12, 8, 8, 8, 11, 12, 9]

    for fms, fvb, diffs in pairs:
        safe = fms.invoice_no.replace("/", "_").replace("\\", "_")[:31]
        ws = wb.create_sheet(safe)
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        r = 1
        r = _write_header_block(ws, r, "FMS", fms, C_FMS_COL, C_FMS_LIGHT)
        r = _write_items_block(ws, r, fms, C_FMS_COL, diffs)
        r = _write_header_block(ws, r, "FVB", fvb, C_FVB_COL, C_FVB_LIGHT)
        r = _write_items_block(ws, r, fvb, C_FVB_COL, diffs)
        _write_ok_row(ws, r, diffs)

        for d in diffs:
            summary_rows.append((d.invoice_no, d.scope, d.field,
                                  d.item_seq, d.fms_val, d.fvb_val))

    ws_s = wb.create_sheet("SUMMARY")
    hdrs = ["Invoice No.", "Scope", "Field", "Item Seq.", "FMS Value", "FVB Value"]
    widths = [22, 10, 20, 10, 28, 28]
    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        c = ws_s.cell(1, ci, h)
        c.font = _font(color="FFFFFF")
        c.fill = _fill(C_SUMMARY)
        c.alignment = Alignment(horizontal="center")
        ws_s.column_dimensions[get_column_letter(ci)].width = w

    if not summary_rows:
        ws_s.cell(2, 1, "✓ No differences found").font = Font(color="2E7D32", bold=True)
    else:
        for ri, row in enumerate(summary_rows, 2):
            for ci, val in enumerate(row, 1):
                c = ws_s.cell(ri, ci, val)
                c.border = _border()
                c.fill = _fill(C_RED)
                c.alignment = Alignment(horizontal="center")
    ws_s.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ── STREAMLIT UI ──────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Invoice Compare — FMS vs FVB",
    page_icon="📊",
    layout="wide"
)

st.markdown("""
<style>
[data-testid="stAppViewContainer"] { background: #0f1117; }
[data-testid="stHeader"] { background: transparent; }
.main-title { font-size:2rem; font-weight:800; color:#fff;
              font-family:monospace; margin-bottom:0; }
.sub-title  { font-size:0.8rem; color:#555; margin-bottom:1.5rem;
              font-family:monospace; letter-spacing:3px; }
.tag-fms    { background:#1e3a5f; color:#90caf9; border-radius:4px;
              padding:1px 8px; font-size:0.72rem; font-weight:700;
              font-family:monospace; margin-left:6px; }
.tag-fvb    { background:#7c4a00; color:#ffd54f; border-radius:4px;
              padding:1px 8px; font-size:0.72rem; font-weight:700;
              font-family:monospace; margin-left:6px; }
.tag-unk    { background:#444; color:#aaa; border-radius:4px;
              padding:1px 8px; font-size:0.72rem; font-weight:700;
              font-family:monospace; margin-left:6px; }
.badge-ok   { background:#14532d; color:#86efac; border-radius:4px;
              padding:2px 10px; font-size:0.72rem; font-weight:700; font-family:monospace; }
.badge-diff { background:#7f1d1d; color:#fca5a5; border-radius:4px;
              padding:2px 10px; font-size:0.72rem; font-weight:700; font-family:monospace; }
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="main-title">📊 Invoice Compare</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">FMS vs FVB — RECONCILIATION TOOL</div>', unsafe_allow_html=True)

# ── SINGLE UPLOAD ZONE ────────────────────────────────────────────────────────
st.markdown("#### Upload Files (FMS & FVB boleh campur, nama bebas)")
uploads = st.file_uploader(
    "Upload semua file sekaligus",
    type=["xlsx", "xls"],
    accept_multiple_files=True,
    label_visibility="collapsed"
)

# Parse & classify immediately after upload
parsed: list[InvoiceData] = []
parse_errors: list[str] = []

if uploads:
    for f in uploads:
        try:
            data = parse_invoice(f, f.name)
            parsed.append(data)
        except Exception as e:
            parse_errors.append(f"{f.name}: {e}")

    # Show file list with auto-detected type
    st.markdown("**File terdeteksi:**")
    cols = st.columns(min(len(parsed), 4))
    for i, d in enumerate(parsed):
        tag = d.file_type
        tag_cls = "tag-fms" if tag == "FMS" else ("tag-fvb" if tag == "FVB" else "tag-unk")
        cols[i % len(cols)].markdown(
            f"`{d.source_file}`<br>"
            f"<span class='{tag_cls}'>{tag}</span> "
            f"<small style='color:#666'>{d.invoice_no}</small>",
            unsafe_allow_html=True
        )

    for err in parse_errors:
        st.error(f"⚠ Gagal parse: {err}")

    # Debug expander — shows raw parsed header so user can diagnose issues
    with st.expander("🔍 Debug: lihat hasil parse (klik jika ada masalah)"):
        for d in parsed:
            st.markdown(f"**{d.source_file}** → Tipe: `{d.file_type}` | Invoice No: `{d.invoice_no}`")
            st.json(d.header, expanded=False)

st.divider()

if not uploads:
    st.info("⬆ Upload file Excel invoice (bisa sekaligus banyak, nama file bebas).")
    st.stop()

if not parsed:
    st.stop()

# Check we have both types
types_found = {d.file_type for d in parsed}
if "FMS" not in types_found:
    st.warning("⚠ Tidak ada file FMS terdeteksi. Pastikan ada invoice dengan SO/PO No berawalan '82...'")
if "FVB" not in types_found:
    st.warning("⚠ Tidak ada file FVB terdeteksi. Pastikan ada invoice dengan SO/PO No berawalan '10...'")
if "UNKNOWN" in types_found:
    unk = [d.source_file for d in parsed if d.file_type == "UNKNOWN"]
    st.warning(f"⚠ Tidak bisa deteksi tipe untuk: {', '.join(unk)}")

run_btn = st.button("▶ COMPARE", type="primary")

if run_btn or st.session_state.get("result"):

    if run_btn:
        fms_map: dict[str, InvoiceData] = {}
        fvb_map: dict[str, InvoiceData] = {}

        for d in parsed:
            if d.file_type == "FMS":
                fms_map[d.invoice_no] = d
            elif d.file_type == "FVB":
                fvb_map[d.invoice_no] = d

        all_keys = sorted(set(fms_map) | set(fvb_map))
        pairs: list[tuple] = []
        missing: list[dict] = []

        for key in all_keys:
            if key not in fms_map:
                missing.append({"invoice": key, "reason": "FMS tidak ditemukan"})
            elif key not in fvb_map:
                missing.append({"invoice": key, "reason": "FVB tidak ditemukan"})
            else:
                diffs = compare_pair(fms_map[key], fvb_map[key])
                pairs.append((fms_map[key], fvb_map[key], diffs))

        st.session_state["result"] = {"pairs": pairs, "missing": missing}

    res    = st.session_state.get("result", {})
    pairs  = res.get("pairs", [])
    missing = res.get("missing", [])

    if not pairs and not missing:
        st.warning("Tidak ada pasangan FMS+FVB yang cocok.")
        st.stop()

    # ── METRICS ───────────────────────────────────────────────────────────────
    total_diffs = sum(len(d) for _, _, d in pairs)
    ok_count    = sum(1 for _, _, d in pairs if not d)

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Invoice Cocok",   len(pairs))
    m2.metric("✅ Semua OK",     ok_count)
    m3.metric("❌ Ada Beda",     len(pairs) - ok_count)
    m4.metric("Total Perbedaan", total_diffs)

    st.divider()

    # ── DOWNLOAD ──────────────────────────────────────────────────────────────
    if pairs:
        excel_bytes = build_excel(pairs)
        st.download_button(
            label="⬇ Download Hasil Compare (.xlsx)",
            data=excel_bytes,
            file_name="Invoice_Compare_Result.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )

    st.divider()

    # ── RESULTS ───────────────────────────────────────────────────────────────
    st.markdown("**Hasil Per Invoice**")

    for fms, fvb, diffs in pairs:
        label = f"{'🔴' if diffs else '🟢'}  {fms.invoice_no}"
        label += f"  —  {len(diffs)} perbedaan" if diffs else "  —  semua sama"

        with st.expander(label):
            if not diffs:
                st.success("Tidak ada perbedaan.")
            else:
                header_diffs = [d for d in diffs if d.scope == "header"]
                item_diffs   = [d for d in diffs if d.scope in ("item", "total")]

                if header_diffs:
                    st.markdown("**Header**")
                    st.dataframe(pd.DataFrame([{
                        "Field": d.field, "FMS": d.fms_val, "FVB": d.fvb_val
                    } for d in header_diffs]), use_container_width=True, hide_index=True)

                if item_diffs:
                    st.markdown("**Items**")
                    st.dataframe(pd.DataFrame([{
                        "Item Seq.": d.item_seq, "Field": d.field,
                        "FMS": d.fms_val, "FVB": d.fvb_val
                    } for d in item_diffs]), use_container_width=True, hide_index=True)

    if missing:
        st.divider()
        st.markdown("**Invoice Tidak Cocok**")
        for m in missing:
            st.warning(f"⚠  **{m['invoice']}** — {m['reason']}")
