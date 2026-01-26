import streamlit as st
import pandas as pd
import io
from openpyxl.styles import Font, Alignment, PatternFill

st.set_page_config(page_title="SAP Material Splitter", page_icon="🧩", layout="wide")
st.title("🧩 SAP (4020 & 5030) Material Description Splitter + Auto Sheet Export")


# =========================
# FUNCTION: Split Material Description
# =========================
def split_material_description(df):
    split_1 = df['Material desctiption'].astype(str).str.split(' ', n=1, expand=True)

    df['Article No.'] = split_1[0]
    df['Gender'] = (
        split_1[1]
        .str.replace('adidas-', '', regex=False)
        .str.strip()
    )

    df = df[[
        'Article',
        'Material desctiption',
        'Article No.',
        'Gender',
        'Sales org.',
        'Distr. Chl',
        'Product Hierarchy',
        'Gen. item cat. grp'
    ]]

    return df


# =========================
# FUNCTION: Split to 4 Sheets
# =========================
def split_to_sheets(df):
    sheets = {
        "4020_TT": df[(df['Sales org.'] == 4020) & (df['Distr. Chl'] == 'TT')],
        "5030_10": df[(df['Sales org.'] == 5030) & (df['Distr. Chl'] == '10')],
        "5030_TT": df[(df['Sales org.'] == 5030) & (df['Distr. Chl'] == 'TT')],
        "4020_10": df[(df['Sales org.'] == 4020) & (df['Distr. Chl'] == '10')],
    }
    return sheets


# =========================
# FUNCTION: Excel Styling
# =========================
def style_worksheet(ws):
    calibri9 = Font(name='Calibri', size=9, color='000000')
    header_font = Font(name='Calibri', size=9, bold=True, color='000000')
    wrap = Alignment(wrap_text=True, vertical='center')
    header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

    # Apply font ke semua cell
    for row in ws.iter_rows():
        for cell in row:
            cell.font = calibri9

    # Styling header (row 1)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = wrap
        cell.fill = header_fill

    # Freeze header
    ws.freeze_panes = "A2"


# =========================
# STREAMLIT UI
# =========================
uploaded_file = st.file_uploader("Upload file Excel SAP", type=["xlsx"])

if uploaded_file:
    df = pd.read_excel(uploaded_file)

    st.subheader("Preview Data Asli")
    st.dataframe(df.head(), use_container_width=True)

    # Step 1: Split material description
    df_processed = split_material_description(df)

    st.subheader("Preview Setelah Split")
    st.dataframe(df_processed.head(), use_container_width=True)

    # Step 2: Split ke kombinasi sheet
    sheet_dict = split_to_sheets(df_processed)

    # Step 3: Export Excel multi-sheet dengan styling
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet Original
        df_processed.to_excel(writer, index=False, sheet_name='Original')
        style_worksheet(writer.sheets['Original'])

        # Sheet lainnya
        for name, data in sheet_dict.items():
            data.to_excel(writer, index=False, sheet_name=name)
            style_worksheet(writer.sheets[name])

    st.download_button(
        label="⬇️ Download Excel (Styled 5 Sheets)",
        data=output.getvalue(),
        file_name="SAP_Material_Split_Styled.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
