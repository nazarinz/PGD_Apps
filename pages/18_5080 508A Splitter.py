import streamlit as st
import pandas as pd
import io
from openpyxl.styles import Font, Alignment, PatternFill

st.set_page_config(page_title="SAP Plant Splitter", page_icon="🏭", layout="wide")
st.title("🏭 SAP Material Extract + Article Main + Plant Split (5080 / 508A)")


# =========================
# FUNCTION: Extract Article No & Gender
# =========================
def extract_material_info(df):
    # Buang size info setelah koma
    base_text = df['Material description'].astype(str).str.split(',', n=1).str[0]

    split_1 = base_text.str.split(' ', n=1, expand=True)

    df['Article No.'] = split_1[0]

    df['Gender'] = (
        split_1[1]
        .str.replace('adidas-', '', regex=False)
        .str.strip()
    )

    return df


# =========================
# FUNCTION: Create Article Main (15 digit)
# =========================
def create_article_main(df):
    df['Article'] = df['Article'].astype(str)
    df['Article main'] = df['Article'].str[:15]
    return df


# =========================
# FUNCTION: Split Sheets by Plant
# =========================
def split_by_plant(df):
    sheets = {
        "5080": df[df['Plant'] == '5080'],
        "508A": df[df['Plant'] == '508A'],
    }
    return sheets


# =========================
# FUNCTION: Excel Styling
# =========================
def style_worksheet(ws):
    calibri9 = Font(name='Calibri', size=9, color='000000')
    header_font = Font(name='Calibri', size=9, bold=True, color='000000')
    wrap = Alignment(wrap_text=True, vertical='center')
    header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

    for row in ws.iter_rows():
        for cell in row:
            cell.font = calibri9

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = wrap
        cell.fill = header_fill

    ws.freeze_panes = "A2"


# =========================
# STREAMLIT UI
# =========================
uploaded_file = st.file_uploader("Upload file Excel SAP (Plant)", type=["xlsx"])

if uploaded_file:
    df = pd.read_excel(uploaded_file)

    st.subheader("Preview Data Asli")
    st.dataframe(df.head(), use_container_width=True)

    # Step 1: Extract Article No & Gender
    df = extract_material_info(df)

    # Step 2: Create Article main
    df = create_article_main(df)

    st.subheader("Preview Setelah Transform")
    st.dataframe(df.head(), use_container_width=True)

    # Step 3: Split sheet by Plant
    sheet_dict = split_by_plant(df)

    # Step 4: Export Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Original')
        style_worksheet(writer.sheets['Original'])

        for name, data in sheet_dict.items():
            data.to_excel(writer, index=False, sheet_name=name)
            style_worksheet(writer.sheets[name])

    st.download_button(
        label="⬇️ Download Excel (Plant Split + Article Main)",
        data=output.getvalue(),
        file_name="SAP_Plant_Split_ArticleMain.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
