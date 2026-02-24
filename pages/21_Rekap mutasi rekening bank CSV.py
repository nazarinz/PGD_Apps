import streamlit as st
import pandas as pd
import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

st.set_page_config(page_title="Rekap Mutasi Rekening", page_icon="💳", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@300;400;500;600;700;800&family=JetBrains+Mono:wght@400;600&display=swap');
html, body, [class*="css"] { font-family: 'Plus Jakarta Sans', sans-serif; }
.stApp { background: linear-gradient(135deg, #0f172a 0%, #1e293b 50%, #0f172a 100%); min-height: 100vh; }
.block-container { padding-top: 1.5rem; padding-bottom: 2rem; max-width: 1200px; }

/* Hero */
.hero { text-align: center; padding: 2rem 2rem 1.5rem; margin-bottom: 1rem; }
.hero-badge { display: inline-block; background: linear-gradient(90deg, #3b82f6, #06b6d4); color: white; font-size: 0.68rem; font-weight: 700; letter-spacing: 0.15em; text-transform: uppercase; padding: 0.28rem 0.9rem; border-radius: 999px; margin-bottom: 1rem; }
.hero h1 { font-size: 2.4rem; font-weight: 800; color: #f1f5f9; margin: 0 0 0.5rem; letter-spacing: -0.02em; }
.hero h1 span { background: linear-gradient(90deg, #3b82f6, #06b6d4); -webkit-background-clip: text; -webkit-text-fill-color: transparent; background-clip: text; }
.hero p { color: #94a3b8; font-size: 0.95rem; margin: 0; }

/* Upload */
.upload-card { background: rgba(30,41,59,0.8); border: 1px solid rgba(59,130,246,0.25); border-radius: 14px; padding: 1.5rem 2rem; margin-bottom: 1.2rem; }

/* Filter bar */
.filter-bar { background: rgba(15,23,42,0.7); border: 1px solid rgba(255,255,255,0.08); border-radius: 12px; padding: 1rem 1.5rem; margin-bottom: 1.2rem; }
.filter-label { font-size: 0.7rem; font-weight: 700; letter-spacing: 0.1em; text-transform: uppercase; color: #475569; margin-bottom: 0.3rem; }

/* Active filter chip */
.chip { display:inline-block; background:rgba(59,130,246,0.18); color:#93c5fd; border:1px solid rgba(59,130,246,0.35); border-radius:999px; font-size:0.72rem; font-weight:600; padding:0.18rem 0.75rem; margin:0.15rem; }

/* Metrics */
.metric-grid { display: grid; grid-template-columns: repeat(4,1fr); gap: 0.85rem; margin-bottom: 1.2rem; }
.metric-card { background: rgba(30,41,59,0.9); border-radius: 12px; padding: 1.1rem 1.3rem; border: 1px solid rgba(255,255,255,0.07); position: relative; overflow: hidden; }
.metric-card::before { content:''; position:absolute; top:0;left:0;right:0;height:3px; }
.metric-card.saldo-awal::before  { background:#475569; }
.metric-card.kredit::before      { background:linear-gradient(90deg,#22c55e,#16a34a); }
.metric-card.debet::before       { background:linear-gradient(90deg,#ef4444,#dc2626); }
.metric-card.saldo-akhir::before { background:linear-gradient(90deg,#3b82f6,#06b6d4); }
.metric-label { font-size:0.68rem; font-weight:700; letter-spacing:0.1em; text-transform:uppercase; color:#64748b; margin-bottom:0.35rem; }
.metric-value { font-family:'JetBrains Mono',monospace; font-size:0.98rem; font-weight:700; color:#f1f5f9; word-break:break-all; }
.metric-value.kredit      { color:#4ade80; }
.metric-value.debet       { color:#f87171; }
.metric-value.saldo-akhir { color:#60a5fa; }
.metric-sub { font-size:0.68rem; color:#475569; margin-top:0.2rem; }
.metric-filtered { font-size:0.65rem; color:#3b82f6; margin-top:0.15rem; font-weight:600; }

/* Day header */
.day-header { font-size:0.72rem; font-weight:700; letter-spacing:0.1em; text-transform:uppercase; color:#3b82f6; border-bottom:1px solid rgba(59,130,246,0.2); padding-bottom:0.35rem; margin:1.2rem 0 0.7rem; }

/* Buttons */
.stDownloadButton > button { background:linear-gradient(135deg,#2563eb,#0891b2) !important; color:white !important; border:none !important; border-radius:10px !important; font-weight:700 !important; font-size:0.9rem !important; padding:0.65rem 2rem !important; width:100% !important; transition:opacity 0.2s !important; }
.stDownloadButton > button:hover { opacity:0.85 !important; }

/* File uploader */
[data-testid="stFileUploader"] { background:rgba(15,23,42,0.5); border:2px dashed rgba(59,130,246,0.35); border-radius:12px; padding:0.8rem; }

/* Sidebar */
section[data-testid="stSidebar"] { background:rgba(15,23,42,0.98) !important; border-right:1px solid rgba(255,255,255,0.07) !important; }
section[data-testid="stSidebar"] .block-container { padding:1.5rem 1rem; max-width:100%; }

/* Streamlit multiselect tags */
[data-baseweb="tag"] { background:rgba(59,130,246,0.25) !important; border:1px solid rgba(59,130,246,0.4) !important; }
[data-baseweb="tag"] span { color:#93c5fd !important; }

.section-title { font-size:0.9rem; font-weight:700; color:#e2e8f0; margin-bottom:0.6rem; }
.info-box { background:rgba(59,130,246,0.08); border:1px solid rgba(59,130,246,0.2); border-radius:8px; padding:0.7rem 1rem; font-size:0.8rem; color:#93c5fd; margin-bottom:0.8rem; }
.footer { text-align:center; color:#334155; font-size:0.72rem; margin-top:2rem; padding-bottom:1rem; }
</style>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════
# PARSING
# ═══════════════════════════════════════════════

def parse_date_dd_mm_yyyy(keterangan: str) -> str:
    """
    Selalu output DD/MM/YYYY.
    BCA kode format:
      - DDMM/FTSCY  → misal 2302/FTSCY = tgl 23 bln 02
      - CR DDMM/    → misal CR 2302/ = tgl 23 bln 02
      - CR MM/DD    → misal CR 02/23 = tgl 23 bln 02 (BCA pakai MM/DD di sini)
      - TANGGAL:DD/MM  → BI-FAST, sudah DD/MM
    """
    k = keterangan.strip()

    # 1. DDMM/FTSCY  (paling reliable)
    m = re.search(r'(\d{2})(\d{2})/FTSCY', k)
    if m:
        dd, mm = m.group(1), m.group(2)
        return f"{dd}/{mm}/2026"

    # 2. CR/DB diikuti 4 digit tanpa slash (DDMM)
    m = re.search(r'(?:CR|DB)\s+(\d{2})(\d{2})/', k)
    if m:
        dd, mm = m.group(1), m.group(2)
        return f"{dd}/{mm}/2026"

    # 3. CR/DB diikuti MM/DD dengan spasi (format terbalik BCA)
    m = re.search(r'(?:CR|DB)\s+(\d{2})/(\d{2})\s', k)
    if m:
        # group1=MM, group2=DD  →  swap jadi DD/MM
        mm, dd = m.group(1), m.group(2)
        return f"{dd}/{mm}/2026"

    # 4. BI-FAST: TANGGAL :DD/MM
    m = re.search(r'TANGGAL\s*:(\d{2})/(\d{2})', k)
    if m:
        dd, mm = m.group(1), m.group(2)
        return f"{dd}/{mm}/2026"

    # 5. Fallback: ambil 4 digit pertama yang ada slash setelahnya
    m = re.search(r'(\d{4})/', k)
    if m:
        c = m.group(1)
        return f"{c[:2]}/{c[2:]}/2026"

    return "N/A"


def parse_name(keterangan: str, tipe: str) -> str:
    k = keterangan.strip()

    # DEBET → ambil nomor rekening tujuan (angka panjang di akhir string)
    if tipe == 'DB':
        m = re.search(r'(\d{8,})\s*$', k)
        if m:
            return m.group(1)
        # Cari setelah pola COD/AREA
        m2 = re.search(r'(?:AREA\d+\s+COD|/)\s+[-\s]+(\d+)\s*$', k)
        if m2:
            return m2.group(1)
        return k[:40]

    # BI-FAST CR
    m = re.search(r'TANGGAL\s*:\S+\s+TRANSFER\s+DR\s+\d+\s+(.+)', k, re.I)
    if m: return m.group(1).strip()
    m = re.search(r'TRANSFER\s+DR\s+\d+\s+(.+)', k, re.I)
    if m: return m.group(1).strip()

    # ESPAY / BI-FAST FinTech
    m = re.search(r'TRFDN-([A-Z][A-Z\s]+?)(?:\s{2,}|ESPAY|$)', k, re.I)
    if m: return m.group(1).strip()

    # GoPay
    m = re.search(r'GoPay Bank Transfe\S*\s+(.+)', k, re.I)
    if m: return m.group(1).strip()

    # AIRPAY
    if re.search(r'AIRPAY', k, re.I): return 'AIRPAY INTERNATIONAL'

    # E-Banking biasa: ambil nama CAPSLOCK setelah double-space (nama bersih di akhir)
    m = re.search(r'\s{2,}([A-Z][A-Z\s\.\',\-]+?)\s*$', k)
    if m:
        candidate = m.group(1).strip()
        if not re.match(r'^\d+$', candidate) and len(candidate) > 2:
            return candidate

    # Fallback: strip prefix angka.angka dari keterangan
    clean = re.sub(r'^\d+(?:\.\d+)?', '', k).strip()
    parts = re.split(r'\s{2,}', clean)
    if len(parts) >= 2:
        last = parts[-1].strip()
        if last and not re.match(r'^\d+$', last):
            return last
    clean2 = re.sub(r'^\d+(?:\.\d+)?', '', clean).strip()
    return clean2[:40] if clean2 else k[:40]


def get_jenis(keterangan: str) -> str:
    k = keterangan.upper()
    if 'BI-FAST' in k: return 'BI-FAST'
    if any(x in k for x in ['ESPAY','DOMPET ANAK BANGSA','AIRPAY','GOPAY']): return 'E-Banking (FinTech)'
    if 'E-BANKING' in k: return 'E-Banking'
    return 'Transfer'


def parse_csv(content: str):
    lines = [l.strip() for l in content.splitlines() if l.strip()]
    summary = {}
    transactions = []
    seq = 0  # original sequence number

    for line in lines:
        low = line.lower()
        if low.startswith('saldo awal'):
            try: summary['saldo_awal'] = float(line.split(',')[-1].replace("'","").strip())
            except: pass
        elif low.startswith('kredit,'):
            try: summary['kredit'] = float(line.split(',')[-1].replace("'","").strip())
            except: pass
        elif low.startswith('debet,'):
            try: summary['debet'] = float(line.split(',')[-1].replace("'","").strip())
            except: pass
        elif low.startswith('saldo akhir'):
            try: summary['saldo_akhir'] = float(line.split(',')[-1].replace("'","").strip())
            except: pass
        elif low.startswith("'pend") or low.startswith("pend"):
            seq += 1
            clean = line.lstrip("'")
            parts = clean.split(',')
            if len(parts) < 5: continue
            keterangan = parts[1]
            try:
                saldo  = float(parts[-1])
                tipe   = parts[-2].strip()
                jumlah = float(parts[-3])
                date_str = parse_date_dd_mm_yyyy(keterangan)
                nama   = parse_name(keterangan, tipe)
                jenis  = get_jenis(keterangan)
                transactions.append({
                    'No_Asli': seq,
                    'Tanggal': date_str,
                    'Jenis': jenis,
                    'Pengirim/Penerima': nama,
                    'Tipe': tipe,
                    'Jumlah': jumlah,
                    'Saldo': saldo,
                })
            except (ValueError, IndexError):
                continue

    return transactions, summary


# ═══════════════════════════════════════════════
# EXCEL
# ═══════════════════════════════════════════════

DARK  = "1F3864"
MED   = "2E75B6"
LGREY = "F7F9FC"
GREEN = "E2EFDA"
RED   = "FCE4D6"
LBLUE = "DBEAFE"
WHITE = "FFFFFF"

def _bdr():
    t = Side(style='thin', color='D1D5DB')
    return Border(left=t, right=t, top=t, bottom=t)

def _title(ws, row, ncols, text, subtitle=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
    c.fill = PatternFill('solid', start_color=DARK)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 28
    if subtitle:
        ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=ncols)
        c2 = ws.cell(row=row+1, column=1, value=subtitle)
        c2.font = Font(name='Calibri', italic=True, size=9, color='FFFFFF')
        c2.fill = PatternFill('solid', start_color=MED)
        c2.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row+1].height = 16

def _hdr(ws, row, headers):
    bdr = _bdr()
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
        c.fill = PatternFill('solid', start_color=MED)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = bdr
    ws.row_dimensions[row].height = 22

def _tx_sheet(ws, rows, filename, label, filter_info="", tipe_filter=None):
    bdr = _bdr()
    df_ = Font(name='Calibri', size=9)
    subtitle = f'File: {filename}   |   Dibuat: {datetime.now().strftime("%d %b %Y %H:%M")}'
    if filter_info:
        subtitle += f'   |   Filter: {filter_info}'
    _title(ws, 1, 7, f'REKAP MUTASI REKENING — {label}', subtitle)
    _hdr(ws, 3, ['No', 'Tanggal', 'Jenis', 'Pengirim / Penerima', 'Debet (Rp)', 'Kredit (Rp)', 'Saldo (Rp)'])
    ws.freeze_panes = 'A4'
    filtered = [r for r in rows if tipe_filter is None or r['Tipe'] == tipe_filter]
    for i, row in enumerate(filtered, 1):
        r = i + 3
        is_db  = row['Tipe'] == 'DB'
        debet  = row['Jumlah'] if is_db else None
        kredit = row['Jumlah'] if not is_db else None
        bg = LGREY if i % 2 == 0 else WHITE
        vals = [row.get('No_Asli', i), row['Tanggal'], row['Jenis'], row['Pengirim/Penerima'], debet, kredit, row['Saldo']]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = df_; cell.border = bdr
            if ci == 5 and val is not None:
                cell.fill = PatternFill('solid', start_color=RED)
                cell.number_format = '#,##0'; cell.alignment = Alignment(horizontal='right')
            elif ci == 6 and val is not None:
                cell.fill = PatternFill('solid', start_color=GREEN)
                cell.number_format = '#,##0'; cell.alignment = Alignment(horizontal='right')
            elif ci == 7:
                cell.fill = PatternFill('solid', start_color=bg)
                cell.number_format = '#,##0'; cell.alignment = Alignment(horizontal='right')
            elif ci == 1:
                cell.fill = PatternFill('solid', start_color=bg)
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.fill = PatternFill('solid', start_color=bg)
    for col, w in zip('ABCDEFG', [6, 12, 20, 30, 16, 16, 18]):
        ws.column_dimensions[col].width = w
    return len(filtered)

def _ringkasan_sheet(ws, transactions, summary, filename, filter_info=""):
    bdr = _bdr()
    df_ = Font(name='Calibri', size=9)
    bf_ = Font(name='Calibri', bold=True, size=9)
    hf_ = Font(name='Calibri', bold=True, color='FFFFFF', size=10)

    subtitle = f'Dibuat dari: {filename}'
    if filter_info: subtitle += f'   |   Filter: {filter_info}'
    _title(ws, 1, 4, 'RINGKASAN MUTASI REKENING', subtitle)

    def kv(row, label, value, bg=WHITE, bold=False):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = bf_ if bold else df_; lc.fill = PatternFill('solid', start_color=bg); lc.border = bdr
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        vc = ws.cell(row=row, column=3, value=value)
        vc.font = bf_ if bold else df_; vc.fill = PatternFill('solid', start_color=bg); vc.border = bdr
        vc.number_format = '#,##0.00'; vc.alignment = Alignment(horizontal='right')

    def sec(row, text):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c = ws.cell(row=row, column=1, value=text)
        c.font = hf_; c.fill = PatternFill('solid', start_color=DARK)
        c.alignment = Alignment(horizontal='center'); ws.row_dimensions[row].height = 20

    cr_total = sum(t['Jumlah'] for t in transactions if t['Tipe']=='CR')
    db_total = sum(t['Jumlah'] for t in transactions if t['Tipe']=='DB')

    sec(4, 'IKHTISAR SALDO (DATA TERFILTER)')
    kv(5, 'Total Kredit (Masuk)',  cr_total, bg=GREEN)
    kv(6, 'Total Debet (Keluar)', db_total, bg=RED)
    kv(7, 'Selisih (CR - DB)',    cr_total - db_total, bg=LBLUE, bold=True)

    # Saldo dari summary asli
    kv(8, 'Saldo Awal (asli file)',  summary.get('saldo_awal', 0), bg=LGREY)
    kv(9, 'Saldo Akhir (asli file)', summary.get('saldo_akhir', 0), bg=LGREY)

    dates = sorted(set(r['Tanggal'] for r in transactions))
    sec(11, 'RINGKASAN PER HARI')
    _hdr(ws, 12, ['Tanggal', 'Tx Masuk', 'Total Masuk (Rp)', 'Total Keluar (Rp)'])
    for ri, d in enumerate(dates):
        r = 13 + ri
        cr_amt = sum(t['Jumlah'] for t in transactions if t['Tanggal']==d and t['Tipe']=='CR')
        db_amt = sum(t['Jumlah'] for t in transactions if t['Tanggal']==d and t['Tipe']=='DB')
        cr_cnt = sum(1 for t in transactions if t['Tanggal']==d and t['Tipe']=='CR')
        bg = LGREY if ri % 2 == 0 else WHITE
        for ci, val in enumerate([d, cr_cnt, cr_amt, db_amt], 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.font = df_; c.border = bdr; c.fill = PatternFill('solid', start_color=bg)
            if ci in (3,4): c.number_format = '#,##0'; c.alignment = Alignment(horizontal='right')
            elif ci==2: c.alignment = Alignment(horizontal='center')

    top_cr = sorted([(t['Jumlah'], t['Pengirim/Penerima'], t['Tanggal'])
                     for t in transactions if t['Tipe']=='CR'], reverse=True)[:10]
    sr = 13 + len(dates) + 2
    sec(sr, 'TOP 10 TRANSAKSI MASUK TERBESAR')
    _hdr(ws, sr+1, ['No','Pengirim','Tanggal','Jumlah (Rp)'])
    for ri, (amt, nm, tgl) in enumerate(top_cr, 1):
        r = sr + 1 + ri
        bg = GREEN if ri % 2 == 0 else WHITE
        for ci, val in enumerate([ri, nm, tgl, amt], 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.font = df_; c.border = bdr
            c.fill = PatternFill('solid', start_color=GREEN if ci==4 else bg)
            if ci==4: c.number_format='#,##0'; c.alignment=Alignment(horizontal='right')
            elif ci==1: c.alignment=Alignment(horizontal='center')

    for col, w in zip('ABCD', [22, 14, 18, 20]):
        ws.column_dimensions[col].width = w


def generate_excel(transactions, summary, filename, filter_info=""):
    wb = Workbook()

    ws_all = wb.active
    ws_all.title = "Semua Transaksi"
    ws_all.sheet_properties.tabColor = "64748b"
    n_all = _tx_sheet(ws_all, transactions, filename, "SEMUA TRANSAKSI", filter_info)

    ws_cr = wb.create_sheet("Kredit (Masuk)")
    ws_cr.sheet_properties.tabColor = "22c55e"
    n_cr = _tx_sheet(ws_cr, transactions, filename, "KREDIT — TRANSAKSI MASUK", filter_info, tipe_filter='CR')

    ws_db = wb.create_sheet("Debet (Keluar)")
    ws_db.sheet_properties.tabColor = "ef4444"
    n_db = _tx_sheet(ws_db, transactions, filename, "DEBET — TRANSAKSI KELUAR", filter_info, tipe_filter='DB')

    ws_r = wb.create_sheet("Ringkasan")
    ws_r.sheet_properties.tabColor = "3b82f6"
    _ringkasan_sheet(ws_r, transactions, summary, filename, filter_info)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), n_cr, n_db


# ═══════════════════════════════════════════════
# UI HELPERS
# ═══════════════════════════════════════════════

def fmt_rp(val):
    return f"Rp {val:,.0f}".replace(",",".")


def render_preview(df_in: pd.DataFrame, tipe=None):
    data = df_in.copy()
    if tipe:
        data = data[data['Tipe']==tipe]
    if data.empty:
        st.info("Tidak ada data untuk ditampilkan.")
        return
    dates = sorted(data['Tanggal'].unique())
    for d in dates:
        sub = data[data['Tanggal']==d]
        cr_d = sub[sub['Tipe']=='CR']['Jumlah'].sum()
        db_d = sub[sub['Tipe']=='DB']['Jumlah'].sum()
        if tipe == 'CR':
            label = f"📅 {d}  ·  Total Masuk: {fmt_rp(cr_d)}  ({len(sub)} tx)"
        elif tipe == 'DB':
            label = f"📅 {d}  ·  Total Keluar: {fmt_rp(db_d)}  ({len(sub)} tx)"
        else:
            label = f"📅 {d}  ·  Masuk: {fmt_rp(cr_d)}  ·  Keluar: {fmt_rp(db_d)}"
        st.markdown(f'<div class="day-header">{label}</div>', unsafe_allow_html=True)
        preview = sub[['No_Asli','Tanggal','Tipe','Pengirim/Penerima','Jenis','Jumlah','Saldo']].copy()
        preview['Tipe'] = preview['Tipe'].map({'CR':'✅ Kredit','DB':'🔴 Debet'})
        preview.columns = ['No','Tanggal','Tipe','Pengirim/Penerima','Jenis','Jumlah (Rp)','Saldo (Rp)']
        preview = preview.reset_index(drop=True)
        st.dataframe(
            preview.style.format({'Jumlah (Rp)':'{:,.0f}','Saldo (Rp)':'{:,.0f}'}),
            use_container_width=True,
            height=min(450, 38*(len(preview)+1))
        )


# ═══════════════════════════════════════════════
# MAIN UI
# ═══════════════════════════════════════════════

st.markdown("""
<div class="hero">
    <div class="hero-badge">💳 Bank Mutation Tools</div>
    <h1>Rekap Mutasi <span>Rekening</span></h1>
    <p>Upload CSV → filter → download Excel 4 sheet: Semua · Kredit · Debet · Ringkasan</p>
</div>
""", unsafe_allow_html=True)

# ── Upload ──
st.markdown('<div class="upload-card">', unsafe_allow_html=True)
st.markdown('<div class="section-title">📂 Upload File CSV Mutasi</div>', unsafe_allow_html=True)
uploaded = st.file_uploader("Upload", type=['csv'], label_visibility='collapsed')
st.markdown('<p style="color:#475569;font-size:0.78rem;margin-top:0.4rem;">Format: Mutasi rekening BCA (e-banking CSV export). Tanggal otomatis dikonversi ke DD/MM/YYYY.</p>', unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)

if not uploaded:
    st.markdown("""
    <div style="text-align:center;padding:3rem 1rem;">
        <div style="font-size:3rem;margin-bottom:0.8rem;">☁️</div>
        <div style="font-size:0.9rem;font-weight:600;color:#64748b;">Belum ada file yang diupload</div>
        <div style="font-size:0.78rem;margin-top:0.35rem;color:#475569;">Upload file CSV mutasi rekening untuk memulai</div>
    </div>""", unsafe_allow_html=True)
    st.stop()

# ── Parse ──
raw   = uploaded.read().decode('utf-8', errors='replace')
fname = uploaded.name

with st.spinner("Memproses data mutasi..."):
    transactions, summary = parse_csv(raw)

if not transactions:
    st.error("⚠️ Tidak ada transaksi yang berhasil diparsing. Pastikan format CSV sesuai.")
    st.stop()

df_raw = pd.DataFrame(transactions)

# ═══════════════════════════════════════════════
# SIDEBAR FILTER
# ═══════════════════════════════════════════════
with st.sidebar:
    st.markdown("""
    <div style="text-align:center;padding:0.5rem 0 1.2rem;">
        <div style="font-size:1.1rem;font-weight:800;color:#f1f5f9;">🔍 Filter</div>
        <div style="font-size:0.7rem;color:#475569;margin-top:0.2rem;">Semua filter berlaku pada preview & Excel</div>
    </div>
    """, unsafe_allow_html=True)

    # 1. Filter Tanggal
    all_dates = sorted(df_raw['Tanggal'].unique())
    st.markdown('<div style="font-size:0.7rem;font-weight:700;letter-spacing:0.08em;text-transform:uppercase;color:#64748b;margin-bottom:0.3rem;">📅 Tanggal</div>', unsafe_allow_html=True)
    sel_dates = st.multiselect(
        "Tanggal",
        options=all_dates,
        default=all_dates,
        label_visibility='collapsed'
    )

    st.markdown('<hr style="border-color:rgba(255,255,255,0.07);margin:0.8rem 0">', unsafe_allow_html=True)

    # 2. Filter Tipe
    st.markdown('<div style="font-size:0.7rem;font-weight:700;letter-spacing:0.08em;text-transform:uppercase;color:#64748b;margin-bottom:0.3rem;">💱 Tipe Transaksi</div>', unsafe_allow_html=True)
    sel_tipe = st.radio(
        "Tipe",
        options=["Semua", "Kredit (Masuk)", "Debet (Keluar)"],
        index=0,
        label_visibility='collapsed'
    )
    tipe_map = {"Semua": None, "Kredit (Masuk)": "CR", "Debet (Keluar)": "DB"}
    tipe_filter = tipe_map[sel_tipe]

    st.markdown('<hr style="border-color:rgba(255,255,255,0.07);margin:0.8rem 0">', unsafe_allow_html=True)

    # 3. Filter Jenis
    all_jenis = sorted(df_raw['Jenis'].unique())
    st.markdown('<div style="font-size:0.7rem;font-weight:700;letter-spacing:0.08em;text-transform:uppercase;color:#64748b;margin-bottom:0.3rem;">🏦 Jenis Transfer</div>', unsafe_allow_html=True)
    sel_jenis = st.multiselect(
        "Jenis",
        options=all_jenis,
        default=all_jenis,
        label_visibility='collapsed'
    )

    st.markdown('<hr style="border-color:rgba(255,255,255,0.07);margin:0.8rem 0">', unsafe_allow_html=True)

    # 4. Cari nama
    st.markdown('<div style="font-size:0.7rem;font-weight:700;letter-spacing:0.08em;text-transform:uppercase;color:#64748b;margin-bottom:0.3rem;">🔎 Cari Pengirim/Penerima</div>', unsafe_allow_html=True)
    search_name = st.text_input("Cari nama", placeholder="Ketik nama...", label_visibility='collapsed')

    st.markdown('<hr style="border-color:rgba(255,255,255,0.07);margin:0.8rem 0">', unsafe_allow_html=True)

    # 5. Filter nominal
    max_jumlah = int(df_raw['Jumlah'].max())
    min_jumlah = int(df_raw['Jumlah'].min())
    st.markdown('<div style="font-size:0.7rem;font-weight:700;letter-spacing:0.08em;text-transform:uppercase;color:#64748b;margin-bottom:0.3rem;">💰 Nominal (Rp)</div>', unsafe_allow_html=True)
    jumlah_range = st.slider(
        "Nominal",
        min_value=min_jumlah,
        max_value=max_jumlah,
        value=(min_jumlah, max_jumlah),
        label_visibility='collapsed'
    )

    st.markdown('<hr style="border-color:rgba(255,255,255,0.07);margin:0.8rem 0">', unsafe_allow_html=True)

    # Reset button
    if st.button("↺  Reset Semua Filter", use_container_width=True):
        st.rerun()

    # Filter summary
    st.markdown(f"""
    <div style="background:rgba(59,130,246,0.08);border:1px solid rgba(59,130,246,0.2);border-radius:8px;padding:0.7rem;margin-top:0.5rem;font-size:0.75rem;color:#93c5fd;">
        <b>Total data:</b> {len(df_raw)} transaksi<br>
        <b>Tanggal dipilih:</b> {len(sel_dates)} dari {len(all_dates)}
    </div>
    """, unsafe_allow_html=True)


# ═══════════════════════════════════════════════
# APPLY FILTERS
# ═══════════════════════════════════════════════
df = df_raw.copy()

if sel_dates:
    df = df[df['Tanggal'].isin(sel_dates)]
if tipe_filter:
    df = df[df['Tipe'] == tipe_filter]
if sel_jenis:
    df = df[df['Jenis'].isin(sel_jenis)]
if search_name:
    df = df[df['Pengirim/Penerima'].str.contains(search_name, case=False, na=False)]
df = df[(df['Jumlah'] >= jumlah_range[0]) & (df['Jumlah'] <= jumlah_range[1])]

is_filtered = len(df) < len(df_raw)
filtered_tx = df.to_dict('records')

# ── Build filter info string for Excel ──
filter_parts = []
if len(sel_dates) < len(all_dates):
    filter_parts.append(f"Tgl: {', '.join(sel_dates)}")
if sel_tipe != "Semua":
    filter_parts.append(f"Tipe: {sel_tipe}")
if len(sel_jenis) < len(all_jenis):
    filter_parts.append(f"Jenis: {', '.join(sel_jenis)}")
if search_name:
    filter_parts.append(f"Nama: '{search_name}'")
if jumlah_range != (min_jumlah, max_jumlah):
    filter_parts.append(f"Nominal: {fmt_rp(jumlah_range[0])} – {fmt_rp(jumlah_range[1])}")
filter_info = " | ".join(filter_parts)

# ═══════════════════════════════════════════════
# METRICS (ikut filter)
# ═══════════════════════════════════════════════
cr_total_raw = summary.get('kredit', df_raw[df_raw['Tipe']=='CR']['Jumlah'].sum())
db_total_raw = summary.get('debet',  df_raw[df_raw['Tipe']=='DB']['Jumlah'].sum())
sa = summary.get('saldo_awal', 0)
se = summary.get('saldo_akhir', 0)

cr_total = df[df['Tipe']=='CR']['Jumlah'].sum()
db_total = df[df['Tipe']=='DB']['Jumlah'].sum()
cr_cnt   = int((df['Tipe']=='CR').sum())
db_cnt   = int((df['Tipe']=='DB').sum())

# Tambahkan badge "terfilter" jika filter aktif
filtered_badge = ' <span style="font-size:0.6rem;background:rgba(59,130,246,0.25);color:#93c5fd;border-radius:999px;padding:0.1rem 0.5rem;vertical-align:middle;">terfilter</span>' if is_filtered else ''

st.markdown(f"""
<div class="metric-grid">
    <div class="metric-card saldo-awal">
        <div class="metric-label">Saldo Awal</div>
        <div class="metric-value">{fmt_rp(sa)}</div>
        <div class="metric-sub">Dari file asli</div>
    </div>
    <div class="metric-card kredit">
        <div class="metric-label">Total Masuk{filtered_badge}</div>
        <div class="metric-value kredit">{fmt_rp(cr_total)}</div>
        <div class="metric-sub">{cr_cnt} transaksi kredit</div>
        {"" if not is_filtered else f'<div class="metric-filtered">Asli: {fmt_rp(cr_total_raw)}</div>'}
    </div>
    <div class="metric-card debet">
        <div class="metric-label">Total Keluar{filtered_badge}</div>
        <div class="metric-value debet">{fmt_rp(db_total)}</div>
        <div class="metric-sub">{db_cnt} transaksi debet</div>
        {"" if not is_filtered else f'<div class="metric-filtered">Asli: {fmt_rp(db_total_raw)}</div>'}
    </div>
    <div class="metric-card saldo-akhir">
        <div class="metric-label">Saldo Akhir</div>
        <div class="metric-value saldo-akhir">{fmt_rp(se)}</div>
        <div class="metric-sub">Dari file asli</div>
    </div>
</div>
""", unsafe_allow_html=True)

# ── Active filter chips ──
if is_filtered:
    chips_html = ''.join([f'<span class="chip">{p}</span>' for p in filter_parts])
    st.markdown(f"""
    <div style="margin-bottom:1rem;">
        <span style="font-size:0.7rem;color:#64748b;font-weight:600;margin-right:0.5rem;">FILTER AKTIF:</span>
        {chips_html}
        <span class="chip" style="background:rgba(239,68,68,0.12);color:#fca5a5;border-color:rgba(239,68,68,0.3);">
            Menampilkan {len(df)} dari {len(df_raw)} transaksi
        </span>
    </div>
    """, unsafe_allow_html=True)

# ── No data ──
if df.empty:
    st.warning("⚠️ Tidak ada transaksi yang cocok dengan filter. Coba ubah filter di sidebar.")
    st.stop()

# ── Tabs preview ──
tab_all, tab_cr, tab_db = st.tabs(["📋 Semua Transaksi", "✅ Kredit (Masuk)", "🔴 Debet (Keluar)"])
with tab_all: render_preview(df)
with tab_cr:  render_preview(df, tipe='CR')
with tab_db:  render_preview(df, tipe='DB')

# ── Download ──
st.markdown("---")
col1, col2, col3 = st.columns([1, 2, 1])
with col2:
    out_name = re.sub(r'\.csv$', '_REKAP.xlsx', fname, flags=re.I)
    with st.spinner("Menyiapkan file Excel..."):
        excel_bytes, n_cr, n_db = generate_excel(filtered_tx, summary, fname, filter_info)

    st.markdown(f"""
    <div style="background:rgba(30,41,59,0.8);border-radius:10px;padding:1rem 1.4rem;margin-bottom:0.8rem;border:1px solid rgba(255,255,255,0.07);">
        <div style="color:#94a3b8;font-size:0.72rem;font-weight:700;letter-spacing:0.08em;text-transform:uppercase;margin-bottom:0.6rem;">
            Output Excel — 4 Sheet {'(Data Terfilter)' if is_filtered else ''}
        </div>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:0.4rem 1.5rem;">
            <span style="color:#94a3b8;font-size:0.82rem;">📋 Semua Transaksi <b style="color:#e2e8f0">({len(df)})</b></span>
            <span style="color:#4ade80;font-size:0.82rem;">✅ Kredit <b>({n_cr})</b></span>
            <span style="color:#f87171;font-size:0.82rem;">🔴 Debet <b>({n_db})</b></span>
            <span style="color:#60a5fa;font-size:0.82rem;">📊 Ringkasan & Top 10</span>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.download_button(
        label=f"⬇️  Download Excel — {out_name}",
        data=excel_bytes,
        file_name=out_name,
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

st.markdown('<div class="footer">Rekap Mutasi Rekening · Streamlit + openpyxl</div>', unsafe_allow_html=True)
