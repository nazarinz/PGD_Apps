import streamlit as st
import pandas as pd
import pdfplumber
import re
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

st.set_page_config(page_title="Infor vs SAP Comparator", page_icon="\U0001f4e6", layout="wide")

# ── Colors ────────────────────────────────────────────────────────────────────
C = dict(
    GREEN="00C853", RED="D50000", ORANGE="E65100",
    LGRAY="F5F5F5", DGRAY="424242", WHITE="FFFFFF",
    TEAL="00695C", DARKGREEN="1B5E20", BLUE_HDR="1A237E",
    INFO_BG="E3F2FD", INFO_FG="0D47A1",
    INFOR_HDR="1565C0",
    INFOR_BG ="BBDEFB",
    SAP_HDR  ="2E7D32",
    SAP_BG   ="C8E6C9",
    CALC_HDR ="6A1B9A",
    CALC_BG  ="E1BEE7",
    KEY_HDR  ="37474F",
    KEY_BG   ="ECEFF1",
)

def hfont(color="FFFFFF", sz=10): return Font(name="Calibri", bold=True, color=color, size=sz)
def cfont(bold=False, color="000000", sz=10): return Font(name="Calibri", bold=bold, color=color, size=sz)
def fill(h): return PatternFill("solid", start_color=h, fgColor=h)
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)
def bdr():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)

def wc(ws, r, c, v, bg=None, fnt=None, aln=None, b=True, nf=None):
    cell = ws.cell(row=r, column=c, value=v)
    if bg:  cell.fill      = fill(bg)
    if fnt: cell.font      = fnt
    if aln: cell.alignment = aln
    if b:   cell.border    = bdr()
    if nf:  cell.number_format = nf
    return cell

def span(ws, row, c1, c2, label, bg, fg="FFFFFF", sz=9):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1, value=label)
    cell.fill = fill(bg); cell.font = hfont(color=fg, sz=sz)
    cell.alignment = center(); cell.border = bdr()
    for cc in range(c1 + 1, c2 + 1):
        ws.cell(row=row, column=cc).fill   = fill(bg)
        ws.cell(row=row, column=cc).border = bdr()


# ── Excel (Infor) Parsing ─────────────────────────────────────────────────────
def load_excel(file):
    df = pd.read_excel(file, dtype=str)
    df.columns = df.columns.str.strip()
    return df

def _find_ctn_qty_col(df):
    """Find the 'Carton Qty per Size' column flexibly — works for L15, L19, etc."""
    for col in df.columns:
        if "Carton Qty per Size" in col:
            return col
    return None

def xl_sizes_for_po(df, po):
    rows = df[df["Order #"].str.strip() == po.strip()].copy()
    ctn_col = _find_ctn_qty_col(df)
    out = []
    for _, r in rows.iterrows():
        try: qty = int(float(r.get("Quantity", 0)))
        except: qty = 0
        try:
            ctn_qty = int(float(r.get(ctn_col, 0))) if ctn_col else None
        except:
            ctn_qty = None
        out.append({
            "UK Size":        str(r.get("Manufacturing Size", "")).strip(),
            "US Size":        str(r.get("Customer Size", "")).strip(),
            "Infor Qty":      qty,
            "Infor Qty/CTN":  ctn_qty,
            "XL Line":        str(r.get("Item Line Number", "")).strip(),
        })
    return pd.DataFrame(out) if out else pd.DataFrame(
        columns=["UK Size", "US Size", "Infor Qty", "Infor Qty/CTN", "XL Line"])

def xl_hdr_for_po(df, po):
    r = df[df["Order #"].str.strip() == po.strip()]
    if r.empty: return {}
    r = r.iloc[0]
    return {
        "PO Number":   po,
        "Article":     str(r.get("Article Number", "")).strip(),
        "Model":       str(r.get("Model Name", "")).strip(),
        "Ship Method": str(r.get("Shipment Method", "")),
        "Pack Mode":   str(r.get("Pack Mode", r.get("VAS/SHAS L15 \u2013 Packing Mode", ""))),
        "Destination": str(r.get("FinalDestinationName", "")),
    }


# ── PDF (SAP) Parsing ─────────────────────────────────────────────────────────
def _f(pat, text, default=""):
    m = re.search(pat, text)
    return m.group(1).strip() if m else default

def normalize_size(s):
    return s.strip()

def parse_pdf(file, filename=""):
    result = {}
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            po = _f(r"Cust\.PO\s*:\s*(\d+)", text)
            if not po:
                continue
            hdr = {
                "PO Number":   po,
                "Article":     _f(r"ART\. NO[:\s]+(\w+)", text).strip(),
                "Model":       _f(r"Model\s*:\s*([A-Z][A-Z0-9 ]+?)(?:\n|Arr|Ship|Coun|Sub|End|Sale|Cust|OUTER)", text).strip(),
                "Ship Method": _f(r"ShipType:\s*\d+\s+(\w+)", text),
                "Pack Mode":   _f(r"(SSP|MSP|USSP|SP)\s+TOTAL", text),
                "Total Pairs": _f(r"Pair[:\s]+([\d,]+)\s+Pairs", text).replace(",", ""),
                "Total CTNs":  _f(r"Ctns[:\s]+([\d,]+)\s+Ctn", text).replace(",", ""),
                "Arr Port":    _f(r"Arr\.Po[:\s]+(\w+)", text),
                "Source File": filename,
            }

            pat = re.compile(
                r"(\d+[-\u2013]\d+|\d+)\s+(\d+)\s+(\d+)\s+(\d+)"
                r"\s+(\d+(?:-K|K|-)?)\s+(\d+(?:-K|K|-)?)\s+[\d.]+"
            )

            rows = [
                {
                    "CTN Range": m.group(1),
                    "CTNs":      int(m.group(2)),
                    "Qty/CTN":   int(m.group(3)),
                    "SAP Qty":   int(m.group(4)),
                    "UK Size":   normalize_size(m.group(5)),
                    "US Size":   normalize_size(m.group(6)),
                }
                for m in pat.finditer(text)
            ]

            result[po] = {
                "header": hdr,
                "sizes":  pd.DataFrame(rows) if rows else pd.DataFrame(),
            }
    return result


# ── Comparison Logic ──────────────────────────────────────────────────────────
def compare_po_fields(xl_df, pdf_data, po):
    xs = xl_sizes_for_po(xl_df, po)
    pg = pdf_data.get(po, {}); ph = pg.get("header", {}); ps = pg.get("sizes", pd.DataFrame())

    rows = []

    # ── Total Qty ──
    infor_total = int(xs["Infor Qty"].sum()) if not xs.empty else 0
    sap_total   = int(ph.get("Total Pairs", 0) or 0)
    rows.append({
        "Field":       "Total Qty (Pairs)",
        "Infor Value": infor_total,
        "SAP Value":   sap_total,
        "Status":      "\u2705 MATCH" if infor_total == sap_total else "\u274c MISMATCH",
    })

    # ── Merge sizes ──
    if not xs.empty and ps is not None and not ps.empty:
        mg = pd.merge(
            xs[["UK Size", "Infor Qty", "Infor Qty/CTN"]],
            ps[["UK Size", "SAP Qty", "Qty/CTN"]],
            on="UK Size", how="outer"
        )
    elif not xs.empty:
        mg = xs[["UK Size", "Infor Qty", "Infor Qty/CTN"]].copy()
        mg["SAP Qty"]  = 0
        mg["Qty/CTN"]  = None
    elif ps is not None and not ps.empty:
        mg = ps[["UK Size", "SAP Qty", "Qty/CTN"]].copy()
        mg["Infor Qty"]     = 0
        mg["Infor Qty/CTN"] = None
    else:
        mg = pd.DataFrame(columns=["UK Size", "Infor Qty", "Infor Qty/CTN", "SAP Qty", "Qty/CTN"])

    if not mg.empty:
        mg["Infor Qty"]     = mg["Infor Qty"].fillna(0).astype(int)
        mg["SAP Qty"]       = mg["SAP Qty"].fillna(0).astype(int)
        mg = mg.sort_values("UK Size").reset_index(drop=True)

        for _, row in mg.iterrows():
            iv, sv = int(row["Infor Qty"]), int(row["SAP Qty"])
            rows.append({
                "Field":       f"Qty Size {row['UK Size']}",
                "Infor Value": iv,
                "SAP Value":   sv,
                "Status":      "\u2705 MATCH" if iv == sv else "\u274c MISMATCH",
            })

        # ── Qty/CTN per size comparison ──
        has_infor_ctn = "Infor Qty/CTN" in mg.columns and mg["Infor Qty/CTN"].notna().any()
        has_sap_ctn   = "Qty/CTN" in mg.columns and mg["Qty/CTN"].notna().any()

        if has_infor_ctn or has_sap_ctn:
            for _, row in mg.iterrows():
                iv_ctn = row.get("Infor Qty/CTN")
                sv_ctn = row.get("Qty/CTN")
                # Only add row if at least one side has data
                if pd.isna(iv_ctn) and pd.isna(sv_ctn):
                    continue
                iv_ctn_i = int(iv_ctn) if pd.notna(iv_ctn) else None
                sv_ctn_i = int(sv_ctn) if pd.notna(sv_ctn) else None
                if iv_ctn_i is None:
                    status = "ONLY IN SAP"
                elif sv_ctn_i is None:
                    status = "ONLY IN INFOR"
                elif iv_ctn_i == sv_ctn_i:
                    status = "\u2705 MATCH"
                else:
                    status = "\u274c MISMATCH"
                rows.append({
                    "Field":       f"Qty/CTN Size {row['UK Size']}",
                    "Infor Value": iv_ctn_i if iv_ctn_i is not None else "",
                    "SAP Value":   sv_ctn_i if sv_ctn_i is not None else "",
                    "Status":      status,
                })

    return pd.DataFrame(rows)


def compare_po_size_detail(xl_df, pdf_data, po):
    xs = xl_sizes_for_po(xl_df, po)
    pg = pdf_data.get(po, {}); ps = pg.get("sizes", pd.DataFrame())
    COLS = ["UK Size", "US Size", "XL Line",
            "Infor Qty", "Infor Qty/CTN",
            "CTN Range", "CTNs", "Qty/CTN", "SAP Qty",
            "Diff", "Diff Qty/CTN", "Status", "Status Qty/CTN"]

    if not xs.empty and ps is not None and not ps.empty:
        xl_c  = xs[["UK Size", "US Size", "Infor Qty", "Infor Qty/CTN", "XL Line"]]
        sap_c = ps[["UK Size", "US Size", "SAP Qty", "Qty/CTN", "CTN Range", "CTNs"]].rename(
            columns={"US Size": "_sap_us"})
        mg = pd.merge(xl_c, sap_c, on="UK Size", how="outer")
        mg["US Size"] = mg["US Size"].combine_first(mg["_sap_us"])
        mg = mg.drop(columns=["_sap_us"], errors="ignore")
    elif not xs.empty:
        mg = xs[["UK Size", "US Size", "Infor Qty", "Infor Qty/CTN", "XL Line"]].copy()
        mg["SAP Qty"] = 0
        mg["Qty/CTN"] = None
        for col in ["CTN Range", "CTNs"]: mg[col] = ""
    elif ps is not None and not ps.empty:
        mg = ps[["UK Size", "US Size", "SAP Qty", "Qty/CTN", "CTN Range", "CTNs"]].copy()
        mg["Infor Qty"]     = 0
        mg["Infor Qty/CTN"] = None
        mg["XL Line"]       = ""
    else:
        return pd.DataFrame(columns=COLS)

    mg["Infor Qty"]     = mg["Infor Qty"].fillna(0).astype(int)
    mg["SAP Qty"]       = mg["SAP Qty"].fillna(0).astype(int)
    mg["Diff"]          = mg["Infor Qty"] - mg["SAP Qty"]

    # Qty/CTN diff
    def _ctn_diff(row):
        iv = row.get("Infor Qty/CTN")
        sv = row.get("Qty/CTN")
        if pd.isna(iv) or pd.isna(sv):
            return None
        try:
            return int(iv) - int(sv)
        except:
            return None

    mg["Diff Qty/CTN"] = mg.apply(_ctn_diff, axis=1)

    def _st(row):
        no_sap   = str(row.get("CTN Range", "")).strip() in ("", "nan", "None", "NaN")
        no_infor = str(row.get("XL Line",   "")).strip() in ("", "nan", "None", "NaN")
        if no_sap:   return "ONLY IN INFOR"
        if no_infor: return "ONLY IN SAP"
        return "\u2705 MATCH" if row["Diff"] == 0 else "\u274c MISMATCH"

    def _st_ctn(row):
        iv = row.get("Infor Qty/CTN")
        sv = row.get("Qty/CTN")
        no_sap   = str(row.get("CTN Range", "")).strip() in ("", "nan", "None", "NaN")
        no_infor = str(row.get("XL Line",   "")).strip() in ("", "nan", "None", "NaN")
        if no_sap:             return "ONLY IN INFOR"
        if no_infor:           return "ONLY IN SAP"
        if pd.isna(iv):        return "NO INFOR DATA"
        if pd.isna(sv):        return "NO SAP DATA"
        try:
            return "\u2705 MATCH" if int(iv) == int(sv) else "\u274c MISMATCH"
        except:
            return "\u274c MISMATCH"

    mg["Status"]         = mg.apply(_st, axis=1)
    mg["Status Qty/CTN"] = mg.apply(_st_ctn, axis=1)

    # Coerce Infor Qty/CTN for display
    mg["Infor Qty/CTN"] = mg["Infor Qty/CTN"].apply(
        lambda x: int(x) if pd.notna(x) else "")
    mg["Qty/CTN"] = mg["Qty/CTN"].apply(
        lambda x: int(x) if pd.notna(x) else "")
    mg["Diff Qty/CTN"] = mg["Diff Qty/CTN"].apply(
        lambda x: int(x) if pd.notna(x) else "")

    for col in COLS:
        if col not in mg.columns: mg[col] = ""
    return mg[COLS].sort_values("UK Size").reset_index(drop=True)


def po_info(xl_df, pdf_data, po):
    xh = xl_hdr_for_po(xl_df, po)
    ph = pdf_data.get(po, {}).get("header", {})
    return {
        "Article (Infor)": xh.get("Article", ""),
        "Article (SAP)":   ph.get("Article", ""),
        "Model (Infor)":   xh.get("Model", ""),
        "Model (SAP)":     ph.get("Model", ""),
    }


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL REPORT
# ══════════════════════════════════════════════════════════════════════════════
def build_report(xl_df, pdf_data, all_pos, xl_filename, pdf_filename):
    wb = Workbook()
    xl_pos_set = set(xl_df["Order #"].str.strip().tolist())

    STATUS_STYLE = {
        "\u2705 MATCH":    (C["GREEN"],  "FFFFFF"),
        "\u274c MISMATCH": (C["RED"],    "FFFFFF"),
        "ONLY IN INFOR":   (C["ORANGE"], "FFFFFF"),
        "ONLY IN SAP":     (C["ORANGE"], "FFFFFF"),
        "NO INFOR DATA":   (C["ORANGE"], "FFFFFF"),
        "NO SAP DATA":     (C["ORANGE"], "FFFFFF"),
    }

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — PO Compare Summary
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active; ws1.title = "PO Compare Summary"
    ws1.sheet_view.showGridLines = False

    ws1.merge_cells("A1:I1")
    c = ws1["A1"]
    c.value = "PO COMPARE SUMMARY \u2014 Infor vs SAP Carton"
    c.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    c.fill = fill(C["DGRAY"]); c.alignment = center(); ws1.row_dimensions[1].height = 30

    ws1.merge_cells("A2:I2")
    c = ws1["A2"]
    c.value = (
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
        f"Infor: {xl_filename}  |  SAP: {pdf_filename}"
    )
    c.font = Font(name="Calibri", size=9, color="616161")
    c.fill = fill(C["LGRAY"]); c.alignment = left(); ws1.row_dimensions[2].height = 14

    ws1.merge_cells("A3:I3")
    c = ws1["A3"]
    c.value = (
        "\u2139\ufe0f  Compared: Total Qty (Pairs) + Qty per Size + Qty/CTN per Size  |  "
        "Info only (not compared): Article \u00b7 Model"
    )
    c.font = Font(name="Calibri", italic=True, size=9, color=C["INFO_FG"])
    c.fill = fill(C["INFO_BG"]); c.alignment = left(); ws1.row_dimensions[3].height = 14

    ws1.merge_cells("A4:I4")
    c = ws1["A4"]
    c.value = (
        "Column color guide:   "
        "\u25a0 DARK BLUE = data dari Infor   "
        "\u25a0 DARK GREEN = data dari SAP   "
        "\u25a0 DARK PURPLE = hasil kalkulasi   "
        "\u25a0 DARK GRAY = key bersama"
    )
    c.font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
    c.fill = fill("263238"); c.alignment = left(); ws1.row_dimensions[4].height = 14

    ws1.row_dimensions[5].height = 18
    span(ws1, 5, 1, 1, "\U0001f511 KEY",             C["KEY_HDR"])
    span(ws1, 5, 2, 3, "\U0001f4ca FROM INFOR",      C["INFOR_HDR"])
    span(ws1, 5, 4, 5, "\U0001f4c4 FROM SAP",        C["SAP_HDR"])
    span(ws1, 5, 6, 8, "\U0001f7e3 COMPARE RESULT",  C["CALC_HDR"])
    span(ws1, 5, 9, 9, "\U0001f7e3 OVERALL",         C["CALC_HDR"])

    ws1.row_dimensions[6].height = 28
    col_defs = [
        (1, "PO Number",         C["KEY_HDR"],   16),
        (2, "Infor File",        C["INFOR_HDR"], 30),
        (3, "Article\n(Infor)",  C["INFOR_HDR"], 14),
        (4, "SAP File(s)",       C["SAP_HDR"],   30),
        (5, "Article\n(SAP)",    C["SAP_HDR"],   14),
        (6, "Total\nFields",     C["CALC_HDR"],  11),
        (7, "\u2705 Match",      C["CALC_HDR"],  10),
        (8, "\u274c Mismatch",   C["CALC_HDR"],  12),
        (9, "Result",            C["CALC_HDR"],  18),
    ]
    for col, lbl, bg, w in col_defs:
        wc(ws1, 6, col, lbl, bg=bg, fnt=hfont(sz=9), aln=center())
        ws1.column_dimensions[get_column_letter(col)].width = w

    for ri, po in enumerate(all_pos, 7):
        df_f    = compare_po_fields(xl_df, pdf_data, po)
        total_f = len(df_f)
        n_m     = (df_f["Status"] == "\u2705 MATCH").sum()    if not df_f.empty else 0
        n_mm    = (df_f["Status"] == "\u274c MISMATCH").sum() if not df_f.empty else 0
        sap_src = pdf_data.get(po, {}).get("header", {}).get("Source File", pdf_filename)
        info    = po_info(xl_df, pdf_data, po)

        if po not in pdf_data:     r_txt, r_bg = "\u26a0\ufe0f NO SAP DATA",   C["ORANGE"]
        elif po not in xl_pos_set: r_txt, r_bg = "\u26a0\ufe0f NO INFOR DATA", C["ORANGE"]
        elif n_mm == 0:            r_txt, r_bg = "\u2705 ALL OK",               C["GREEN"]
        else:                      r_txt, r_bg = f"\u274c {n_mm} ISSUE(S)",     C["RED"]

        alt = C["LGRAY"] if ri % 2 == 0 else C["WHITE"]
        ws1.row_dimensions[ri].height = 15

        wc(ws1, ri, 1, po,                      bg=C["KEY_BG"],   fnt=cfont(bold=True, sz=9), aln=left())
        wc(ws1, ri, 2, xl_filename,              bg=C["INFOR_BG"], fnt=cfont(sz=8),            aln=left())
        wc(ws1, ri, 3, info["Article (Infor)"],  bg=C["INFOR_BG"], fnt=cfont(sz=9),            aln=center())
        wc(ws1, ri, 4, sap_src,                  bg=C["SAP_BG"],   fnt=cfont(sz=8),            aln=left())
        wc(ws1, ri, 5, info["Article (SAP)"],    bg=C["SAP_BG"],   fnt=cfont(sz=9),            aln=center())
        wc(ws1, ri, 6, total_f,                  bg=alt,           fnt=cfont(sz=9),            aln=center())
        wc(ws1, ri, 7, n_m,
           bg=alt, fnt=cfont(bold=True, color=C["DARKGREEN"], sz=9), aln=center())
        wc(ws1, ri, 8, n_mm if n_mm > 0 else 0,
           bg=alt, fnt=cfont(bold=n_mm > 0, color=C["RED"] if n_mm > 0 else "000000", sz=9), aln=center())
        wc(ws1, ri, 9, r_txt, bg=r_bg, fnt=hfont(sz=9), aln=center())

    ws1.freeze_panes = "A7"

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2 — PO Compare Detail
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("PO Compare Detail")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:F1")
    c = ws2["A1"]
    c.value = "PO COMPARE DETAIL \u2014 Field-by-Field"
    c.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    c.fill = fill(C["DGRAY"]); c.alignment = center(); ws2.row_dimensions[1].height = 26

    ws2.merge_cells("A2:F2")
    c = ws2["A2"]
    c.value = (
        "\u2139\ufe0f  Compared: Total Qty (Pairs) + Qty per UK Size + Qty/CTN per UK Size  |  "
        "Blue = Infor value  |  Green = SAP value  |  Purple = Diff (Infor \u2212 SAP)"
    )
    c.font = Font(name="Calibri", italic=True, size=9, color=C["INFO_FG"])
    c.fill = fill(C["INFO_BG"]); c.alignment = left(); ws2.row_dimensions[2].height = 14

    ws2.row_dimensions[3].height = 18
    span(ws2, 3, 1, 2, "\U0001f511 KEY",        C["KEY_HDR"])
    span(ws2, 3, 3, 3, "\U0001f4ca FROM INFOR", C["INFOR_HDR"])
    span(ws2, 3, 4, 4, "\U0001f4c4 FROM SAP",   C["SAP_HDR"])
    span(ws2, 3, 5, 6, "\U0001f7e3 CALCULATED", C["CALC_HDR"])

    ws2.row_dimensions[4].height = 18
    ws2_cols = [
        (1, "PO Number",                C["KEY_HDR"],   16),
        (2, "Field",                    C["KEY_HDR"],   32),
        (3, "Infor Value",              C["INFOR_HDR"], 16),
        (4, "SAP Value",                C["SAP_HDR"],   16),
        (5, "Diff\n(Infor\u2212SAP)",   C["CALC_HDR"],  12),
        (6, "Status",                   C["CALC_HDR"],  18),
    ]
    for col, lbl, bg, w in ws2_cols:
        wc(ws2, 4, col, lbl, bg=bg, fnt=hfont(sz=9), aln=center())
        ws2.column_dimensions[get_column_letter(col)].width = w

    ri2 = 5
    for po in all_pos:
        df_f = compare_po_fields(xl_df, pdf_data, po)
        if df_f.empty: continue
        for _, row in df_f.iterrows():
            st = row["Status"]; is_m = (st == "\u2705 MATCH")
            alt = C["LGRAY"] if ri2 % 2 == 0 else C["WHITE"]
            iv, sv = row["Infor Value"], row["SAP Value"]
            try:    diff = int(iv) - int(sv)
            except: diff = ""

            ws2.row_dimensions[ri2].height = 14
            wc(ws2, ri2, 1, po,           bg=C["KEY_BG"],  fnt=cfont(bold=True, sz=9), aln=left())
            wc(ws2, ri2, 2, row["Field"], bg=alt,          fnt=cfont(sz=9),            aln=left())
            wc(ws2, ri2, 3, iv,
               bg="BBDEFB" if not is_m else C["INFOR_BG"],
               fnt=cfont(bold=not is_m, color="0D47A1", sz=9), aln=center())
            wc(ws2, ri2, 4, sv,
               bg="C8E6C9" if not is_m else C["SAP_BG"],
               fnt=cfont(bold=not is_m, color="1B5E20", sz=9), aln=center())
            diff_val = f"{diff:+}" if isinstance(diff, int) and diff != 0 else (0 if diff == 0 else "")
            wc(ws2, ri2, 5, diff_val,
               bg="FFEBEE" if not is_m else C["CALC_BG"],
               fnt=cfont(bold=not is_m, color=C["RED"] if not is_m else "6A1B9A", sz=9), aln=center())
            wc(ws2, ri2, 6, st,
               bg=C["GREEN"] if is_m else C["RED"],
               fnt=hfont(sz=9), aln=center())
            ri2 += 1

    ws2.freeze_panes = "A5"

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3 — Size Detail
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Size Detail")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:Q1")
    c = ws3["A1"]
    c.value = (
        "SIZE DETAIL \u2014 Semua PO  |  "
        "Warna kolom: \u25a0 Biru=Infor  \u25a0 Hijau=SAP  \u25a0 Ungu=Kalkulasi  \u25a0 Abu=Key bersama"
    )
    c.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    c.fill = fill(C["DGRAY"]); c.alignment = left(); ws3.row_dimensions[1].height = 22

    ws3.merge_cells("A2:Q2")
    c = ws3["A2"]
    c.value = (
        "PO / UK Size / US Size = key bersama  |  "
        "Article(Infor), Model(Infor), XL Line, Infor Qty, Infor Qty/CTN = FROM INFOR  |  "
        "Article(SAP), Model(SAP), CTN Range, CTNs, Qty/CTN, SAP Qty = FROM SAP  |  "
        "Diff, Diff Qty/CTN, Status, Status Qty/CTN = CALCULATED"
    )
    c.font = Font(name="Calibri", italic=True, size=8, color="37474F")
    c.fill = fill("FAFAFA"); c.alignment = left(); ws3.row_dimensions[2].height = 14

    # Column spans — now 17 columns total
    ws3.row_dimensions[3].height = 20
    span(ws3, 3,  1,  1, "\U0001f511 KEY",                                C["KEY_HDR"])
    span(ws3, 3,  2,  2, "\U0001f4ca INFOR \u2014 info only",             C["INFOR_HDR"])
    span(ws3, 3,  3,  3, "\U0001f4c4 SAP \u2014 info only",               C["SAP_HDR"])
    span(ws3, 3,  4,  4, "\U0001f4ca INFOR \u2014 info only",             C["INFOR_HDR"])
    span(ws3, 3,  5,  5, "\U0001f4c4 SAP \u2014 info only",               C["SAP_HDR"])
    span(ws3, 3,  6,  7, "\U0001f511 KEY (shared)",                       C["KEY_HDR"])
    span(ws3, 3,  8,  9, "\U0001f4ca FROM INFOR \u2014 compared",         C["INFOR_HDR"])
    span(ws3, 3, 10, 13, "\U0001f4c4 FROM SAP \u2014 compared",           C["SAP_HDR"])
    span(ws3, 3, 14, 17, "\U0001f7e3 CALCULATED",                         C["CALC_HDR"])

    ws3.row_dimensions[4].height = 32
    sd_cols = [
        ( 1, "PO\nNumber",              C["KEY_HDR"],   14),
        ( 2, "Article\n(Infor)",        C["INFOR_HDR"], 13),
        ( 3, "Article\n(SAP)",          C["SAP_HDR"],   13),
        ( 4, "Model\n(Infor)",          C["INFOR_HDR"], 22),
        ( 5, "Model\n(SAP)",            C["SAP_HDR"],   22),
        ( 6, "UK\nSize",                C["KEY_HDR"],    9),
        ( 7, "US\nSize",                C["KEY_HDR"],    9),
        ( 8, "XL Line\n\u2190Infor",   C["INFOR_HDR"],  9),
        ( 9, "Infor Qty\n(pairs)",      C["INFOR_HDR"], 12),
        (10, "CTN Range\n\u2190SAP",    C["SAP_HDR"],   13),
        (11, "CTNs\n\u2190SAP",         C["SAP_HDR"],    8),
        (12, "Qty/CTN\n\u2190SAP",      C["SAP_HDR"],    9),
        (13, "SAP Qty\n(pairs)",        C["SAP_HDR"],   12),
        (14, "Diff\n(Infor\u2212SAP)",  C["CALC_HDR"],  10),
        (15, "Status\nQty",             C["CALC_HDR"],  16),
        (16, "Infor\nQty/CTN\u2190",    C["INFOR_HDR"], 12),
        # col 16 header needs to appear in INFOR band — we'll override span below
        # Actually easier to keep CALCULATED band for cols 14-17 and add Infor Qty/CTN separate
    ]
    # Override: col 16 is Infor Qty/CTN (Infor), col 17 is Diff + Status Qty/CTN (CALC)
    # Redefine cleanly:
    sd_cols = [
        ( 1, "PO\nNumber",               C["KEY_HDR"],   14),
        ( 2, "Article\n(Infor)",         C["INFOR_HDR"], 13),
        ( 3, "Article\n(SAP)",           C["SAP_HDR"],   13),
        ( 4, "Model\n(Infor)",           C["INFOR_HDR"], 22),
        ( 5, "Model\n(SAP)",             C["SAP_HDR"],   22),
        ( 6, "UK\nSize",                 C["KEY_HDR"],    9),
        ( 7, "US\nSize",                 C["KEY_HDR"],    9),
        ( 8, "XL Line\n\u2190Infor",    C["INFOR_HDR"],  9),
        ( 9, "Infor Qty\n(pairs)",       C["INFOR_HDR"], 13),
        (10, "Infor\nQty/CTN",           C["INFOR_HDR"], 13),
        (11, "CTN Range\n\u2190SAP",     C["SAP_HDR"],   13),
        (12, "CTNs\n\u2190SAP",          C["SAP_HDR"],    8),
        (13, "SAP\nQty/CTN",             C["SAP_HDR"],   12),
        (14, "SAP Qty\n(pairs)",         C["SAP_HDR"],   12),
        (15, "Diff Qty\n(I\u2212S)",     C["CALC_HDR"],  10),
        (16, "Status\nQty",              C["CALC_HDR"],  16),
        (17, "Diff\nQty/CTN",            C["CALC_HDR"],  12),
        (18, "Status\nQty/CTN",          C["CALC_HDR"],  18),
    ]

    # Re-do row 3 spans for 18 columns
    ws3.merge_cells("A1:R1")
    ws3.merge_cells("A2:R2")
    # Redo span row 3 with correct column groupings for 18 cols
    # Clear previous spans
    for cc in range(1, 19):
        ws3.cell(row=3, column=cc).value = None
    span(ws3, 3,  1,  1, "\U0001f511 KEY",                                C["KEY_HDR"])
    span(ws3, 3,  2,  2, "\U0001f4ca INFOR \u2014 info only",             C["INFOR_HDR"])
    span(ws3, 3,  3,  3, "\U0001f4c4 SAP \u2014 info only",               C["SAP_HDR"])
    span(ws3, 3,  4,  4, "\U0001f4ca INFOR \u2014 info only",             C["INFOR_HDR"])
    span(ws3, 3,  5,  5, "\U0001f4c4 SAP \u2014 info only",               C["SAP_HDR"])
    span(ws3, 3,  6,  7, "\U0001f511 KEY (shared)",                       C["KEY_HDR"])
    span(ws3, 3,  8, 10, "\U0001f4ca FROM INFOR \u2014 compared",         C["INFOR_HDR"])
    span(ws3, 3, 11, 14, "\U0001f4c4 FROM SAP \u2014 compared",           C["SAP_HDR"])
    span(ws3, 3, 15, 18, "\U0001f7e3 CALCULATED",                         C["CALC_HDR"])

    for col, lbl, bg, w in sd_cols:
        wc(ws3, 4, col, lbl, bg=bg, fnt=hfont(sz=9), aln=center())
        ws3.column_dimensions[get_column_letter(col)].width = w

    ri3 = 5
    for po in all_pos:
        info = po_info(xl_df, pdf_data, po)
        sd   = compare_po_size_detail(xl_df, pdf_data, po)
        if sd.empty: continue
        for _, row in sd.iterrows():
            st         = row.get("Status", "")
            st_ctn     = row.get("Status Qty/CTN", "")
            s_bg, s_fc = STATUS_STYLE.get(st, (C["WHITE"], "000000"))
            sc_bg, sc_fc = STATUS_STYLE.get(st_ctn, (C["WHITE"], "000000"))
            alt        = C["LGRAY"] if ri3 % 2 == 0 else C["WHITE"]
            ws3.row_dimensions[ri3].height = 14

            try:    diff_i = int(row.get("Diff", 0))
            except: diff_i = 0

            try:    diff_ctn = int(row.get("Diff Qty/CTN", 0))
            except: diff_ctn = 0

            is_mm      = (st == "\u274c MISMATCH")
            is_mm_ctn  = (st_ctn == "\u274c MISMATCH")
            only_infor = (st == "ONLY IN INFOR")
            only_sap   = (st == "ONLY IN SAP")

            iq_bg   = "BBDEFB" if is_mm      else (C["INFOR_BG"] if not only_sap   else "FFCDD2")
            iqc_bg  = "BBDEFB" if is_mm_ctn  else (C["INFOR_BG"] if not only_sap   else "FFCDD2")
            sq_bg   = "C8E6C9" if is_mm      else (C["SAP_BG"]   if not only_infor else "FFCDD2")
            sqc_bg  = "C8E6C9" if is_mm_ctn  else (C["SAP_BG"]   if not only_infor else "FFCDD2")
            ctn_bg  = C["SAP_BG"]   if not only_infor else "FFCDD2"
            xl_bg   = C["INFOR_BG"] if not only_sap   else "FFCDD2"

            infor_ctn_val = row.get("Infor Qty/CTN", "")
            sap_ctn_val   = row.get("Qty/CTN", "")
            diff_ctn_disp = f"{diff_ctn:+}" if isinstance(diff_ctn, int) and diff_ctn != 0 else (0 if diff_ctn == 0 else "")

            wc(ws3, ri3,  1, po,                              bg=C["KEY_BG"],   fnt=cfont(bold=True, sz=9),            aln=left())
            wc(ws3, ri3,  2, info["Article (Infor)"],         bg=C["INFOR_BG"], fnt=cfont(sz=9, color=C["INFO_FG"]),   aln=left())
            wc(ws3, ri3,  3, info["Article (SAP)"],           bg=C["SAP_BG"],   fnt=cfont(sz=9, color="1B5E20"),       aln=left())
            wc(ws3, ri3,  4, info["Model (Infor)"],           bg=C["INFOR_BG"], fnt=cfont(sz=9, color=C["INFO_FG"]),   aln=left())
            wc(ws3, ri3,  5, info["Model (SAP)"],             bg=C["SAP_BG"],   fnt=cfont(sz=9, color="1B5E20"),       aln=left())
            wc(ws3, ri3,  6, row.get("UK Size", ""),          bg=C["KEY_BG"],   fnt=cfont(bold=True, sz=9),            aln=center())
            wc(ws3, ri3,  7, row.get("US Size", ""),          bg=C["KEY_BG"],   fnt=cfont(sz=9),                       aln=center())
            wc(ws3, ri3,  8, row.get("XL Line", ""),          bg=xl_bg,         fnt=cfont(sz=9, color=C["INFO_FG"]),   aln=center())
            wc(ws3, ri3,  9, int(row.get("Infor Qty", 0)),    bg=iq_bg,         fnt=cfont(bold=is_mm, color="0D47A1", sz=9), aln=center())
            wc(ws3, ri3, 10, infor_ctn_val,                   bg=iqc_bg,        fnt=cfont(bold=is_mm_ctn, color="0D47A1", sz=9), aln=center())
            wc(ws3, ri3, 11, row.get("CTN Range", ""),        bg=ctn_bg,        fnt=cfont(sz=9, color="1B5E20"),       aln=center())
            wc(ws3, ri3, 12, row.get("CTNs", ""),             bg=ctn_bg,        fnt=cfont(sz=9, color="1B5E20"),       aln=center())
            wc(ws3, ri3, 13, sap_ctn_val,                     bg=sqc_bg,        fnt=cfont(bold=is_mm_ctn, color="1B5E20", sz=9), aln=center())
            wc(ws3, ri3, 14, int(row.get("SAP Qty", 0)),      bg=sq_bg,         fnt=cfont(bold=is_mm, color="1B5E20", sz=9), aln=center())
            wc(ws3, ri3, 15,
               f"{diff_i:+}" if diff_i != 0 else 0,
               bg="FFEBEE" if diff_i != 0 else C["CALC_BG"],
               fnt=cfont(bold=diff_i != 0, color=C["RED"] if diff_i != 0 else "000000", sz=9), aln=center())
            wc(ws3, ri3, 16, st,   bg=s_bg,  fnt=hfont(color=s_fc, sz=9),  aln=center())
            wc(ws3, ri3, 17, diff_ctn_disp,
               bg="FFEBEE" if diff_ctn != 0 else C["CALC_BG"],
               fnt=cfont(bold=diff_ctn != 0, color=C["RED"] if diff_ctn != 0 else "000000", sz=9), aln=center())
            wc(ws3, ri3, 18, st_ctn, bg=sc_bg, fnt=hfont(color=sc_fc, sz=9), aln=center())
            ri3 += 1

    ws3.freeze_panes = "A5"

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 4 — Discrepancies Only
    # ══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Discrepancies Only")
    ws4.sheet_view.showGridLines = False

    ws4.merge_cells("A1:F1")
    c = ws4["A1"]
    c.value = "\u274c DISCREPANCIES ONLY \u2014 Hanya baris Infor \u2260 SAP"
    c.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    c.fill = fill(C["ORANGE"]); c.alignment = center(); ws4.row_dimensions[1].height = 24

    ws4.merge_cells("A2:F2")
    c = ws4["A2"]
    c.value = "Biru = nilai Infor  |  Hijau = nilai SAP  |  Ungu = selisih (Infor \u2212 SAP)  |  Includes Qty and Qty/CTN mismatches"
    c.font = Font(name="Calibri", italic=True, size=9, color="37474F")
    c.fill = fill("FFF3E0"); c.alignment = left(); ws4.row_dimensions[2].height = 13

    ws4.row_dimensions[3].height = 18
    span(ws4, 3, 1, 2, "\U0001f511 KEY",        C["KEY_HDR"])
    span(ws4, 3, 3, 3, "\U0001f4ca FROM INFOR", C["INFOR_HDR"])
    span(ws4, 3, 4, 4, "\U0001f4c4 FROM SAP",   C["SAP_HDR"])
    span(ws4, 3, 5, 6, "\U0001f7e3 CALCULATED", C["CALC_HDR"])

    ws4.row_dimensions[4].height = 18
    disc_cols = [
        (1, "PO Number",             C["KEY_HDR"],   16),
        (2, "Field",                 C["KEY_HDR"],   32),
        (3, "Infor Value",           C["INFOR_HDR"], 14),
        (4, "SAP Value",             C["SAP_HDR"],   14),
        (5, "Diff (Infor\u2212SAP)", C["CALC_HDR"],  12),
        (6, "Status",                C["CALC_HDR"],  18),
    ]
    for col, lbl, bg, w in disc_cols:
        wc(ws4, 4, col, lbl, bg=bg, fnt=hfont(sz=9), aln=center())
        ws4.column_dimensions[get_column_letter(col)].width = w

    ri4 = 5
    for po in all_pos:
        df_f   = compare_po_fields(xl_df, pdf_data, po)
        if df_f.empty: continue
        issues = df_f[df_f["Status"] != "\u2705 MATCH"]
        if issues.empty: continue
        for _, row in issues.iterrows():
            alt = C["LGRAY"] if ri4 % 2 == 0 else C["WHITE"]
            iv, sv = row["Infor Value"], row["SAP Value"]
            try:    diff = int(iv) - int(sv)
            except: diff = ""
            ws4.row_dimensions[ri4].height = 14
            wc(ws4, ri4, 1, po,           bg=C["KEY_BG"],  fnt=cfont(bold=True, sz=9),            aln=left())
            wc(ws4, ri4, 2, row["Field"], bg=alt,          fnt=cfont(sz=9),                       aln=left())
            wc(ws4, ri4, 3, iv,           bg="BBDEFB",     fnt=cfont(bold=True, color="0D47A1", sz=9), aln=center())
            wc(ws4, ri4, 4, sv,           bg="C8E6C9",     fnt=cfont(bold=True, color="1B5E20", sz=9), aln=center())
            wc(ws4, ri4, 5,
               f"{diff:+}" if isinstance(diff, int) else "",
               bg=C["CALC_BG"], fnt=cfont(bold=True, color="6A1B9A", sz=9), aln=center())
            wc(ws4, ri4, 6, row["Status"], bg=C["RED"], fnt=hfont(sz=9), aln=center())
            ri4 += 1

    if ri4 == 5:
        c = ws4.cell(5, 1, "\u2705 Tidak ada discrepancy \u2014 semua field qty cocok!")
        c.font = Font(name="Calibri", bold=True, color=C["GREEN"], size=11)

    ws4.freeze_panes = "A5"

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 5 — Raw Infor Data
    # ══════════════════════════════════════════════════════════════════════════
    ws_r = wb.create_sheet("Raw Infor Data")
    ws_r.sheet_view.showGridLines = False

    ws_r.merge_cells(f"A1:{get_column_letter(len(xl_df.columns))}1")
    c = ws_r["A1"]
    c.value = f"RAW INFOR DATA \u2014 source: {xl_filename}"
    c.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    c.fill = fill(C["INFOR_HDR"]); c.alignment = left(); ws_r.row_dimensions[1].height = 20

    ws_r.merge_cells(f"A2:{get_column_letter(len(xl_df.columns))}2")
    c = ws_r["A2"]
    c.value = "Semua kolom asli dari file Infor Excel \u2014 tidak ada modifikasi"
    c.font = Font(name="Calibri", italic=True, size=8, color=C["INFO_FG"])
    c.fill = fill(C["INFOR_BG"]); c.alignment = left(); ws_r.row_dimensions[2].height = 13

    for ci, col in enumerate(xl_df.columns, 1):
        wc(ws_r, 3, ci, col, bg=C["INFOR_HDR"], fnt=hfont(sz=9), aln=center())
    ws_r.row_dimensions[3].height = 18

    for ri, row in xl_df.iterrows():
        for ci, v in enumerate(row, 1):
            wc(ws_r, ri + 4, ci, v, fnt=cfont(sz=8), aln=left(),
               bg=C["INFOR_BG"] if ri % 2 == 0 else C["WHITE"])

    for ci, col in enumerate(xl_df.columns, 1):
        ws_r.column_dimensions[get_column_letter(ci)].width = max(10, min(28, len(str(col)) + 2))
    ws_r.freeze_panes = "A4"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


# ── Streamlit UI ───────────────────────────────────────────────────────────────
st.markdown(
    "<style>.stApp{font-family:Arial,sans-serif}.block-container{padding-top:1.2rem}</style>",
    unsafe_allow_html=True,
)

st.title("\U0001f4e6 Infor vs SAP Carton \u2014 Field-by-Field PO Comparator")
st.caption(
    "**Compared:** Total Qty (Pairs) \u00b7 Qty per Size \u00b7 **Qty/CTN per Size (NEW)**  |  "
    "**Info only (not compared):** Article \u00b7 Model  |  "
    "**Size detail columns:** UK/US Size \u00b7 XL Line \u00b7 CTN Range \u00b7 CTNs \u00b7 Infor Qty/CTN \u00b7 SAP Qty/CTN"
)

col1, col2 = st.columns(2)
with col1:
    st.subheader("\U0001f4ca Infor File \u2014 Order List (Excel)")
    xl_file = st.file_uploader("Upload Excel (.xlsx / .xls)", type=["xlsx", "xls"])
with col2:
    st.subheader("\U0001f4c4 SAP Carton Form (PDF)")
    pdf_file = st.file_uploader("Upload PDF (multi-page)", type=["pdf"])

if xl_file and pdf_file:
    with st.spinner("Processing & matching data..."):
        xl_df    = load_excel(xl_file)
        pdf_data = parse_pdf(pdf_file, filename=pdf_file.name)
        xl_pos     = xl_df["Order #"].str.strip().unique().tolist()
        pdf_pos    = list(pdf_data.keys())
        all_pos    = sorted(set(xl_pos + pdf_pos))
        match_pos  = [p for p in pdf_pos if p in set(xl_pos)]
        xl_pos_set = set(xl_pos)

        # Detect Infor Qty/CTN column
        ctn_col_found = _find_ctn_qty_col(xl_df)

    if ctn_col_found:
        st.success(f"\u2705 Kolom Infor Qty/CTN ditemukan: **`{ctn_col_found}`**")
    else:
        st.warning(
            "\u26a0\ufe0f Kolom 'Carton Qty per Size' tidak ditemukan di file Infor. "
            "Comparison Qty/CTN akan dilewati. Pastikan nama kolom mengandung 'Carton Qty per Size'."
        )

    sum_rows = []
    for po in all_pos:
        df_f    = compare_po_fields(xl_df, pdf_data, po)
        total_f = len(df_f)
        n_m     = (df_f["Status"] == "\u2705 MATCH").sum()    if not df_f.empty else 0
        n_mm    = (df_f["Status"] == "\u274c MISMATCH").sum() if not df_f.empty else 0
        sap_src = pdf_data.get(po, {}).get("header", {}).get("Source File", pdf_file.name)
        if po not in pdf_data:     result = "\u26a0\ufe0f NO SAP DATA"
        elif po not in xl_pos_set: result = "\u26a0\ufe0f NO INFOR DATA"
        elif n_mm == 0:            result = "\u2705 ALL OK"
        else:                      result = f"\u274c {n_mm} ISSUE(S)"
        sum_rows.append({
            "PO Number":      po,
            "Infor File":     xl_file.name,
            "SAP File(s)":    sap_src,
            "Total Fields":   total_f,
            "\u2705 Match":   n_m,
            "\u274c Mismatch": n_mm,
            "Result":         result,
        })
    sum_df = pd.DataFrame(sum_rows)

    k1, k2, k3, k4, k5, k6 = st.columns(6)
    k1.metric("Matched POs",           len(match_pos))
    k2.metric("\u2705 All OK",          (sum_df["Result"] == "\u2705 ALL OK").sum())
    k3.metric("\u274c POs w/ Issues",   (sum_df["\u274c Mismatch"] > 0).sum())
    k4.metric("Total Field Mismatches", int(sum_df["\u274c Mismatch"].sum()))
    k5.metric("SAP only",              len([p for p in pdf_pos if p not in xl_pos_set]))
    k6.metric("Infor only",            len([p for p in xl_pos  if p not in set(pdf_pos)]))

    st.divider()
    st.subheader("\U0001f4cb PO Compare Summary \u2014 adidas Infor vs SAP Carton")
    st.info(
        "\u2139\ufe0f **Fields compared:** Total Qty (Pairs) + Qty per Size + **Qty/CTN per Size**.  "
        "Article & Model **not compared** \u2014 info only.",
        icon=None,
    )

    def _cr(val):
        if "ALL OK" in str(val): return "background-color:#c8f7c5;color:#1B5E20;font-weight:bold"
        if "ISSUE"  in str(val): return "background-color:#ffcdd2;color:#B71C1C;font-weight:bold"
        if "NO"     in str(val): return "background-color:#ffe0b2;color:#E65100;font-weight:bold"
        return ""

    def _cmm(val):
        if isinstance(val, (int, float)) and val > 0:
            return "color:#B71C1C;font-weight:bold"
        return "color:#1B5E20;font-weight:bold"

    st.dataframe(
        sum_df.style.map(_cr, subset=["Result"]).map(_cmm, subset=["\u274c Mismatch"]),
        use_container_width=True, hide_index=True,
        height=min(600, 80 + len(sum_df) * 36),
    )

    st.divider()
    tab1, tab2, tab3 = st.tabs([
        "\U0001f50d Field-by-Field Detail",
        "\U0001f4d0 Size Detail (All Columns)",
        "\u274c Discrepancies Only",
    ])

    with tab1:
        st.caption(
            "**Compared fields:** `Total Qty (Pairs)` + `Qty Size X` + `Qty/CTN Size X` per UK Size. "
            "Article & Model **not shown here**."
        )
        det_rows = []
        for po in all_pos:
            df_f = compare_po_fields(xl_df, pdf_data, po)
            if df_f.empty: continue
            df_f.insert(0, "PO Number", po)
            det_rows.append(df_f)
        if det_rows:
            det_df = pd.concat(det_rows, ignore_index=True)

            def _cs(val):
                if val == "\u2705 MATCH":    return "background-color:#c8f7c5;color:#1B5E20;font-weight:bold"
                if val == "\u274c MISMATCH": return "background-color:#ffcdd2;color:#B71C1C;font-weight:bold"
                if "ONLY" in str(val):       return "background-color:#ffe0b2;color:#E65100;font-weight:bold"
                return ""

            def _cv(row):
                base = [""] * len(row); idx = list(row.index)
                if row.get("Status") == "\u274c MISMATCH":
                    for f in ["Infor Value", "SAP Value"]:
                        if f in idx:
                            base[idx.index(f)] = "background-color:#ffebee;color:#B71C1C;font-weight:bold"
                return base

            st.dataframe(
                det_df.style.map(_cs, subset=["Status"]).apply(_cv, axis=1),
                use_container_width=True, hide_index=True,
                height=min(700, 80 + len(det_df) * 34),
            )

    with tab2:
        st.caption(
            "**All columns per size row.** Article & Model = \U0001f535 info (not compared). "
            "XL Line + Infor Qty/CTN = Infor · CTN Range/CTNs/SAP Qty/CTN/SAP Qty = SAP · "
            "Diff/Status = Calculated."
        )
        sd_rows = []
        for po in all_pos:
            info = po_info(xl_df, pdf_data, po)
            sd   = compare_po_size_detail(xl_df, pdf_data, po)
            if sd.empty: continue
            sd.insert(0, "PO Number",       po)
            sd.insert(1, "Article (Infor)", info["Article (Infor)"])
            sd.insert(2, "Article (SAP)",   info["Article (SAP)"])
            sd.insert(3, "Model (Infor)",   info["Model (Infor)"])
            sd.insert(4, "Model (SAP)",     info["Model (SAP)"])
            sd_rows.append(sd)
        if sd_rows:
            sd_df = pd.concat(sd_rows, ignore_index=True)
            INFO_COLS = ["Article (Infor)", "Article (SAP)", "Model (Infor)", "Model (SAP)"]

            def _sst(val):
                if val == "\u2705 MATCH":    return "background-color:#c8f7c5;color:#1B5E20;font-weight:bold"
                if val == "\u274c MISMATCH": return "background-color:#ffcdd2;color:#B71C1C;font-weight:bold"
                if "ONLY" in str(val):       return "background-color:#ffe0b2;color:#E65100;font-weight:bold"
                return ""

            def _sdiff(val):
                try:
                    if int(val) != 0: return "color:#B71C1C;font-weight:bold"
                except: pass
                return ""

            def _info_col(val): return "background-color:#E3F2FD;color:#0D47A1"

            present_info = [c for c in INFO_COLS if c in sd_df.columns]
            status_cols  = [c for c in ["Status", "Status Qty/CTN"] if c in sd_df.columns]
            diff_cols    = [c for c in ["Diff", "Diff Qty/CTN"] if c in sd_df.columns]

            st.dataframe(
                sd_df.style
                    .map(_sst, subset=status_cols)
                    .map(_sdiff, subset=diff_cols)
                    .map(_info_col, subset=present_info),
                use_container_width=True, hide_index=True,
                height=min(700, 80 + len(sd_df) * 34),
            )
        else:
            st.info("No size data available.")

    with tab3:
        disc_rows = []
        for po in all_pos:
            df_f = compare_po_fields(xl_df, pdf_data, po)
            if df_f.empty: continue
            iss = df_f[df_f["Status"] != "\u2705 MATCH"].copy()
            if iss.empty: continue
            iss.insert(0, "PO Number", po)
            disc_rows.append(iss)
        if disc_rows:
            disc_df = pd.concat(disc_rows, ignore_index=True)

            def _ds(row):
                base = [""] * len(row); idx = list(row.index)
                for f in ["Infor Value", "SAP Value"]:
                    if f in idx:
                        base[idx.index(f)] = "background-color:#ffebee;color:#B71C1C;font-weight:bold"
                if "Status" in idx:
                    base[idx.index("Status")] = "background-color:#ffcdd2;color:#B71C1C;font-weight:bold"
                return base

            st.dataframe(
                disc_df.style.apply(_ds, axis=1),
                use_container_width=True, hide_index=True,
                height=min(500, 80 + len(disc_df) * 34),
            )
        else:
            st.success("\u2705 No discrepancies found \u2014 all matched POs are 100% OK!")

    st.divider()
    st.subheader("\u2b07\ufe0f Download Excel Report")
    with st.spinner("Building Excel report..."):
        report_buf = build_report(xl_df, pdf_data, all_pos, xl_file.name, pdf_file.name)
    fname = f"InforVsSAP_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    st.download_button(
        "\U0001f4e5 Download Excel Report",
        data=report_buf,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
    )

    pdf_only = [p for p in pdf_pos if p not in xl_pos_set]
    xl_only  = [p for p in xl_pos  if p not in set(pdf_pos)]
    if pdf_only:
        st.warning(f"\u26a0\ufe0f {len(pdf_only)} PO in SAP only: {', '.join(pdf_only)}")
    if xl_only:
        with st.expander(f"\u2139\ufe0f {len(xl_only)} PO in Infor only (not in SAP PDF)"):
            st.write(xl_only)

elif xl_file and not pdf_file:
    st.info("\U0001f4c4 Please upload the SAP Carton Form PDF.")
elif pdf_file and not xl_file:
    st.info("\U0001f4ca Please upload the Infor Order List Excel.")
else:
    st.info("\U0001f446 Upload both files above to start.")
    with st.expander("\u2139\ufe0f How it works"):
        st.markdown("""
        **Match key**: `Order #` (Infor) = `Cust.PO` (SAP)

        | Field | Compared? |
        |---|---|
        | Total Qty (Pairs) | ✅ Yes |
        | Qty Size X (per UK Size) | ✅ Yes |
        | **Qty/CTN Size X (per UK Size)** | ✅ **Yes (NEW)** |
        | Article | ❌ No — info only |
        | Model | ❌ No — info only |

        **Infor column for Qty/CTN**: any column whose name contains `Carton Qty per Size`
        (e.g. `VAS/SHAS L19 – Carton Qty per Size`, `VAS/SHAS L15 – Carton Qty per Size`).

        **SAP column for Qty/CTN**: `PER PRS` column parsed as `Qty/CTN`.

        **Excel sheets:** `PO Compare Summary` · `PO Compare Detail` · `Size Detail` · `Discrepancies Only` · `Raw Infor Data`

        **Size fix note:** Parser sekarang menangkap suffix `-K` dengan benar
        (contoh: `10-K`, `11-K`, `12-K`, `13-K`) sehingga merge Infor ↔ SAP akurat.
        """)
