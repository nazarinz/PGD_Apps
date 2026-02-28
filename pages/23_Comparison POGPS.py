# ============================================================
# PGD Comparison Tracking — SAP vs Infor  |  PO Splitter 5.000
# Supply Chain Data Integrity Control Tower v2.0
# ============================================================

import io
import sys
import re
import zipfile
from datetime import datetime, timedelta
from contextlib import nullcontext

import numpy as np
import pandas as pd
import streamlit as st

# ==== OpenPyXL (opsional, untuk ekspor Excel yang di-styling) ====
EXCEL_EXPORT_AVAILABLE = True
try:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:
    EXCEL_EXPORT_AVAILABLE = False

# ================== Streamlit Config ==================
st.set_page_config(
    page_title="PGD Control Tower — SAP vs Infor",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ================== GLOBAL CSS INJECTION ==================
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap');

/* ── Base ── */
html, body, [class*="css"] {
    font-family: 'DM Sans', sans-serif;
}

/* ── Hide Streamlit chrome ── */
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
.block-container {padding-top: 1.5rem; padding-bottom: 2rem;}

/* ── Executive Header Banner ── */
.ct-banner {
    background: linear-gradient(135deg, #0f172a 0%, #1e3a5f 60%, #0c4a6e 100%);
    border-radius: 12px;
    padding: 28px 36px;
    margin-bottom: 24px;
    border-left: 4px solid #38bdf8;
    position: relative;
    overflow: hidden;
}
.ct-banner::before {
    content: '';
    position: absolute;
    top: -40px; right: -40px;
    width: 180px; height: 180px;
    border-radius: 50%;
    background: rgba(56,189,248,0.06);
}
.ct-banner h1 {
    color: #f0f9ff;
    font-size: 1.6rem;
    font-weight: 700;
    margin: 0 0 4px 0;
    letter-spacing: -0.02em;
}
.ct-banner p {
    color: #7dd3fc;
    font-size: 0.82rem;
    margin: 0;
    font-weight: 400;
    letter-spacing: 0.04em;
    text-transform: uppercase;
}
.ct-badge {
    display: inline-block;
    background: #0ea5e9;
    color: white;
    font-size: 0.68rem;
    font-weight: 600;
    padding: 2px 8px;
    border-radius: 20px;
    letter-spacing: 0.06em;
    text-transform: uppercase;
    margin-left: 10px;
    vertical-align: middle;
}

/* ── KPI Cards ── */
.kpi-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
    gap: 12px;
    margin-bottom: 20px;
}
.kpi-card {
    background: #ffffff;
    border-radius: 10px;
    padding: 16px 18px;
    border-top: 3px solid #e2e8f0;
    box-shadow: 0 1px 6px rgba(0,0,0,0.07);
    transition: box-shadow .2s;
}
.kpi-card:hover { box-shadow: 0 4px 16px rgba(0,0,0,0.12); }
.kpi-card.green  { border-top-color: #22c55e; }
.kpi-card.yellow { border-top-color: #eab308; }
.kpi-card.red    { border-top-color: #ef4444; }
.kpi-card.blue   { border-top-color: #3b82f6; }
.kpi-card.purple { border-top-color: #a855f7; }
.kpi-label {
    font-size: 0.68rem;
    font-weight: 600;
    color: #64748b;
    text-transform: uppercase;
    letter-spacing: 0.06em;
    margin-bottom: 6px;
}
.kpi-value {
    font-size: 1.75rem;
    font-weight: 700;
    color: #0f172a;
    line-height: 1;
    font-family: 'DM Mono', monospace;
}
.kpi-sub {
    font-size: 0.72rem;
    color: #94a3b8;
    margin-top: 4px;
}
.kpi-delta-up   { color: #22c55e; font-size: 0.75rem; font-weight: 600; }
.kpi-delta-down { color: #ef4444; font-size: 0.75rem; font-weight: 600; }
.kpi-delta-neu  { color: #64748b; font-size: 0.75rem; font-weight: 600; }

/* ── Section Headers ── */
.section-header {
    display: flex;
    align-items: center;
    gap: 10px;
    margin: 28px 0 14px 0;
    padding-bottom: 10px;
    border-bottom: 1px solid #e2e8f0;
}
.section-header .sh-icon {
    width: 32px; height: 32px;
    border-radius: 8px;
    display: flex; align-items: center; justify-content: center;
    font-size: 1rem;
}
.section-header h2 {
    font-size: 1.05rem;
    font-weight: 700;
    color: #0f172a;
    margin: 0;
}
.section-header span.sub {
    font-size: 0.78rem;
    color: #64748b;
    margin-left: 6px;
    font-weight: 400;
}

/* ── Heatmap table ── */
.heatmap-table { width: 100%; border-collapse: collapse; font-size: 0.82rem; }
.heatmap-table th {
    background: #f8fafc;
    color: #475569;
    font-weight: 600;
    padding: 9px 12px;
    text-align: left;
    border-bottom: 2px solid #e2e8f0;
    font-size: 0.72rem;
    text-transform: uppercase;
    letter-spacing: 0.05em;
}
.heatmap-table td {
    padding: 8px 12px;
    border-bottom: 1px solid #f1f5f9;
    color: #334155;
}
.heatmap-table tr:hover td { background: #f8fafc; }
.bar-track {
    background: #e2e8f0;
    border-radius: 4px;
    height: 8px;
    width: 120px;
    display: inline-block;
    vertical-align: middle;
    margin-right: 8px;
}
.bar-fill {
    height: 8px;
    border-radius: 4px;
    display: inline-block;
}

/* ── Risk badges ── */
.risk-high   { background:#fee2e2; color:#b91c1c; padding:2px 10px; border-radius:20px; font-size:0.72rem; font-weight:700; }
.risk-medium { background:#fef3c7; color:#92400e; padding:2px 10px; border-radius:20px; font-size:0.72rem; font-weight:700; }
.risk-low    { background:#dcfce7; color:#166534; padding:2px 10px; border-radius:20px; font-size:0.72rem; font-weight:700; }

/* ── Insight cards ── */
.insight-card {
    background: #f8fafc;
    border-left: 3px solid #3b82f6;
    border-radius: 0 8px 8px 0;
    padding: 13px 16px;
    margin-bottom: 10px;
    font-size: 0.85rem;
    color: #1e293b;
    line-height: 1.55;
}
.insight-card.warn { border-left-color: #f59e0b; }
.insight-card.good { border-left-color: #22c55e; }
.insight-card.crit { border-left-color: #ef4444; background: #fff5f5; }
.insight-card strong { color: #0f172a; }

/* ── Divider ── */
.ct-divider {
    height: 1px;
    background: linear-gradient(90deg, #e2e8f0 0%, transparent 100%);
    margin: 24px 0;
}

/* ── Trend pill ── */
.trend-pill {
    display: inline-block;
    border-radius: 20px;
    padding: 2px 9px;
    font-size: 0.7rem;
    font-weight: 700;
    margin-left: 6px;
}
.trend-up   { background: #fee2e2; color: #b91c1c; }
.trend-down { background: #dcfce7; color: #166534; }
.trend-same { background: #e2e8f0; color: #475569; }
</style>
""", unsafe_allow_html=True)

# ================== Warna, Kolom, Format ==================
INFOR_COLOR  = "FFF9F16D"
RESULT_COLOR = "FFC6EFCE"
OTHER_COLOR  = "FFD9D9D9"
DATE_FMT     = "m/d/yyyy"

INFOR_COLUMNS_FIXED = [
    "Order Status Infor","Infor Quantity","Infor Model Name","Infor Article No",
    "Infor Classification Code","Infor Delay/Early - Confirmation CRD",
    "Infor Delay - PO PSDD Update","Infor Lead time","Infor GPS Country",
    "Infor Ship-to Country","Infor FPD","Infor LPD","Infor CRD","Infor PSDD",
    "Infor PODD","Infor PD","Infor Delay - PO PD Update",
    "Infor Shipment Method", "Infor Market PO Number",
]

BLANK_ON_EXPORT_COLUMNS = [
    "Delay/Early - Confirmation CRD","Infor Delay/Early - Confirmation CRD",
    "Result_Delay_CRD","Delay - PO PSDD Update","Infor Delay - PO PSDD Update",
    "Delay - PO PD Update","Infor Delay - PO PD Update","Shipment Method",
]

_NAN_STRINGS = {"NAN","NaN","nan","NULL","null","None","NONE","--","N/A","NAT","NAT"}

DATE_COLUMNS_PREF = [
    "Document Date","FPD","LPD","CRD","PSDD","FCR Date","PODD","PD","PO Date","Actual PGI",
    "Infor FPD","Infor LPD","Infor CRD","Infor PSDD","Infor PODD","Infor PD",
]

# ================== Country Name Normalization ==================
COUNTRY_NAME_MAP = {
    "USA":"UNITED STATES","U.S.A.":"UNITED STATES","US":"UNITED STATES",
    "UNITED STATES OF AMERICA":"UNITED STATES","UTD.ARAB EMIR.":"UNITED ARAB EMIRATES",
    "U.A.E.":"UNITED ARAB EMIRATES","UAE":"UNITED ARAB EMIRATES",
    "SOUTH KOREA":"KOREA","REPUBLIC OF KOREA":"KOREA","KOREA, REPUBLIC OF":"KOREA",
    "KOREA, SOUTH":"KOREA","HONG KONG-CHINA":"CHINA","HONG KONG":"CHINA","HK":"CHINA",
    "PEOPLES REP. OF CHINA":"CHINA","CHINA, PEOPLES REP.":"CHINA",
    "PEOPLE'S REPUBLIC OF CHINA":"CHINA","P.R. CHINA":"CHINA","MACAU":"CHINA","MACAO":"CHINA",
    "VIET NAM":"VIETNAM","VIET NAM, SOC. REP.":"VIETNAM",
    "PHILIPPINEN":"PHILIPPINES","PHILLIPINES":"PHILIPPINES",
    "TURKEY":"TURKIYE","TÜRKIYE":"TURKIYE","TURKEI":"TURKIYE",
    "GREAT BRITAIN":"UNITED KINGDOM","UK":"UNITED KINGDOM","ENGLAND":"UNITED KINGDOM",
    "CZECH REPUBLIC":"CZECHIA","CZECH REP.":"CZECHIA",
    "SAUDI-ARABIA":"SAUDI ARABIA","SAUDI ARABIEN":"SAUDI ARABIA",
}

def normalize_country(x):
    if pd.isna(x): return ""
    s = str(x).strip().upper()
    return COUNTRY_NAME_MAP.get(s, s)

def _vec_normalize_country(series: pd.Series) -> pd.Series:
    upper = series.fillna("").astype(str).str.strip().str.upper()
    return upper.map(COUNTRY_NAME_MAP).fillna(upper)

# ================== Helpers (Umum) ==================
def today_str_id():
    return (datetime.utcnow() + timedelta(hours=7)).strftime("%Y%m%d")

def status_ctx(label="Processing...", expanded=True):
    if hasattr(st, "status"):
        return st.status(label, expanded=expanded)
    st.info(label)
    return nullcontext()

def _status_update(ctx, label=None, state=None):
    if hasattr(ctx, "update"):
        ctx.update(label=label, state=state)
    else:
        if state == "error":      st.error(label or "")
        elif state == "complete": st.success(label or "")
        else:                     st.info(label or "")

@st.cache_data(show_spinner=False)
def read_excel_file(file):
    return pd.read_excel(file, engine="openpyxl")

@st.cache_data(show_spinner=False)
def read_csv_file(file):
    for enc in ("utf-8","utf-8-sig","latin1"):
        try:
            file.seek(0)
            return pd.read_csv(file, encoding=enc)
        except Exception:
            continue
    file.seek(0)
    return pd.read_csv(file)

def convert_date_columns(df):
    date_cols = [
        'Document Date','FPD','LPD','CRD','PSDD','FCR Date','PODD','PD','PO Date','Actual PGI',
        'Infor CRD','Infor PD','Infor PSDD','Infor FPD','Infor LPD','Infor PODD',
    ]
    for col in date_cols:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')
    return df

def load_sap(sap_df):
    df = sap_df.copy()
    if "Quanity" in df.columns and "Quantity" not in df.columns:
        df.rename(columns={'Quanity': 'Quantity'}, inplace=True)
    if "PO No.(Full)" in df.columns:
        df["PO No.(Full)"] = df["PO No.(Full)"].astype(str).str.strip()
    df = convert_date_columns(df)
    return df

def load_infor_from_many_csv(csv_dfs):
    data_list = []
    required_cols = ['PO Statistical Delivery Date (PSDD)','Customer Request Date (CRD)','Line Aggregator']
    for i, df in enumerate(csv_dfs, start=1):
        df.columns = df.columns.str.strip()
        if all(col in df.columns for col in required_cols):
            data_list.append(df)
            st.success(f"Dibaca ✅ CSV ke-{i} (kolom wajib lengkap)")
        else:
            miss = [c for c in required_cols if c not in df.columns]
            st.warning(f"CSV ke-{i} dilewati ⚠️ (kolom wajib hilang: {miss})")
    if not data_list: return pd.DataFrame()
    return pd.concat(data_list, ignore_index=True)

_RE_NON_DIGIT = re.compile(r"\D")
_NAN_STRINGS_UPPER: frozenset = frozenset(s.upper() for s in _NAN_STRINGS)

def normalize_po(x):
    if pd.isna(x): return ""
    s = str(x).strip()
    if not s or s.lower() in ("nan","none","null"): return ""
    try: s = str(int(float(s)))
    except (ValueError, TypeError): pass
    digits = re.sub(r"\D","",s)
    return digits.zfill(10)

def _vec_normalize_po(series: pd.Series) -> pd.Series:
    s        = series.fillna("").astype(str).str.strip()
    bad_mask = s.str.lower().isin({"nan","none","null",""})
    numeric  = pd.to_numeric(s, errors="coerce")
    result   = pd.Series("", index=series.index, dtype=object)
    has_num = numeric.notna() & ~bad_mask
    if has_num.any():
        result.loc[has_num] = numeric.loc[has_num].astype("int64").astype(str).str.zfill(10)
    needs_re = ~has_num & ~bad_mask
    if needs_re.any():
        result.loc[needs_re] = s.loc[needs_re].str.replace(_RE_NON_DIGIT,"",regex=True).str.zfill(10)
    return result

def normalize_market_po(x):
    if pd.isna(x): return ""
    s = str(x).strip()
    if not s or s.lower() in ("nan","none","null",""): return ""
    try: s = str(int(float(s)))
    except (ValueError, TypeError): pass
    digits = re.sub(r"\D","",s)
    if not digits: return ""
    if all(c == "0" for c in digits): return ""
    return digits.zfill(10)

def _vec_normalize_market_po(series: pd.Series) -> pd.Series:
    s        = series.fillna("").astype(str).str.strip()
    bad_mask = s.str.lower().isin({"nan","none","null",""})
    numeric  = pd.to_numeric(s, errors="coerce")
    result   = pd.Series("", index=series.index, dtype=object)
    has_num   = numeric.notna() & ~bad_mask
    all_zeros = numeric.notna() & (numeric == 0)
    valid_num = has_num & ~all_zeros
    if valid_num.any():
        result.loc[valid_num] = numeric.loc[valid_num].astype("int64").astype(str).str.zfill(10)
    needs_re = ~has_num & ~bad_mask
    if needs_re.any():
        digits = s.loc[needs_re].str.replace(_RE_NON_DIGIT,"",regex=True)
        is_zero_re = digits.str.match(r"^0*$")
        non_zero   = ~is_zero_re
        if non_zero.any():
            result.loc[digits[non_zero].index] = digits[non_zero].str.zfill(10)
    return result

# ================== Proses Infor PO Level ==================
def process_infor_po_level(df_all):
    df_all = df_all.copy()
    df_all.columns = df_all.columns.str.strip()
    selected_columns = [
        'Order #','Order Status','Model Name','Article Number',
        'Gps Customer Number','Country/Region',
        'Customer Request Date (CRD)','Plan Date',
        'PO Statistical Delivery Date (PSDD)',
        'First Production Date','Last Production Date',
        'PODD','Production Lead Time','Class Code',
        'Delay - Confirmation','Delay - PO Del Update',
        'Delivery Delay Pd','Quantity','Shipment Method','Market PO Number',
    ]
    missing = [c for c in selected_columns if c not in df_all.columns]
    if missing:
        st.error(f"Kolom Infor hilang: {missing}")
        return pd.DataFrame()
    df = df_all[selected_columns].copy()
    df["Order #"] = _vec_normalize_po(df["Order #"])
    df_po = (
        df.groupby("Order #", as_index=False)
        .agg({
            'Order Status':'first','Model Name':'first','Article Number':'first',
            'Gps Customer Number':'first','Country/Region':'first',
            'Customer Request Date (CRD)':'first','Plan Date':'first',
            'PO Statistical Delivery Date (PSDD)':'first',
            'First Production Date':'first','Last Production Date':'first',
            'PODD':'first','Production Lead Time':'first','Class Code':'first',
            'Delay - Confirmation':'first','Delay - PO Del Update':'first',
            'Delivery Delay Pd':'first','Quantity':'sum',
            'Shipment Method':'first','Market PO Number':'first',
        })
    )
    df_po.rename(columns={
        'Order Status':'Order Status Infor','Model Name':'Infor Model Name',
        'Article Number':'Infor Article No','Gps Customer Number':'Infor GPS Country',
        'Country/Region':'Infor Ship-to Country',
        'Customer Request Date (CRD)':'Infor CRD','Plan Date':'Infor PD',
        'PO Statistical Delivery Date (PSDD)':'Infor PSDD',
        'First Production Date':'Infor FPD','Last Production Date':'Infor LPD',
        'PODD':'Infor PODD','Production Lead Time':'Infor Lead time',
        'Class Code':'Infor Classification Code',
        'Delay - Confirmation':'Infor Delay/Early - Confirmation CRD',
        'Delay - PO Del Update':'Infor Delay - PO PSDD Update',
        'Delivery Delay Pd':'Infor Delay - PO PD Update',
        'Quantity':'Infor Quantity','Shipment Method':'Infor Shipment Method',
        'Market PO Number':'Infor Market PO Number',
    }, inplace=True)
    return convert_date_columns(df_po)

# ================== Fill Missing Dates ==================
def fill_missing_dates(df):
    df['Order Status Infor'] = (
        df.get('Order Status Infor', pd.Series(dtype=str))
        .astype(str).str.strip().str.upper()
    )
    for col in ['LPD','FPD','CRD','PD','PSDD','PODD']:
        if col not in df.columns:
            df[col] = pd.NaT
        elif not pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = pd.to_datetime(df[col], errors='coerce')
    mask_open = df['Order Status Infor'].eq('OPEN')
    min_dates = df[['CRD','PD']].min(axis=1)
    df.loc[mask_open & df['LPD'].isna(),'LPD']  = min_dates
    df.loc[mask_open & df['FPD'].isna(),'FPD']  = min_dates
    df.loc[mask_open & df['PSDD'].isna(),'PSDD'] = df['CRD']
    df.loc[mask_open & df['PODD'].isna(),'PODD'] = df['CRD']
    return df

# ================== Clean & Compare ==================
def clean_and_compare(df_merged):
    for col in ["Quantity","Infor Quantity","Production Lead Time","Infor Lead time","Article Lead time"]:
        if col in df_merged.columns:
            df_merged[col] = pd.to_numeric(df_merged[col], errors="coerce").fillna(0).round(2)

    code_mapping = {
        '161':'01-0161','84':'03-0084','68':'02-0068','64':'04-0064','62':'02-0062',
        '61':'01-0061','51':'03-0051','46':'03-0046','7':'02-0007','3':'03-0003',
        '2':'01-0002','1':'01-0001','4':'04-0004','8':'02-0008','10':'04-0010',
        '49':'03-0049','90':'04-0090','63':'03-0063','27':'04-0027',
    }

    def _vec_map_code(series: pd.Series) -> pd.Series:
        cleaned  = series.replace(["--","N/A","NULL"], pd.NA)
        numeric  = pd.to_numeric(cleaned, errors="coerce")
        has_num  = numeric.notna()
        int_keys = numeric.where(has_num).dropna().astype("int64").astype(str)
        mapped   = int_keys.map(code_mapping)
        result   = cleaned.copy()
        result.loc[mapped.dropna().index] = mapped.dropna()
        return result

    for col in ["Infor Delay/Early - Confirmation CRD","Infor Delay - PO PSDD Update","Infor Delay - PO PD Update"]:
        if col in df_merged.columns:
            df_merged[col] = _vec_map_code(df_merged[col])

    string_cols = [
        "Model Name","Infor Model Name","Article No","Infor Article No",
        "Classification Code","Infor Classification Code","Infor Ship-to Country",
        "Ship-to-Sort1","Infor GPS Country",
        "Delay/Early - Confirmation CRD","Infor Delay/Early - Confirmation CRD",
        "Delay - PO PSDD Update","Infor Delay - PO PSDD Update",
        "Delay - PO PD Update","Infor Delay - PO PD Update",
        "Shipment Method","Infor Shipment Method",
    ]
    for col in string_cols:
        if col in df_merged.columns:
            df_merged[col] = df_merged[col].astype(str).str.strip().str.upper()

    if "Ship-to Country" in df_merged.columns:
        df_merged["Ship-to Country"] = _vec_normalize_country(df_merged["Ship-to Country"])
    if "Ship-to-Sort1" in df_merged.columns:
        df_merged["Ship-to-Sort1"] = df_merged["Ship-to-Sort1"].astype(str).str.replace(".0","",regex=False)
    if "Infor GPS Country" in df_merged.columns:
        df_merged["Infor GPS Country"] = df_merged["Infor GPS Country"].astype(str).str.replace(".0","",regex=False)
    if "Cust Ord No" in df_merged.columns:
        df_merged["Cust Ord No"] = _vec_normalize_market_po(df_merged["Cust Ord No"])
    if "Infor Market PO Number" in df_merged.columns:
        df_merged["Infor Market PO Number"] = _vec_normalize_market_po(df_merged["Infor Market PO Number"])

    nan_clean_cols = [
        "Shipment Method","Infor Shipment Method","Delay - PO PSDD Update","Infor Delay - PO PSDD Update",
        "Delay - PO PD Update","Infor Delay - PO PD Update",
        "Delay/Early - Confirmation CRD","Infor Delay/Early - Confirmation CRD",
        "Ship-to Country","Infor Ship-to Country","Ship-to-Sort1","Infor GPS Country",
        "Model Name","Infor Model Name","Article No","Infor Article No",
        "Classification Code","Infor Classification Code","Cust Ord No","Infor Market PO Number",
    ]
    for col in nan_clean_cols:
        if col in df_merged.columns:
            df_merged[col] = df_merged[col].where(~df_merged[col].isin(_NAN_STRINGS_UPPER),"")

    cols_set = frozenset(df_merged.columns)
    n = len(df_merged)

    def safe_result(c1, c2):
        if c1 in cols_set and c2 in cols_set:
            return np.where(df_merged[c1] == df_merged[c2], "TRUE", "FALSE")
        return np.full(n, "", dtype=object)

    df_merged["Result_Quantity"]            = safe_result("Quantity",                       "Infor Quantity")
    df_merged["Result_Model Name"]          = safe_result("Model Name",                     "Infor Model Name")
    df_merged["Result_Article No"]          = safe_result("Article No",                     "Infor Article No")
    df_merged["Result_Classification Code"] = safe_result("Classification Code",            "Infor Classification Code")
    df_merged["Result_Delay_CRD"]           = safe_result("Delay/Early - Confirmation CRD","Infor Delay/Early - Confirmation CRD")
    df_merged["Result_Delay_PSDD"]          = safe_result("Delay - PO PSDD Update",         "Infor Delay - PO PSDD Update")
    df_merged["Result_Delay_PD"]            = safe_result("Delay - PO PD Update",           "Infor Delay - PO PD Update")
    df_merged["Result_Lead Time"]           = safe_result("Article Lead time",              "Infor Lead time")
    df_merged["Result_Country"]             = safe_result("Ship-to Country",                "Infor Ship-to Country")
    df_merged["Result_Sort1"]               = safe_result("Ship-to-Sort1",                  "Infor GPS Country")
    df_merged["Result_FPD"]                 = safe_result("FPD",                            "Infor FPD")
    df_merged["Result_LPD"]                 = safe_result("LPD",                            "Infor LPD")
    df_merged["Result_CRD"]                 = safe_result("CRD",                            "Infor CRD")
    df_merged["Result_PSDD"]                = safe_result("PSDD",                           "Infor PSDD")
    df_merged["Result_PODD"]                = safe_result("PODD",                           "Infor PODD")
    df_merged["Result_PD"]                  = safe_result("PD",                             "Infor PD")
    df_merged["Result_Market PO"]           = safe_result("Cust Ord No",                    "Infor Market PO Number")
    df_merged["Result_Shipment Method"]     = safe_result("Shipment Method",                "Infor Shipment Method")

    for res_col, c1, c2 in [
        ("Result_Shipment Method","Shipment Method","Infor Shipment Method"),
        ("Result_Delay_PSDD","Delay - PO PSDD Update","Infor Delay - PO PSDD Update"),
        ("Result_Delay_PD","Delay - PO PD Update","Infor Delay - PO PD Update"),
        ("Result_Delay_CRD","Delay/Early - Confirmation CRD","Infor Delay/Early - Confirmation CRD"),
    ]:
        if res_col in df_merged.columns and c1 in df_merged.columns and c2 in df_merged.columns:
            both_empty = (df_merged[c1].astype(str).str.strip() == "") & \
                         (df_merged[c2].astype(str).str.strip() == "")
            df_merged.loc[both_empty, res_col] = ""

    return df_merged

# ================== Desired Column Order ==================
DESIRED_ORDER = [
    'Client No','Site','Brand FTY Name','SO','Order Type','Order Type Description',
    'PO No.(Full)','Customer PO item','Order Status Infor',
    'Cust Ord No','Infor Market PO Number','Result_Market PO',
    'PO No.(Short)','Merchandise Category 2',
    'Quantity','Infor Quantity','Result_Quantity',
    'Model Name','Infor Model Name','Result_Model Name',
    'Article No','Infor Article No','Result_Article No',
    'SAP Material','Pattern Code(Up.No.)','Model No','Outsole Mold',
    'Gender','Category 1','Category 2','Category 3','Unit Price',
    'Classification Code','Infor Classification Code','Result_Classification Code',
    'DRC','Delay/Early - Confirmation PD',
    'Delay/Early - Confirmation CRD','Infor Delay/Early - Confirmation CRD','Result_Delay_CRD',
    'MDP','PDP','SDP',
    'Article Lead time','Infor Lead time','Result_Lead Time',
    'Ship-to-Sort1','Infor GPS Country','Result_Sort1',
    'Ship-to Country','Infor Ship-to Country','Result_Country',
    'Ship to Name',
    'Shipment Method','Infor Shipment Method','Result_Shipment Method',
    'Delay - PO PSDD Update','Infor Delay - PO PSDD Update','Result_Delay_PSDD',
    'Delay - PO PD Update','Infor Delay - PO PD Update','Result_Delay_PD',
    'Document Date',
    'PODD','Infor PODD','Result_PODD',
    'LPD','Infor LPD','Result_LPD',
    'PSDD','Infor PSDD','Result_PSDD',
    'FPD','Infor FPD','Result_FPD',
    'CRD','Infor CRD','Result_CRD',
    'PD','Infor PD','Result_PD',
    'FCR Date','PO Date','Actual PGI','Segment','S&P LPD','Currency'
]

DESIRED_ORDER_SET = frozenset(DESIRED_ORDER)

def reorder_columns(df, desired_order):
    col_set  = set(df.columns)
    existing = [c for c in desired_order if c in col_set]
    tail     = [c for c in df.columns if c not in DESIRED_ORDER_SET]
    return df[existing + tail]

# ================== Build Report ==================
def build_report(df_sap, df_infor_raw):
    df_sap2 = load_sap(df_sap)
    df_sap2["PO No.(Full)"] = _vec_normalize_po(df_sap2["PO No.(Full)"])
    df_infor = process_infor_po_level(df_infor_raw)
    if df_infor.empty: return pd.DataFrame()
    df = df_sap2.merge(df_infor, how="left", left_on="PO No.(Full)", right_on="Order #")
    fill_missing_dates(df)
    clean_and_compare(df)
    return reorder_columns(df, DESIRED_ORDER)

# ================== Export Helpers ==================
def _blank_export_columns(df):
    out = df.copy()
    blank_vals = {
        np.nan:"", pd.NA:"", None:"",
        "NaN":"","NAN":"","nan":"","NULL":"","null":"","--":"","N/A":"","NAT":"",
        0:"",0.0:"","0":"",
    }
    for col in BLANK_ON_EXPORT_COLUMNS:
        if col in out.columns:
            out[col] = out[col].replace(blank_vals)
    return out

def _export_excel_styled(df, sheet_name="Report"):
    if not EXCEL_EXPORT_AVAILABLE:
        raise RuntimeError("Fitur ekspor Excel (styled) butuh 'openpyxl'")
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.fill = PatternFill(fill_type=None)
        header_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))[0]
        idx_by_name  = {c.value: i + 1 for i, c in enumerate(header_cells)}
        for cell in header_cells:
            col_name = str(cell.value)
            if col_name in INFOR_COLUMNS_FIXED:
                fill = PatternFill("solid", fgColor=INFOR_COLOR)
            elif col_name.startswith("Result_"):
                fill = PatternFill("solid", fgColor=RESULT_COLOR)
            else:
                fill = PatternFill("solid", fgColor=OTHER_COLOR)
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font      = Font(name="Calibri", size=9, bold=True)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.fill      = PatternFill(fill_type=None)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font      = Font(name="Calibri", size=9)
        for date_col in DATE_COLUMNS_PREF:
            if date_col in idx_by_name:
                cidx = idx_by_name[date_col]
                for r in range(2, ws.max_row + 1):
                    cell = ws.cell(row=r, column=cidx)
                    if cell.value not in ("", None):
                        cell.number_format = DATE_FMT
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            maxlen = 0
            for cell in ws[col_letter]:
                v = "" if cell.value is None else str(cell.value)
                maxlen = max(maxlen, len(v))
            ws.column_dimensions[col_letter].width = min(max(9, maxlen + 2), 40)
        ws.freeze_panes    = "A2"
        ws.auto_filter.ref = ws.dimensions
    bio.seek(0)
    return bio


# ══════════════════════════════════════════════════════════════════════════
# ███  INSIGHT ENGINE — all vectorized, no backend logic changed  ███
# ══════════════════════════════════════════════════════════════════════════

# Canonical list of all Result columns + human-readable labels
ALL_RESULT_META = [
    ("Result_Quantity",            "Quantity"),
    ("Result_Model Name",          "Model Name"),
    ("Result_Article No",          "Article No"),
    ("Result_Classification Code", "Class Code"),
    ("Result_Lead Time",           "Lead Time"),
    ("Result_Sort1",               "GPS Country"),
    ("Result_Country",             "Ship-to Country"),
    ("Result_Shipment Method",     "Shipment Method"),
    ("Result_Market PO",           "Market PO"),
    ("Result_Delay_CRD",           "Delay / CRD"),
    ("Result_Delay_PSDD",          "Delay / PSDD"),
    ("Result_Delay_PD",            "Delay / PD"),
    ("Result_FPD",                 "FPD"),
    ("Result_LPD",                 "LPD"),
    ("Result_CRD",                 "CRD"),
    ("Result_PSDD",                "PSDD"),
    ("Result_PODD",                "PODD"),
    ("Result_PD",                  "PD"),
]
CRITICAL_FIELDS = {"Result_CRD","Result_PSDD","Result_LPD","Result_PODD"}


def _build_match_stats(df: pd.DataFrame) -> pd.DataFrame:
    """Vectorized: compute TRUE/FALSE/blank per Result col for given df slice."""
    rows = []
    for col, label in ALL_RESULT_META:
        if col not in df.columns:
            continue
        s          = df[col]
        true_cnt   = int(s.eq("TRUE").sum())
        false_cnt  = int(s.eq("FALSE").sum())
        blank_cnt  = int((~s.isin(["TRUE","FALSE"])).sum())
        evaluated  = true_cnt + false_cnt
        match_pct  = round(true_cnt  / evaluated * 100, 1) if evaluated > 0 else 0.0
        false_pct  = round(false_cnt / evaluated * 100, 1) if evaluated > 0 else 0.0
        rows.append({
            "col":       col,
            "label":     label,
            "true":      true_cnt,
            "false":     false_cnt,
            "blank":     blank_cnt,
            "evaluated": evaluated,
            "match_pct": match_pct,
            "false_pct": false_pct,
            "critical":  col in CRITICAL_FIELDS,
        })
    return pd.DataFrame(rows)


def _mismatch_count_per_po(df: pd.DataFrame) -> pd.DataFrame:
    """Per-PO mismatch count across all Result cols. Vectorized sum."""
    existing = [c for c, _ in ALL_RESULT_META if c in df.columns]
    if not existing or "PO No.(Full)" not in df.columns:
        return pd.DataFrame()
    bool_matrix = (df[existing] == "FALSE").astype(int)
    df2 = df[["PO No.(Full)","Order Status Infor"]].copy()
    df2["Mismatch_Count"] = bool_matrix.sum(axis=1)
    return (
        df2[df2["Mismatch_Count"] > 0]
        .sort_values("Mismatch_Count", ascending=False)
        .drop_duplicates("PO No.(Full)")
        .reset_index(drop=True)
    )


def _risk_score(df: pd.DataFrame) -> pd.DataFrame:
    """
    Rule-based risk engine (all vectorized):
      HIGH   → OPEN/UNCONFIRMED + any date mismatch (CRD/PSDD/PODD/LPD)
      MEDIUM → OPEN/UNCONFIRMED + non-date mismatch only
      LOW    → other statuses with mismatch
    """
    date_result_cols = [c for c in ["Result_CRD","Result_PSDD","Result_PODD","Result_LPD"] if c in df.columns]
    other_result_cols = [
        c for c, _ in ALL_RESULT_META
        if c in df.columns and c not in date_result_cols
    ]
    if "Order Status Infor" not in df.columns:
        return pd.DataFrame()

    status_upper = df["Order Status Infor"].astype(str).str.upper()
    is_active    = status_upper.isin(["OPEN","UNCONFIRMED"])

    has_date_mismatch  = (df[date_result_cols] == "FALSE").any(axis=1) if date_result_cols else pd.Series(False, index=df.index)
    has_other_mismatch = (df[other_result_cols] == "FALSE").any(axis=1) if other_result_cols else pd.Series(False, index=df.index)

    conditions = [
        is_active & has_date_mismatch,
        is_active & ~has_date_mismatch & has_other_mismatch,
        ~is_active & (has_date_mismatch | has_other_mismatch),
    ]
    choices = ["HIGH", "MEDIUM", "LOW"]

    df2 = df[["PO No.(Full)","Order Status Infor"]].copy() if "PO No.(Full)" in df.columns else df[["Order Status Infor"]].copy()
    df2["Risk"]          = np.select(conditions, choices, default="NONE")
    df2["Date_Mismatch"] = has_date_mismatch
    # Add mismatch count
    all_result = [c for c, _ in ALL_RESULT_META if c in df.columns]
    df2["Mismatch_Count"] = (df[all_result] == "FALSE").sum(axis=1) if all_result else 0
    return df2


# ─────────────────────────────────────────────────────────────
#  1. EXECUTIVE SUMMARY BAR
# ─────────────────────────────────────────────────────────────
def render_executive_summary(df_view: pd.DataFrame):
    """Top-of-page KPI bar — color-coded, business readable."""

    stats = _build_match_stats(df_view)
    total_po = len(df_view)

    # % fully matched PO (every evaluated Result col is TRUE)
    result_cols_present = [c for c, _ in ALL_RESULT_META if c in df_view.columns]
    if result_cols_present:
        has_any_false    = (df_view[result_cols_present] == "FALSE").any(axis=1)
        fully_matched    = int((~has_any_false).sum())
        mismatch_po      = int(has_any_false.sum())
        pct_matched      = round(fully_matched / total_po * 100, 1) if total_po > 0 else 0.0
        pct_mismatch     = round(mismatch_po   / total_po * 100, 1) if total_po > 0 else 0.0
    else:
        fully_matched = mismatch_po = 0
        pct_matched = pct_mismatch = 0.0

    # Critical fields accuracy (average match_pct of critical fields)
    crit_stats = stats[stats["critical"] == True]
    crit_acc   = round(crit_stats["match_pct"].mean(), 1) if not crit_stats.empty else 0.0

    # Highest error field
    if not stats.empty:
        worst_row    = stats.sort_values("false_pct", ascending=False).iloc[0]
        worst_field  = worst_row["label"]
        worst_fpct   = worst_row["false_pct"]
    else:
        worst_field, worst_fpct = "—", 0.0

    # Open orders with mismatch
    if "Order Status Infor" in df_view.columns and result_cols_present:
        open_mask      = df_view["Order Status Infor"].astype(str).str.upper() == "OPEN"
        open_mismatch  = int((df_view.loc[open_mask, result_cols_present] == "FALSE").any(axis=1).sum())
        total_open     = int(open_mask.sum())
    else:
        open_mismatch = total_open = 0

    # Operational risk score (0-100): weighted average of critical false%
    risk_weight = {"Result_CRD":3,"Result_PSDD":3,"Result_LPD":2,"Result_PODD":2}
    ors_num = ors_den = 0
    for rc, w in risk_weight.items():
        row = stats[stats["col"] == rc]
        if not row.empty:
            ors_num += row.iloc[0]["false_pct"] * w
            ors_den += w
    op_risk_score = round(ors_num / ors_den, 1) if ors_den > 0 else 0.0

    # Color helpers
    def kpi_color_match(pct):
        if pct >= 90: return "green"
        if pct >= 70: return "yellow"
        return "red"

    def kpi_color_false(pct):
        if pct <= 5:  return "green"
        if pct <= 25: return "yellow"
        return "red"

    def kpi_color_risk(score):
        if score <= 10: return "green"
        if score <= 35: return "yellow"
        return "red"

    cards = [
        {
            "label": "Total PO Compared",
            "value": f"{total_po:,}",
            "sub":   "active in filter",
            "color": "blue",
            "delta": None,
        },
        {
            "label": "Fully Matched PO",
            "value": f"{pct_matched}%",
            "sub":   f"{fully_matched:,} PO all fields correct",
            "color": kpi_color_match(pct_matched),
            "delta": f"{fully_matched:,} PO",
        },
        {
            "label": "Mismatch PO",
            "value": f"{pct_mismatch}%",
            "sub":   f"{mismatch_po:,} PO with ≥1 mismatch",
            "color": kpi_color_false(pct_mismatch),
            "delta": f"{mismatch_po:,} PO",
        },
        {
            "label": "Critical Fields Accuracy",
            "value": f"{crit_acc}%",
            "sub":   "CRD · PSDD · LPD · PODD avg",
            "color": kpi_color_match(crit_acc),
            "delta": None,
        },
        {
            "label": "Highest Error Field",
            "value": worst_field,
            "sub":   f"{worst_fpct}% FALSE rate",
            "color": kpi_color_false(worst_fpct),
            "delta": f"{worst_fpct}%",
        },
        {
            "label": "OPEN w/ Mismatch",
            "value": f"{open_mismatch:,}",
            "sub":   f"of {total_open:,} total OPEN orders",
            "color": kpi_color_false((open_mismatch/total_open*100) if total_open else 0),
            "delta": None,
        },
        {
            "label": "Operational Risk Score",
            "value": f"{op_risk_score}",
            "sub":   "0 = safe · 100 = critical",
            "color": kpi_color_risk(op_risk_score),
            "delta": None,
        },
    ]

    html_cards = ""
    for c in cards:
        delta_html = ""
        if c["delta"]:
            delta_html = f'<div class="kpi-delta-neu">{c["delta"]}</div>'
        html_cards += f"""
        <div class="kpi-card {c['color']}">
            <div class="kpi-label">{c['label']}</div>
            <div class="kpi-value">{c['value']}</div>
            <div class="kpi-sub">{c['sub']}</div>
            {delta_html}
        </div>"""

    st.markdown(f'<div class="kpi-grid">{html_cards}</div>', unsafe_allow_html=True)
    return stats  # reuse downstream


# ─────────────────────────────────────────────────────────────
#  2. DATA QUALITY HEATMAP
# ─────────────────────────────────────────────────────────────
def render_data_quality_heatmap(stats: pd.DataFrame):
    if stats.empty:
        return

    sh = """
    <div class="section-header">
        <div class="sh-icon" style="background:#eff6ff;">📊</div>
        <h2>Data Quality Heatmap<span class="sub">Match rate per field — sorted worst → best</span></h2>
    </div>"""
    st.markdown(sh, unsafe_allow_html=True)

    sorted_stats = stats.sort_values("match_pct", ascending=True).reset_index(drop=True)

    def bar_color(pct):
        if pct >= 90: return "#22c55e"
        if pct >= 70: return "#eab308"
        return "#ef4444"

    def rating(pct):
        if pct >= 95: return ("EXCELLENT", "#166534", "#dcfce7")
        if pct >= 85: return ("GOOD",      "#1e40af", "#dbeafe")
        if pct >= 70: return ("FAIR",      "#92400e", "#fef3c7")
        if pct >= 50: return ("POOR",      "#9a3412", "#ffedd5")
        return ("CRITICAL", "#b91c1c", "#fee2e2")

    rows_html = ""
    for _, row in sorted_stats.iterrows():
        pct   = row["match_pct"]
        fpct  = row["false_pct"]
        color = bar_color(pct)
        rat, rat_fg, rat_bg = rating(pct)
        crit_badge = ' <span style="background:#dbeafe;color:#1e40af;padding:1px 6px;border-radius:10px;font-size:0.65rem;font-weight:700;">CRITICAL</span>' if row["critical"] else ""
        rows_html += f"""
        <tr>
          <td><strong>{row['label']}</strong>{crit_badge}</td>
          <td>
            <span class="bar-track"><span class="bar-fill" style="width:{pct*1.2}px;max-width:120px;background:{color};display:inline-block;height:8px;border-radius:4px;"></span></span>
            <strong style="color:{color};font-family:'DM Mono',monospace;">{pct}%</strong>
          </td>
          <td style="color:#ef4444;font-family:'DM Mono',monospace;">{fpct}%</td>
          <td><strong style="color:#0f172a;font-family:'DM Mono',monospace;">{row['true']:,}</strong></td>
          <td style="color:#ef4444;font-family:'DM Mono',monospace;">{row['false']:,}</td>
          <td><span style="background:{rat_bg};color:{rat_fg};padding:2px 8px;border-radius:10px;font-size:0.7rem;font-weight:700;">{rat}</span></td>
        </tr>"""

    table_html = f"""
    <table class="heatmap-table">
      <thead>
        <tr>
          <th>Field</th>
          <th>Match Rate</th>
          <th>FALSE %</th>
          <th>TRUE Count</th>
          <th>FALSE Count</th>
          <th>Rating</th>
        </tr>
      </thead>
      <tbody>{rows_html}</tbody>
    </table>"""

    st.markdown(table_html, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────
#  3. MISMATCH ROOT CAUSE — TOP 10 PO
# ─────────────────────────────────────────────────────────────
def render_mismatch_root_cause(df_view: pd.DataFrame):
    sh = """
    <div class="section-header">
        <div class="sh-icon" style="background:#fff1f2;">🎯</div>
        <h2>Mismatch Root Cause<span class="sub">Top 10 PO by mismatch count</span></h2>
    </div>"""
    st.markdown(sh, unsafe_allow_html=True)

    mismatch_df = _mismatch_count_per_po(df_view)
    if mismatch_df.empty:
        st.success("✅ No mismatch found across all PO.")
        return

    top10 = mismatch_df.head(10).copy()

    # Enrich: which fields are FALSE per PO
    result_cols_present = [c for c, _ in ALL_RESULT_META if c in df_view.columns]
    label_map = {c: l for c, l in ALL_RESULT_META}

    if "PO No.(Full)" in df_view.columns:
        po_detail = df_view.set_index("PO No.(Full)")[result_cols_present]
        failed_fields = []
        for po in top10["PO No.(Full)"].values:
            if po in po_detail.index:
                row    = po_detail.loc[po]
                fields = [label_map.get(c, c) for c in result_cols_present if row[c] == "FALSE"]
                failed_fields.append(", ".join(fields[:6]) + ("…" if len(fields) > 6 else ""))
            else:
                failed_fields.append("—")
        top10["Failed Fields"] = failed_fields

    max_cnt = top10["Mismatch_Count"].max() or 1
    rows_html = ""
    for i, (_, r) in enumerate(top10.iterrows(), 1):
        pct   = r["Mismatch_Count"] / len(result_cols_present) * 100
        color = "#ef4444" if pct >= 50 else ("#f59e0b" if pct >= 25 else "#64748b")
        status = str(r.get("Order Status Infor","")).upper()
        sbadge = f'<span class="risk-high">{status}</span>' if status in ("OPEN","UNCONFIRMED") else \
                 f'<span class="risk-low">{status}</span>'
        bar_w = int(r["Mismatch_Count"] / max_cnt * 100)
        rows_html += f"""
        <tr>
          <td style="color:#64748b;font-size:0.75rem;">#{i}</td>
          <td><strong style="font-family:'DM Mono',monospace;">{r.get('PO No.(Full)','—')}</strong></td>
          <td>{sbadge}</td>
          <td>
            <div style="display:flex;align-items:center;gap:8px;">
              <div style="width:{bar_w}px;max-width:80px;height:6px;background:{color};border-radius:3px;"></div>
              <strong style="color:{color};font-family:'DM Mono',monospace;">{r['Mismatch_Count']}</strong>
            </div>
          </td>
          <td style="font-size:0.78rem;color:#475569;">{r.get('Failed Fields','—')}</td>
        </tr>"""

    table_html = f"""
    <table class="heatmap-table">
      <thead>
        <tr><th>#</th><th>PO Number</th><th>Status</th><th>Mismatch Count</th><th>Failed Fields</th></tr>
      </thead>
      <tbody>{rows_html}</tbody>
    </table>"""
    st.markdown(table_html, unsafe_allow_html=True)

    with st.expander("📋 View full mismatch list", expanded=False):
        st.dataframe(mismatch_df, use_container_width=True, hide_index=True)
        st.download_button(
            "⬇️ Download Mismatch List (CSV)",
            data=mismatch_df.to_csv(index=False).encode("utf-8"),
            file_name=f"mismatch_list_{today_str_id()}.csv",
            mime="text/csv",
        )


# ─────────────────────────────────────────────────────────────
#  4. OPERATIONAL RISK PANEL
# ─────────────────────────────────────────────────────────────
def render_operational_risk(df_view: pd.DataFrame):
    sh = """
    <div class="section-header">
        <div class="sh-icon" style="background:#fff7ed;">⚠️</div>
        <h2>Operational Risk Panel<span class="sub">Rule-based risk detection — OPEN/UNCONFIRMED orders</span></h2>
    </div>"""
    st.markdown(sh, unsafe_allow_html=True)

    risk_df = _risk_score(df_view)
    if risk_df.empty:
        st.info("Risk analysis requires 'Order Status Infor' column.")
        return

    dist = risk_df["Risk"].value_counts()
    high   = int(dist.get("HIGH",   0))
    medium = int(dist.get("MEDIUM", 0))
    low    = int(dist.get("LOW",    0))
    none_  = int(dist.get("NONE",   0))
    total  = len(risk_df)

    # KPI mini row
    c1, c2, c3, c4 = st.columns(4)
    c1.markdown(f"""
    <div class="kpi-card red">
        <div class="kpi-label">🔴 High Risk</div>
        <div class="kpi-value">{high}</div>
        <div class="kpi-sub">OPEN + date mismatch</div>
    </div>""", unsafe_allow_html=True)
    c2.markdown(f"""
    <div class="kpi-card yellow">
        <div class="kpi-label">🟡 Medium Risk</div>
        <div class="kpi-value">{medium}</div>
        <div class="kpi-sub">OPEN + non-date mismatch</div>
    </div>""", unsafe_allow_html=True)
    c3.markdown(f"""
    <div class="kpi-card green">
        <div class="kpi-label">🟢 Low / Closed</div>
        <div class="kpi-value">{low}</div>
        <div class="kpi-sub">closed status mismatch</div>
    </div>""", unsafe_allow_html=True)
    c4.markdown(f"""
    <div class="kpi-card blue">
        <div class="kpi-label">✅ No Risk</div>
        <div class="kpi-value">{none_}</div>
        <div class="kpi-sub">all fields match or irrelevant</div>
    </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Risk distribution bar
    if total > 0:
        bar_html = '<div style="display:flex;height:14px;border-radius:7px;overflow:hidden;margin-bottom:6px;">'
        for cnt, color in [(high,"#ef4444"),(medium,"#f59e0b"),(low,"#22c55e"),(none_,"#e2e8f0")]:
            w = cnt / total * 100
            if w > 0:
                bar_html += f'<div style="width:{w}%;background:{color};"></div>'
        bar_html += '</div>'
        bar_html += f'<div style="font-size:0.72rem;color:#64748b;">Risk distribution across {total:,} PO</div>'
        st.markdown(bar_html, unsafe_allow_html=True)

    # High-risk table
    high_risk = risk_df[risk_df["Risk"] == "HIGH"].head(50)
    if not high_risk.empty:
        with st.expander(f"🔴 HIGH RISK Orders — {high} PO (OPEN + date mismatch)", expanded=True):
            cols_show = [c for c in ["PO No.(Full)","Order Status Infor","Mismatch_Count","Date_Mismatch"] if c in high_risk.columns]
            st.dataframe(high_risk[cols_show].reset_index(drop=True), use_container_width=True, hide_index=True)
            st.download_button(
                "⬇️ Download High Risk PO (CSV)",
                data=high_risk[cols_show].to_csv(index=False).encode("utf-8"),
                file_name=f"high_risk_PO_{today_str_id()}.csv",
                mime="text/csv",
            )

    medium_risk = risk_df[risk_df["Risk"] == "MEDIUM"].head(50)
    if not medium_risk.empty:
        with st.expander(f"🟡 MEDIUM RISK Orders — {medium} PO", expanded=False):
            cols_show = [c for c in ["PO No.(Full)","Order Status Infor","Mismatch_Count"] if c in medium_risk.columns]
            st.dataframe(medium_risk[cols_show].reset_index(drop=True), use_container_width=True, hide_index=True)


# ─────────────────────────────────────────────────────────────
#  5. FIELD FAILURE TREND (session memory)
# ─────────────────────────────────────────────────────────────
def render_field_failure_trend(current_stats: pd.DataFrame):
    sh = """
    <div class="section-header">
        <div class="sh-icon" style="background:#f0fdf4;">📈</div>
        <h2>Field Failure Trend<span class="sub">Compare current vs previous upload (session memory)</span></h2>
    </div>"""
    st.markdown(sh, unsafe_allow_html=True)

    prev_stats = st.session_state.get("prev_stats_snapshot", None)

    if prev_stats is None:
        st.info("📌 Upload a second file or re-execute to compare against previous run. Snapshot saved for next comparison.")
        # Save current as baseline
        st.session_state["prev_stats_snapshot"] = current_stats.copy()
        return

    merged = current_stats.merge(
        prev_stats[["col","false_pct"]].rename(columns={"false_pct": "prev_false_pct"}),
        on="col", how="left"
    )
    merged["delta"] = (merged["false_pct"] - merged["prev_false_pct"]).round(1)
    merged["trend"] = merged["delta"].apply(
        lambda d: "▲" if d > 0.5 else ("▼" if d < -0.5 else "→")
    )

    rows_html = ""
    for _, r in merged.sort_values("delta", ascending=False).iterrows():
        d = r["delta"]
        if pd.isna(d):
            trend_html = '<span class="trend-pill trend-same">NEW</span>'
        elif d > 0.5:
            trend_html = f'<span class="trend-pill trend-up">▲ +{d}%</span>'
        elif d < -0.5:
            trend_html = f'<span class="trend-pill trend-down">▼ {d}%</span>'
        else:
            trend_html = f'<span class="trend-pill trend-same">→ {d}%</span>'

        prev_str = f"{r['prev_false_pct']}%" if not pd.isna(r.get('prev_false_pct',float('nan'))) else "—"
        rows_html += f"""
        <tr>
          <td><strong>{r['label']}</strong></td>
          <td style="font-family:'DM Mono',monospace;">{r['false_pct']}%</td>
          <td style="font-family:'DM Mono',monospace;">{prev_str}</td>
          <td>{trend_html}</td>
        </tr>"""

    table_html = f"""
    <table class="heatmap-table">
      <thead>
        <tr><th>Field</th><th>Current FALSE %</th><th>Previous FALSE %</th><th>Trend</th></tr>
      </thead>
      <tbody>{rows_html}</tbody>
    </table>"""
    st.markdown(table_html, unsafe_allow_html=True)

    if st.button("🔄 Set current as new baseline", key="set_baseline"):
        st.session_state["prev_stats_snapshot"] = current_stats.copy()
        st.success("Baseline updated.")


# ─────────────────────────────────────────────────────────────
#  6. SMART BUSINESS INSIGHTS (auto-generated narrative)
# ─────────────────────────────────────────────────────────────
def render_smart_insights(df_view: pd.DataFrame, stats: pd.DataFrame):
    sh = """
    <div class="section-header">
        <div class="sh-icon" style="background:#faf5ff;">💡</div>
        <h2>Smart Business Insights<span class="sub">Auto-generated from data — updated on every filter change</span></h2>
    </div>"""
    st.markdown(sh, unsafe_allow_html=True)

    if stats.empty:
        return

    total_po   = len(df_view)
    insights   = []

    # ── Sort for context ──
    worst       = stats.sort_values("false_pct", ascending=False)
    best        = stats.sort_values("false_pct", ascending=True)
    top_problem = worst.iloc[0]  if len(worst) > 0 else None
    top_good    = best.iloc[0]   if len(best)  > 0 else None

    # Open / mismatch context
    result_cols_present = [c for c, _ in ALL_RESULT_META if c in df_view.columns]
    if "Order Status Infor" in df_view.columns and result_cols_present:
        open_mask      = df_view["Order Status Infor"].astype(str).str.upper() == "OPEN"
        total_open     = int(open_mask.sum())
        open_mismatch  = int((df_view.loc[open_mask, result_cols_present] == "FALSE").any(axis=1).sum())
        open_pct       = round(open_mismatch / total_open * 100, 1) if total_open > 0 else 0
    else:
        total_open = open_mismatch = open_pct = 0

    # Critical field status
    crit_issues = stats[(stats["critical"] == True) & (stats["false_pct"] > 10)]

    # ── Generate insight strings ──
    if top_problem is not None and top_problem["false_pct"] > 20:
        lvl = "crit" if top_problem["false_pct"] > 50 else "warn"
        insights.append((
            lvl,
            f"<strong>{top_problem['label']}</strong> is the dominant mismatch field with "
            f"<strong>{top_problem['false_pct']}% FALSE rate</strong> ({top_problem['false']:,} records). "
            f"Investigate synchronization between SAP and Infor for this attribute immediately."
        ))

    if open_pct > 30:
        insights.append((
            "crit",
            f"<strong>{open_pct}% of OPEN orders ({open_mismatch:,} of {total_open:,})</strong> have at least one "
            f"field mismatch. This represents direct operational exposure — open orders cannot be "
            f"shipped or invoiced correctly until discrepancies are resolved."
        ))
    elif open_pct > 10:
        insights.append((
            "warn",
            f"<strong>{open_pct}% of OPEN orders</strong> ({open_mismatch:,} PO) carry data mismatches. "
            f"Prioritize resolution before order execution to avoid downstream delays."
        ))
    else:
        insights.append((
            "good",
            f"Only <strong>{open_pct}%</strong> of OPEN orders have mismatches. "
            f"Data quality for active orders is in good standing."
        ))

    if not crit_issues.empty:
        crit_names = ", ".join(crit_issues["label"].tolist())
        avg_err    = round(crit_issues["false_pct"].mean(), 1)
        insights.append((
            "crit",
            f"Critical date fields <strong>{crit_names}</strong> show average <strong>{avg_err}% error rate</strong>. "
            f"Date mismatches directly impact delivery promises, planning accuracy, and customer satisfaction."
        ))
    else:
        insights.append((
            "good",
            f"All critical date fields (CRD, PSDD, LPD, PODD) have <10% error rate. "
            f"Scheduling alignment between SAP and Infor is healthy."
        ))

    if top_good is not None and top_good["false_pct"] <= 2:
        insights.append((
            "good",
            f"<strong>{top_good['label']}</strong> is your most reliable field with only "
            f"<strong>{top_good['false_pct']}% FALSE rate</strong>. "
            f"This mapping/sync process can serve as a benchmark for other fields."
        ))

    # Check for systematic issues (multiple fields > 30%)
    systemic = stats[stats["false_pct"] > 30]
    if len(systemic) >= 3:
        insights.append((
            "crit",
            f"<strong>{len(systemic)} fields</strong> exceed 30% error rate simultaneously "
            f"({', '.join(systemic['label'].tolist()[:4])}{'…' if len(systemic)>4 else ''}). "
            f"This pattern suggests a <strong>systematic integration failure</strong>, not isolated data entry errors. "
            f"Escalate to integration / middleware team."
        ))

    # Mismatch volume
    total_false_events = int(stats["false"].sum())
    if total_false_events > 0:
        insights.append((
            "warn" if total_false_events > 500 else "good",
            f"There are <strong>{total_false_events:,} total field-level mismatches</strong> across "
            f"{total_po:,} PO in the current filter. "
            f"Each mismatch is a potential manual correction effort — resolving the top 3 fields by false count "
            f"would eliminate approximately {round(stats.nlargest(3,'false')['false'].sum()/total_false_events*100)}% of all issues."
        ))

    # Render
    for kind, text in insights:
        st.markdown(f'<div class="insight-card {kind}">{text}</div>', unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────
#  7. FALSE INSIGHT OPEN (detailed drill-down — kept from v1)
# ─────────────────────────────────────────────────────────────
def render_false_insight_open(df_view: pd.DataFrame):
    sh = """
    <div class="section-header">
        <div class="sh-icon" style="background:#fff1f2;">🔴</div>
        <h2>FALSE Detail Drill-Down<span class="sub">OPEN orders only — expandable per-field breakdown</span></h2>
    </div>"""
    st.markdown(sh, unsafe_allow_html=True)

    if "Order Status Infor" in df_view.columns:
        df_open = df_view[df_view["Order Status Infor"].astype(str).str.upper() == "OPEN"].copy()
    else:
        st.warning("Kolom 'Order Status Infor' tidak ditemukan.")
        return

    total_open = len(df_open)
    if total_open == 0:
        st.info("Tidak ada data dengan Order Status **OPEN** pada filter saat ini.")
        return

    existing_result_cols = [c for c, _ in ALL_RESULT_META if c in df_open.columns]
    if not existing_result_cols:
        return

    open_stats = _build_match_stats(df_open)

    # Mini KPI
    total_false_open = int(open_stats["false"].sum())
    worst_open       = open_stats.sort_values("false_pct", ascending=False).iloc[0] if not open_stats.empty else None
    c1, c2, c3 = st.columns(3)
    c1.metric("Total OPEN Orders", f"{total_open:,}")
    c2.metric("Total FALSE events (OPEN)", f"{total_false_open:,}")
    c3.metric("Worst field (OPEN)", f"{worst_open['label']} — {worst_open['false_pct']}%" if worst_open is not None else "—")

    # Field pair map
    field_pair_map = {
        "Result_Quantity":            ("Quantity",                       "Infor Quantity"),
        "Result_Model Name":          ("Model Name",                     "Infor Model Name"),
        "Result_Article No":          ("Article No",                     "Infor Article No"),
        "Result_Classification Code": ("Classification Code",            "Infor Classification Code"),
        "Result_Lead Time":           ("Article Lead time",              "Infor Lead time"),
        "Result_Sort1":               ("Ship-to-Sort1",                  "Infor GPS Country"),
        "Result_Country":             ("Ship-to Country",                "Infor Ship-to Country"),
        "Result_Shipment Method":     ("Shipment Method",                "Infor Shipment Method"),
        "Result_Market PO":           ("Cust Ord No",                    "Infor Market PO Number"),
        "Result_Delay_CRD":           ("Delay/Early - Confirmation CRD", "Infor Delay/Early - Confirmation CRD"),
        "Result_Delay_PSDD":          ("Delay - PO PSDD Update",         "Infor Delay - PO PSDD Update"),
        "Result_Delay_PD":            ("Delay - PO PD Update",           "Infor Delay - PO PD Update"),
        "Result_FPD":                 ("FPD",                            "Infor FPD"),
        "Result_LPD":                 ("LPD",                            "Infor LPD"),
        "Result_CRD":                 ("CRD",                            "Infor CRD"),
        "Result_PSDD":                ("PSDD",                           "Infor PSDD"),
        "Result_PODD":                ("PODD",                           "Infor PODD"),
        "Result_PD":                  ("PD",                             "Infor PD"),
    }
    id_cols = ["PO No.(Full)","Order Status Infor"]
    if "Cust Ord No" in df_open.columns:
        id_cols.append("Cust Ord No")

    for col, label in ALL_RESULT_META:
        if col not in df_open.columns:
            continue
        false_rows = df_open[df_open[col] == "FALSE"]
        cnt = len(false_rows)
        if cnt == 0:
            continue
        sap_col, infor_col = field_pair_map.get(col, (None, None))
        show_cols = [c for c in id_cols if c in false_rows.columns]
        if sap_col and sap_col in false_rows.columns:
            show_cols.append(sap_col)
        if infor_col and infor_col in false_rows.columns:
            show_cols.append(infor_col)
        show_cols.append(col)
        with st.expander(f"❌ {label} — {cnt:,} FALSE", expanded=False):
            st.dataframe(false_rows[show_cols].reset_index(drop=True), use_container_width=True, hide_index=True)
            st.download_button(
                label=f"⬇️ Download FALSE rows — {label}",
                data=false_rows[show_cols].to_csv(index=False).encode("utf-8"),
                file_name=f"FALSE_{label.replace(' ','_')}_{today_str_id()}.csv",
                mime="text/csv",
                key=f"dl_{col}",
            )


# ================== Helpers (PO Splitter) ==================
def parse_input(text: str, split_mode: str = "auto"):
    text = text.strip()
    if not text: return []
    if split_mode == "newline":     raw = text.splitlines()
    elif split_mode == "comma":     raw = text.split(",")
    elif split_mode == "semicolon": raw = text.split(";")
    elif split_mode == "whitespace":raw = re.split(r"\s+", text)
    else:
        if "\n" in text:
            raw = re.split(r"[\r\n]+", text)
            split_more = []
            for line in raw:
                line = line.strip()
                if not line: continue
                if ("," in line) or (";" in line):
                    split_more.extend(re.split(r"[,;]", line))
                else:
                    split_more.append(line)
            raw = split_more
        elif ("," in text) or (";" in text):
            raw = re.split(r"[,;]", text)
        else:
            raw = re.split(r"\s+", text)
    return [x.strip() for x in raw if str(x).strip() != ""]

def normalize_items(items, keep_only_digits=False, upper_case=False, strip_prefix_suffix=False):
    normed = []
    for it in items:
        s = str(it)
        if strip_prefix_suffix: s = re.sub(r"^\W+|\W+$", "", s)
        if keep_only_digits:    s = re.sub(r"\D+", "", s)
        if upper_case:          s = s.upper()
        s = s.strip()
        if s != "": normed.append(s)
    return normed

def chunk_list(items, size):
    return [items[i:i + size] for i in range(0, len(items), size)]

def to_txt_bytes(lines):
    buf = io.StringIO()
    for ln in lines: buf.write(f"{ln}\n")
    return buf.getvalue().encode("utf-8")

def df_from_list(items, col_name="PO"):
    return pd.DataFrame({col_name: items})

def make_zip_bytes(chunks, basename="chunk", as_csv=True, col_name="PO"):
    mem = io.BytesIO()
    with zipfile.ZipFile(mem, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for idx, part in enumerate(chunks, start=1):
            if as_csv:
                df        = df_from_list(part, col_name=col_name)
                csv_bytes = df.to_csv(index=False).encode("utf-8")
                zf.writestr(f"{basename}_{idx:02d}.csv", csv_bytes)
            else:
                zf.writestr(f"{basename}_{idx:02d}.txt", to_txt_bytes(part))
    mem.seek(0)
    return mem


# ══════════════════════════════════════════════════════════════════════
# ███  TABS
# ══════════════════════════════════════════════════════════════════════
tab1, tab2 = st.tabs(["📦 PGD Control Tower", "🧩 PO Splitter"])

# ════════════════════════════════════════════════════════════
#  TAB 1 — CONTROL TOWER
# ════════════════════════════════════════════════════════════
with tab1:

    # ── Executive header banner ──
    now_str = (datetime.utcnow() + timedelta(hours=7)).strftime("%d %b %Y, %H:%M WIB")
    st.markdown(f"""
    <div class="ct-banner">
        <h1>Supply Chain Data Integrity Control Tower
            <span class="ct-badge">LIVE</span>
        </h1>
        <p>SAP vs Infor PGD Comparison &nbsp;·&nbsp; Last refreshed: {now_str} &nbsp;·&nbsp; Data Quality & Operational Risk Monitoring</p>
    </div>
    """, unsafe_allow_html=True)

    with st.sidebar:
        st.header("📤 Upload Files")
        sap_file    = st.file_uploader("SAP Excel (.xlsx)", type=["xlsx"], key="sap_upload")
        infor_files = st.file_uploader("Infor CSV (multi-file ok)", type=["csv"], accept_multiple_files=True, key="infor_upload")
        st.caption("SAP needs `PO No.(Full)` & `Quantity`. Infor CSV needs `PSDD`, `CRD`, `Line Aggregator`.")

    if sap_file and infor_files:
        with status_ctx("Reading & merging files…", expanded=True) as status:
            try:
                sap_df        = read_excel_file(sap_file)
                infor_csv_dfs = [read_csv_file(f) for f in infor_files]
                infor_all     = load_infor_from_many_csv(infor_csv_dfs)

                if infor_all.empty:
                    _status_update(status, label="Failed: no valid Infor CSV.", state="error")
                else:
                    _status_update(status, label="Files read. Building report…", state="running")
                    final_df = build_report(sap_df, infor_all)

                    if final_df.empty:
                        _status_update(status, label="Failed to build report — check required columns.", state="error")
                    else:
                        _status_update(status, label="Report ready ✅", state="complete")

                        # ── Country map expander ──
                        with st.expander("🗺️ Country Normalization Map", expanded=False):
                            map_df = pd.DataFrame(list(COUNTRY_NAME_MAP.items()), columns=["SAP Original","Normalized"])
                            st.dataframe(map_df, use_container_width=True, hide_index=True)

                        # ── Sidebar filters ──
                        with st.sidebar.form("filters_form"):
                            st.header("🔎 Filters")

                            def uniq_vals(df, col):
                                if col in df.columns:
                                    return sorted([str(x) for x in df[col].dropna().unique().tolist()])
                                return []

                            status_opts     = uniq_vals(final_df, "Order Status Infor")
                            selected_status = st.multiselect("Order Status Infor", options=status_opts, default=status_opts)
                            po_opts         = uniq_vals(final_df, "PO No.(Full)")
                            selected_pos    = st.multiselect("PO No.(Full)", options=po_opts, placeholder="All PO (optional filter)")

                            result_cols_filter = [
                                "Result_Quantity","Result_FPD","Result_LPD","Result_CRD",
                                "Result_PSDD","Result_PODD","Result_PD","Result_Market PO",
                                "Result_Shipment Method","Result_Country",
                            ]
                            result_selections = {}
                            for col in result_cols_filter:
                                opts = uniq_vals(final_df, col)
                                if opts:
                                    result_selections[col] = st.multiselect(col, options=opts, default=opts)

                            mode      = st.radio("Data view mode", ["All Columns","Analyse LPD PODD","Analyse FPD PSDD"], horizontal=False)
                            submitted = st.form_submit_button("🔄 Apply Filters")

                        if submitted or "df_view" in st.session_state:
                            if submitted:
                                st.session_state["selected_status"]   = selected_status
                                st.session_state["selected_pos"]      = selected_pos
                                st.session_state["result_selections"] = result_selections
                                st.session_state["mode"]              = mode

                            selected_status   = st.session_state.get("selected_status",   status_opts)
                            selected_pos      = st.session_state.get("selected_pos",      [])
                            result_selections = st.session_state.get("result_selections", {})
                            mode              = st.session_state.get("mode",              "All Columns")

                            df_view = final_df.copy()
                            if selected_status:
                                df_view = df_view[df_view["Order Status Infor"].astype(str).isin(selected_status)]
                            if selected_pos:
                                df_view = df_view[df_view["PO No.(Full)"].astype(str).isin(selected_pos)]
                            for col, sel in result_selections.items():
                                base_opts = uniq_vals(final_df, col)
                                if sel and set(sel) != set(base_opts):
                                    df_view = df_view[df_view[col].astype(str).isin(sel)]

                            st.session_state["df_view"]  = df_view
                            st.session_state["final_df"] = final_df

                            # ════════════════════════════════════════
                            #  1️⃣  EXECUTIVE SUMMARY KPI BAR
                            # ════════════════════════════════════════
                            stats = render_executive_summary(df_view)

                            st.markdown('<div class="ct-divider"></div>', unsafe_allow_html=True)

                            # Two-column layout for heatmap + insights
                            col_left, col_right = st.columns([3, 2], gap="large")

                            with col_left:
                                # ════════════════════════════════════════
                                #  2️⃣  DATA QUALITY HEATMAP
                                # ════════════════════════════════════════
                                render_data_quality_heatmap(stats)

                            with col_right:
                                # ════════════════════════════════════════
                                #  6️⃣  SMART BUSINESS INSIGHTS
                                # ════════════════════════════════════════
                                render_smart_insights(df_view, stats)

                            st.markdown('<div class="ct-divider"></div>', unsafe_allow_html=True)

                            # ════════════════════════════════════════
                            #  3️⃣  MISMATCH ROOT CAUSE
                            # ════════════════════════════════════════
                            render_mismatch_root_cause(df_view)

                            st.markdown('<div class="ct-divider"></div>', unsafe_allow_html=True)

                            # ════════════════════════════════════════
                            #  4️⃣  OPERATIONAL RISK PANEL
                            # ════════════════════════════════════════
                            render_operational_risk(df_view)

                            st.markdown('<div class="ct-divider"></div>', unsafe_allow_html=True)

                            # ════════════════════════════════════════
                            #  5️⃣  FIELD FAILURE TREND
                            # ════════════════════════════════════════
                            render_field_failure_trend(stats)

                            st.markdown('<div class="ct-divider"></div>', unsafe_allow_html=True)

                            # ════════════════════════════════════════
                            #  7️⃣  FALSE DETAIL DRILL-DOWN (OPEN)
                            # ════════════════════════════════════════
                            render_false_insight_open(df_view)

                            st.markdown('<div class="ct-divider"></div>', unsafe_allow_html=True)

                            # ════════════════════════════════════════
                            #  RAW DATA PREVIEW
                            # ════════════════════════════════════════
                            sh_raw = """
                            <div class="section-header">
                                <div class="sh-icon" style="background:#f8fafc;">📋</div>
                                <h2>Raw Data Preview<span class="sub">filtered result set</span></h2>
                            </div>"""
                            st.markdown(sh_raw, unsafe_allow_html=True)

                            def subset(df, cols):
                                existing = [c for c in cols if c in df.columns]
                                return df[existing] if existing else pd.DataFrame()

                            if mode == "All Columns":
                                st.dataframe(df_view.head(100), use_container_width=True)
                            elif mode == "Analyse LPD PODD":
                                st.dataframe(subset(df_view, [
                                    "PO No.(Full)","Order Status Infor","DRC","Delay/Early - Confirmation PD",
                                    "Delay/Early - Confirmation CRD","Infor Delay/Early - Confirmation CRD","Result_Delay_CRD",
                                    "Delay - PO PSDD Update","Infor Delay - PO PSDD Update","Result_Delay_PSDD",
                                    "Delay - PO PD Update","LPD","Infor LPD","Result_LPD","PODD","Infor PODD","Result_PODD",
                                ]).head(2000), use_container_width=True)
                            elif mode == "Analyse FPD PSDD":
                                st.dataframe(subset(df_view, [
                                    "PO No.(Full)","Order Status Infor","DRC","Delay/Early - Confirmation PD",
                                    "Delay/Early - Confirmation CRD","Infor Delay/Early - Confirmation CRD","Result_Delay_CRD",
                                    "Delay - PO PSDD Update","Infor Delay - PO PSDD Update","Result_Delay_PSDD",
                                    "Delay - PO PD Update","FPD","Infor FPD","Result_FPD","PSDD","Infor PSDD","Result_PSDD",
                                ]).head(2000), use_container_width=True)

                            # ── Downloads ──
                            st.markdown('<div class="ct-divider"></div>', unsafe_allow_html=True)
                            st.markdown("""
                            <div class="section-header">
                                <div class="sh-icon" style="background:#f0fdf4;">⬇️</div>
                                <h2>Export Report</h2>
                            </div>""", unsafe_allow_html=True)

                            df_export     = _blank_export_columns(df_view)
                            out_name_xlsx = f"PGD Control Tower Report - {today_str_id()}.xlsx"
                            out_name_csv  = f"PGD Control Tower Report - {today_str_id()}.csv"

                            dc1, dc2 = st.columns(2)
                            dc1.download_button(
                                label="⬇️ Download CSV (Filtered)",
                                data=df_export.to_csv(index=False).encode("utf-8"),
                                file_name=out_name_csv, mime="text/csv",
                                use_container_width=True,
                            )
                            try:
                                excel_bytes = _export_excel_styled(df_export, sheet_name="Report")
                                dc2.download_button(
                                    label="⬇️ Download Excel (Styled)",
                                    data=excel_bytes, file_name=out_name_xlsx,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True,
                                )
                            except Exception as ex_excel:
                                dc2.warning(f"Excel export error: {ex_excel}")

                        else:
                            st.info("Configure filters in the sidebar and click **🔄 Apply Filters** to load the dashboard.")

            except Exception as e:
                _status_update(status, label="Application error.", state="error")
                st.exception(e)
    else:
        st.markdown("""
        <div style="background:#f8fafc;border-radius:12px;padding:40px;text-align:center;border:2px dashed #e2e8f0;margin-top:20px;">
            <div style="font-size:2.5rem;margin-bottom:12px;">📂</div>
            <div style="font-size:1.1rem;font-weight:600;color:#0f172a;margin-bottom:8px;">Upload files to start monitoring</div>
            <div style="color:#64748b;font-size:0.85rem;">Use the sidebar to upload your SAP Excel and Infor CSV files</div>
        </div>
        """, unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════
#  TAB 2 — PO SPLITTER
# ════════════════════════════════════════════════════════════
with tab2:
    st.markdown("""
    <div class="ct-banner" style="background:linear-gradient(135deg,#0f172a 0%,#1e3a2f 100%);border-left-color:#34d399;">
        <h1>PO Splitter <span class="ct-badge" style="background:#10b981;">TOOL</span></h1>
        <p>Split large PO lists into chunks of 5,000 for system import</p>
    </div>
    """, unsafe_allow_html=True)

    with st.expander("⚙️ Parsing & Normalization Options", expanded=False):
        c1, c2, c3, c4, c5 = st.columns(5)
        split_mode          = c1.selectbox("Separator mode", ["auto","newline","comma","semicolon","whitespace"])
        chunk_size          = c2.number_input("Max PO per chunk", min_value=1, max_value=1_000_000, value=5000, step=1)
        drop_duplicates     = c3.checkbox("Remove duplicates", value=False)
        keep_only_digits    = c4.checkbox("Keep only digits", value=False)
        upper_case          = c5.checkbox("Uppercase", value=False)
        strip_prefix_suffix = st.checkbox("Strip non-alphanumeric prefix/suffix", value=False)

    input_text = st.text_area(
        "Paste PO list here:",
        height=220,
        placeholder="PO001\nPO002\nPO003\n— or —\nPO001, PO002, PO003",
        key="po_splitter_text",
    )

    if st.button("🚀 Split PO List", key="po_splitter_btn"):
        items          = parse_input(input_text, split_mode=split_mode)
        original_count = len(items)
        if keep_only_digits or upper_case or strip_prefix_suffix:
            items = normalize_items(items, keep_only_digits=keep_only_digits,
                                    upper_case=upper_case, strip_prefix_suffix=strip_prefix_suffix)
        if drop_duplicates:
            items = list(dict.fromkeys(items))

        total = len(items)
        st.divider()
        c1, c2, c3 = st.columns(3)
        c1.metric("Input (before processing)",  original_count)
        c2.metric("After processing",           total)
        c3.metric("Chunk size",                 chunk_size)

        if total == 0:
            st.warning("No PO detected. Check input & parsing options.")
        else:
            parts = chunk_list(items, int(chunk_size))
            st.success(f"Split into **{len(parts)}** chunks.")
            col_zip1, col_zip2 = st.columns(2)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            col_zip1.download_button("⬇️ Download ZIP (CSV)", data=make_zip_bytes(parts, as_csv=True),
                file_name=f"PO_chunks_csv_{timestamp}.zip", mime="application/zip", use_container_width=True)
            col_zip2.download_button("⬇️ Download ZIP (TXT)", data=make_zip_bytes(parts, as_csv=False),
                file_name=f"PO_chunks_txt_{timestamp}.zip", mime="application/zip", use_container_width=True)
            for idx, part in enumerate(parts, start=1):
                with st.expander(f"Chunk {idx} — {len(part):,} PO", expanded=False):
                    df = df_from_list(part, col_name="PO")
                    st.dataframe(df, use_container_width=True, hide_index=True)
                    cdl1, cdl2 = st.columns(2)
                    cdl1.download_button(f"CSV chunk {idx}", data=df.to_csv(index=False).encode("utf-8"),
                        file_name=f"PO_chunk_{idx:02d}.csv", mime="text/csv", use_container_width=True)
                    cdl2.download_button(f"TXT chunk {idx}", data=to_txt_bytes(part),
                        file_name=f"PO_chunk_{idx:02d}.txt", mime="text/plain", use_container_width=True)
    else:
        st.caption("Ready — click **🚀 Split PO List** to begin.")

# ── Debug ──
with st.expander("🛠 Debug Info", expanded=False):
    try:
        import platform
        st.write("Python:", sys.version)
        st.write("Platform:", platform.platform())
        st.write("Streamlit:", st.__version__)
        st.write("Pandas:", pd.__version__)
        st.write("NumPy:", np.__version__)
        st.write("openpyxl available:", EXCEL_EXPORT_AVAILABLE)
    except Exception as e:
        st.write("Debug error:", e)
